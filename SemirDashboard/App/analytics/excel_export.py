"""
App/analytics/excel_export.py

Excel export functionality for analytics data.
Generates formatted Excel workbooks with multiple sheets.

Header Abbreviations:
    INV(RET) = Returning Invoices (invoices from customers who made return visits)
    AMT(RET) = Returning Amount (total amount from returning customers)
    INV(CUS) = Customer Invoices (all invoices excluding VIP ID = 0)
    AMT(CUS) = Customer Amount (total amount from customers excluding VIP ID = 0)

Version: 3.4 - Added shop details and comparison sheets
Version: 3.5 - Standardized headers to use abbreviations consistently
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Common Excel style helpers ────────────────────────────────────────────────
# Importable by other modules (coupon_analytics, etc.) for consistent styling.

# Sales / CNV workbook: navy blue headers
XL_HDR_FILL  = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
XL_HDR_FONT  = Font(bold=True, color="FFFFFF")
XL_HDR_ALIGN = Alignment(horizontal="center", vertical="center")

# Section title style (bold, no background)
XL_TITLE_FONT = Font(bold=True, size=14)

# Filter annotation style (small italic grey)
XL_FILTER_FONT = Font(italic=True, color="888888", size=9)


def xl_write_header(ws, headers, fill=None, font=None, align=None, row=1):
    """Write a styled header row to ws starting at the given row number."""
    fill  = fill  or XL_HDR_FILL
    font  = font  or XL_HDR_FONT
    align = align or XL_HDR_ALIGN
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill
        c.font = font
        c.alignment = align


def xl_set_col_widths(ws, widths, start_col=1):
    """
    Set column widths from a list of ints/floats.

    widths: list of widths; position 0 → column start_col.
    """
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + i)].width = w


def xl_prepend_filter_row(wb, parts):
    """
    Insert a filter-info row at the top of every sheet in wb.

    parts: list of strings, e.g. ["Period: 2025-01-01 → 2025-12-31", "Shop: Bala Group"]
    No-op if parts is empty.
    """
    if not parts:
        return
    line = "  |  ".join(parts)
    for ws in wb.worksheets:
        ws.insert_rows(1)
        ws["A1"] = line
        ws["A1"].font = XL_FILTER_FONT


# ─────────────────────────────────────────────────────────────────────────────


def export_analytics_to_excel(data, date_from=None, date_to=None, shop_group=None):
    """
    Export analytics data to Excel workbook.
    
    Creates 8 sheets:
    1. Overview (includes filter info)
    2. By VIP Grade
    3. By Season
    4. By Shop
    5. By Shop - Detail
    6. By Grade - All Shops
    7. By Season - All Shops
    8. Customer Details
    
    Args:
        data: Analytics data dict
        date_from: Start date filter (for display)
        date_to: End date filter (for display)
        shop_group: Shop group filter (for display)
    """
    wb = Workbook()
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    _create_overview_sheet(wb, data, header_fill, header_font, header_align, date_from, date_to, shop_group)
    _create_grade_sheet(wb, data, header_fill, header_font, header_align)
    _create_season_sheet(wb, data, header_fill, header_font, header_align)
    _create_month_sheet(wb, data, header_fill, header_font, header_align)
    _create_week_sheet(wb, data, header_fill, header_font, header_align)
    _create_shop_sheet(wb, data, header_fill, header_font, header_align)
    _create_shop_detail_sheet(wb, data, header_fill, header_font, header_align)
    _create_grade_comparison_sheet(wb, data, header_fill, header_font, header_align)
    _create_season_comparison_sheet(wb, data, header_fill, header_font, header_align)
    _create_month_comparison_sheet(wb, data, header_fill, header_font, header_align)
    _create_week_comparison_sheet(wb, data, header_fill, header_font, header_align)
    _create_details_sheet(wb, data, header_fill, header_font, header_align)
    _create_buyer_without_info_sheet(wb, data, header_fill, header_font, header_align)
    _create_reconciliation_sheet(wb, data, header_fill, header_font, header_align)

    return wb


def _create_overview_sheet(wb, data, header_fill, header_font, header_align, date_from=None, date_to=None, shop_group=None):
    """Overview sheet with filter information."""
    ws = wb.active
    ws.title = "Overview"
    
    ws['A1'] = "Customer Analytics Overview"
    ws['A1'].font = Font(bold=True, size=14)
    
    # Show active filters
    row = 2
    if date_from and date_to:
        ws[f'A{row}'] = "Period Filter:"
        ws[f'B{row}'] = f"{date_from} to {date_to}"
        ws[f'B{row}'].font = Font(bold=True, color="0066CC")
        row += 1
    
    if shop_group:
        ws[f'A{row}'] = "Shop Group Filter:"
        ws[f'B{row}'] = shop_group
        ws[f'B{row}'].font = Font(bold=True, color="0066CC")
        row += 1
    
    row += 1  # Empty row
    
    ov = data['overview']
    for label, value in [
        ("New Members (Period)", ov['new_members_in_period']),
        ("Returning (Period)", ov['returning_customers']),
        ("Active (Period)", ov['active_customers']),
        ("Return Rate (Period)", f"{ov['return_rate']}%"),
        ("INV(RET)", ov.get('returning_invoices', 0)),
        ("AMT(RET)", ov.get('returning_amount', 0)),
        ("INV(CUS)", ov['total_invoices_without_vip0']),
        ("AMT(CUS)", ov['total_amount_without_vip0']),
        ("Total Invoices", ov['total_invoices_with_vip0']),
        ("Total Amount", ov['total_amount_with_vip0']),
        (None, None),  # separator
        ("Total Customers (All Time)", ov['total_customers_in_db']),
        ("Member Active (All Time) - Has Invoice", ov['member_active_all_time']),
        ("Member Inactive (All Time) - No Invoice", ov['member_inactive_all_time']),
        ("Return Rate (All Time)", f"{ov['return_rate_all_time']}%"),
    ]:
        if label is None:
            row += 1
            continue
        ws[f'A{row}'] = label
        # Format amounts with number format instead of string
        if 'AMT' in label or 'Amount' in label:
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = '#,##0'
        else:
            ws[f'B{row}'] = value
        row += 1
    
    # Add abbreviations explanation
    row += 1
    ws[f'A{row}'] = "Column Abbreviations:"
    ws[f'A{row}'].font = Font(bold=True, color="0066CC")
    row += 1
    
    abbrev_info = [
        ("INV(RET)", "= Returning Invoices"),
        ("AMT(RET)", "= Returning Amount"),
        ("INV(CUS)", "= Customer Invoices"),
        ("AMT(CUS)", "= Customer Amount"),
    ]
    
    for abbrev, meaning in abbrev_info:
        ws[f'A{row}'] = abbrev
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = meaning
        row += 1
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25


def _create_grade_sheet(wb, data, header_fill, header_font, header_align):
    """By VIP Grade sheet."""
    ws = wb.create_sheet("By VIP Grade")

    headers = ["Grade", "Active", "Returning", "Return Rate", "Total in DB", "Return Rate (AT)", "INV(RET)", "AMT(RET)", "Total Invoices", "Total Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, g in enumerate(data['by_grade'], 2):
        ws.cell(row=row_num, column=1, value=g['grade'])
        ws.cell(row=row_num, column=2, value=g['total_customers'])
        ws.cell(row=row_num, column=3, value=g['returning_customers'])
        ws.cell(row=row_num, column=4, value=f"{g['return_rate']}%")
        ws.cell(row=row_num, column=5, value=g.get('total_in_db', 0))
        ws.cell(row=row_num, column=6, value=f"{g.get('return_rate_all_time', 0)}%")
        ws.cell(row=row_num, column=7, value=g.get('returning_invoices', 0))
        ws.cell(row=row_num, column=8, value=g.get('returning_amount', 0))
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        ws.cell(row=row_num, column=9, value=g['total_invoices'])
        ws.cell(row=row_num, column=10, value=g['total_amount'])
        ws.cell(row=row_num, column=10).number_format = '#,##0'

    ws.column_dimensions['A'].width = 12  # Grade
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 14  # Numeric
    for col in range(7, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16  # Amount


def _create_season_sheet(wb, data, header_fill, header_font, header_align):
    """By Season sheet."""
    ws = wb.create_sheet("By Season")

    headers = ["Season", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, s in enumerate(data['by_session'], 2):
        ws.cell(row=row_num, column=1, value=s['session'])
        ws.cell(row=row_num, column=2, value=s['total_customers'])
        ws.cell(row=row_num, column=3, value=s['returning_customers'])
        ws.cell(row=row_num, column=4, value=f"{s['return_rate']}%")
        ws.cell(row=row_num, column=5, value=s.get('returning_invoices', 0))
        ws.cell(row=row_num, column=6, value=s.get('returning_amount', 0))
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7, value=s['total_invoices'])
        ws.cell(row=row_num, column=8, value=s['total_amount'])
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        ws.cell(row=row_num, column=9, value=s.get('total_invoices_with_vip0', 0))
        ws.cell(row=row_num, column=10, value=s.get('total_amount_with_vip0', 0))
        ws.cell(row=row_num, column=10).number_format = '#,##0'

    ws.column_dimensions['A'].width = 15  # Season
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14  # Numeric
    for col in range(9, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16  # Amount


def _create_month_sheet(wb, data, header_fill, header_font, header_align):
    """By Month sheet."""
    ws = wb.create_sheet("By Month")

    headers = ["Month", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, m in enumerate(data.get('by_month', []), 2):
        ws.cell(row=row_num, column=1, value=m['month'])
        ws.cell(row=row_num, column=2, value=m['total_customers'])
        ws.cell(row=row_num, column=3, value=m['returning_customers'])
        ws.cell(row=row_num, column=4, value=f"{m['return_rate']}%")
        ws.cell(row=row_num, column=5, value=m.get('returning_invoices', 0))
        ws.cell(row=row_num, column=6, value=m.get('returning_amount', 0))
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7, value=m['total_invoices'])
        ws.cell(row=row_num, column=8, value=m['total_amount'])
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        ws.cell(row=row_num, column=9, value=m.get('total_invoices_with_vip0', 0))
        ws.cell(row=row_num, column=10, value=m.get('total_amount_with_vip0', 0))
        ws.cell(row=row_num, column=10).number_format = '#,##0'

    ws.column_dimensions['A'].width = 12  # Month YYYY-MM
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(9, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _create_shop_sheet(wb, data, header_fill, header_font, header_align):
    """By Shop summary."""
    ws = wb.create_sheet("By Shop")

    headers = ["Shop", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    sorted_shops = sorted(data['by_shop'], key=lambda x: x['shop_name'])

    for row_num, shop in enumerate(sorted_shops, 2):
        ws.cell(row=row_num, column=1, value=shop['shop_name'])
        ws.cell(row=row_num, column=2, value=shop['total_customers'])
        ws.cell(row=row_num, column=3, value=shop['returning_customers'])
        ws.cell(row=row_num, column=4, value=f"{shop['return_rate']}%")
        ws.cell(row=row_num, column=5, value=shop.get('returning_invoices', 0))
        ws.cell(row=row_num, column=6, value=shop.get('returning_amount', 0))
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7, value=shop['total_invoices'])
        ws.cell(row=row_num, column=8, value=shop['total_amount'])
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        ws.cell(row=row_num, column=9, value=shop.get('total_invoices_with_vip0', 0))
        ws.cell(row=row_num, column=10, value=shop.get('total_amount_with_vip0', 0))
        ws.cell(row=row_num, column=10).number_format = '#,##0'

    ws.column_dimensions['A'].width = 30  # Shop name
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14  # Numeric
    for col in range(9, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16  # Amount


def _create_shop_detail_sheet(wb, data, header_fill, header_font, header_align):
    """By Shop - Detail with grade and season breakdowns."""
    ws = wb.create_sheet("By Shop - Detail")
    
    current_row = 1
    sorted_shops = sorted(data['by_shop'], key=lambda x: x['shop_name'])
    
    for shop in sorted_shops:
        ws.cell(row=current_row, column=1, value=f"SHOP: {shop['shop_name']}")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:M{current_row}')
        current_row += 1

        # By Grade
        ws.cell(row=current_row, column=1, value="By VIP Grade").font = Font(bold=True)
        current_row += 1

        grade_headers = ["Grade", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)"]
        for col_num, header in enumerate(grade_headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for g in shop.get('by_grade', []):
            ws.cell(row=current_row, column=1, value=g['grade'])
            ws.cell(row=current_row, column=2, value=g['total_customers'])
            ws.cell(row=current_row, column=3, value=g['returning_customers'])
            ws.cell(row=current_row, column=4, value=f"{g['return_rate']}%")
            ws.cell(row=current_row, column=5, value=g.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=g.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=g['total_invoices'])
            ws.cell(row=current_row, column=8, value=g['total_amount'])
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            current_row += 1

        current_row += 1

        # By Season
        ws.cell(row=current_row, column=1, value="By Season").font = Font(bold=True)
        current_row += 1

        season_headers = ["Season", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
        for col_num, header in enumerate(season_headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for s in shop.get('by_session', []):
            if not s.get('total_invoices_with_vip0', 0):
                continue
            ws.cell(row=current_row, column=1, value=s['session'])
            ws.cell(row=current_row, column=2, value=s['total_customers'])
            ws.cell(row=current_row, column=3, value=s['returning_customers'])
            ws.cell(row=current_row, column=4, value=f"{s['return_rate']}%")
            ws.cell(row=current_row, column=5, value=s.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=s.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=s['total_invoices'])
            ws.cell(row=current_row, column=8, value=s['total_amount'])
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            ws.cell(row=current_row, column=9, value=s.get('total_invoices_with_vip0', 0))
            ws.cell(row=current_row, column=10, value=s.get('total_amount_with_vip0', 0))
            ws.cell(row=current_row, column=10).number_format = '#,##0'
            current_row += 1

        current_row += 1

        # By Month
        ws.cell(row=current_row, column=1, value="By Month").font = Font(bold=True)
        current_row += 1

        month_headers = ["Month", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
        for col_num, header in enumerate(month_headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for m in shop.get('by_month', []):
            if not m.get('total_invoices_with_vip0', 0):
                continue
            ws.cell(row=current_row, column=1, value=m['month'])
            ws.cell(row=current_row, column=2, value=m['total_customers'])
            ws.cell(row=current_row, column=3, value=m['returning_customers'])
            ws.cell(row=current_row, column=4, value=f"{m['return_rate']}%")
            ws.cell(row=current_row, column=5, value=m.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=m.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=m['total_invoices'])
            ws.cell(row=current_row, column=8, value=m['total_amount'])
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            ws.cell(row=current_row, column=9, value=m.get('total_invoices_with_vip0', 0))
            ws.cell(row=current_row, column=10, value=m.get('total_amount_with_vip0', 0))
            ws.cell(row=current_row, column=10).number_format = '#,##0'
            current_row += 1

        current_row += 1

        # By Week
        ws.cell(row=current_row, column=1, value="By Week").font = Font(bold=True)
        current_row += 1

        week_headers = ["Week", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
        for col_num, header in enumerate(week_headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for w in shop.get('by_week', []):
            if not w.get('total_invoices_with_vip0', 0):
                continue
            ws.cell(row=current_row, column=1, value=w['week_label'])
            ws.cell(row=current_row, column=2, value=w['total_customers'])
            ws.cell(row=current_row, column=3, value=w['returning_customers'])
            ws.cell(row=current_row, column=4, value=f"{w['return_rate']}%")
            ws.cell(row=current_row, column=5, value=w.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=w.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=w['total_invoices'])
            ws.cell(row=current_row, column=8, value=w['total_amount'])
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            ws.cell(row=current_row, column=9, value=w.get('total_invoices_with_vip0', 0))
            ws.cell(row=current_row, column=10, value=w.get('total_amount_with_vip0', 0))
            ws.cell(row=current_row, column=10).number_format = '#,##0'
            current_row += 1

        current_row += 2

    # Column widths
    ws.column_dimensions['A'].width = 22  # Week label is longer
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 14  # Numeric
    for col in range(8, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16  # Amount


def _create_grade_comparison_sheet(wb, data, header_fill, header_font, header_align):
    """By Grade - All Shops comparison. Period = one table, all shops included."""
    ws = wb.create_sheet("By Grade - All Shops")

    all_grades = set()
    for shop in data['by_shop']:
        for g in shop.get('by_grade', []):
            all_grades.add(g['grade'])

    GRADE_ORDER = {'No Grade': 0, 'Member': 1, 'Silver': 2, 'Gold': 3, 'Diamond': 4}
    sorted_grades = sorted(all_grades, key=lambda x: GRADE_ORDER.get(x, 99))
    sorted_shops = sorted(data['by_shop'], key=lambda x: x['shop_name'])

    shop_grade_map = {
        shop['shop_name']: {g['grade']: g for g in shop.get('by_grade', [])}
        for shop in sorted_shops
    }

    headers = ["Shop", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "Total INV", "Total Amount"]
    current_row = 1

    for grade in sorted_grades:
        ws.cell(row=current_row, column=1, value=f"GRADE: {grade}").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for shop in sorted_shops:
            g = shop_grade_map[shop['shop_name']].get(grade, {})
            ws.cell(row=current_row, column=1, value=shop['shop_name'])
            ws.cell(row=current_row, column=2, value=g.get('total_customers', 0))
            ws.cell(row=current_row, column=3, value=g.get('returning_customers', 0))
            ws.cell(row=current_row, column=4, value=f"{g.get('return_rate', 0)}%")
            ws.cell(row=current_row, column=5, value=g.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=g.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=g.get('total_invoices', 0))
            ws.cell(row=current_row, column=7).number_format = '#,##0'
            ws.cell(row=current_row, column=8, value=g.get('total_amount', 0))
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            current_row += 1

        current_row += 1  # blank row between periods

    ws.column_dimensions['A'].width = 32
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_season_comparison_sheet(wb, data, header_fill, header_font, header_align):
    """By Season - All Shops comparison. Period = one table, all shops included."""
    ws = wb.create_sheet("By Season - All Shops")

    all_seasons = [s['session'] for s in data['by_session']]
    sorted_shops = sorted(data['by_shop'], key=lambda x: x['shop_name'])

    shop_session_map = {
        shop['shop_name']: {s['session']: s for s in shop.get('by_session', [])}
        for shop in sorted_shops
    }

    headers = ["Shop", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "Total INV", "Total Amount"]
    current_row = 1

    for season in all_seasons:
        ws.cell(row=current_row, column=1, value=f"SEASON: {season}").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for shop in sorted_shops:
            s = shop_session_map[shop['shop_name']].get(season, {})
            ws.cell(row=current_row, column=1, value=shop['shop_name'])
            ws.cell(row=current_row, column=2, value=s.get('total_customers', 0))
            ws.cell(row=current_row, column=3, value=s.get('returning_customers', 0))
            ws.cell(row=current_row, column=4, value=f"{s.get('return_rate', 0)}%")
            ws.cell(row=current_row, column=5, value=s.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=s.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=s.get('total_invoices_with_vip0', 0))
            ws.cell(row=current_row, column=7).number_format = '#,##0'
            ws.cell(row=current_row, column=8, value=s.get('total_amount_with_vip0', 0))
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            current_row += 1

        current_row += 1

    ws.column_dimensions['A'].width = 32
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_month_comparison_sheet(wb, data, header_fill, header_font, header_align):
    """By Month - All Shops comparison. Period = one table, all shops included."""
    ws = wb.create_sheet("By Month - All Shops")

    all_months = [m['month'] for m in data.get('by_month', [])]
    sorted_shops = sorted(data['by_shop'], key=lambda x: x['shop_name'])

    shop_month_map = {
        shop['shop_name']: {m['month']: m for m in shop.get('by_month', [])}
        for shop in sorted_shops
    }

    headers = ["Shop", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "Total INV", "Total Amount"]
    current_row = 1

    for month in all_months:
        ws.cell(row=current_row, column=1, value=f"MONTH: {month}").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for shop in sorted_shops:
            m = shop_month_map[shop['shop_name']].get(month, {})
            ws.cell(row=current_row, column=1, value=shop['shop_name'])
            ws.cell(row=current_row, column=2, value=m.get('total_customers', 0))
            ws.cell(row=current_row, column=3, value=m.get('returning_customers', 0))
            ws.cell(row=current_row, column=4, value=f"{m.get('return_rate', 0)}%")
            ws.cell(row=current_row, column=5, value=m.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=m.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=m.get('total_invoices_with_vip0', 0))
            ws.cell(row=current_row, column=7).number_format = '#,##0'
            ws.cell(row=current_row, column=8, value=m.get('total_amount_with_vip0', 0))
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            current_row += 1

        current_row += 1

    ws.column_dimensions['A'].width = 32
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15


def _create_week_sheet(wb, data, header_fill, header_font, header_align):
    """By Week sheet."""
    ws = wb.create_sheet("By Week")

    headers = ["Week", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "INV(CUS)", "AMT(CUS)", "Total Invoices", "Total Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, w in enumerate(data.get('by_week', []), 2):
        ws.cell(row=row_num, column=1, value=w['week_label'])
        ws.cell(row=row_num, column=2, value=w['total_customers'])
        ws.cell(row=row_num, column=3, value=w['returning_customers'])
        ws.cell(row=row_num, column=4, value=f"{w['return_rate']}%")
        ws.cell(row=row_num, column=5, value=w.get('returning_invoices', 0))
        ws.cell(row=row_num, column=6, value=w.get('returning_amount', 0))
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        ws.cell(row=row_num, column=7, value=w['total_invoices'])
        ws.cell(row=row_num, column=8, value=w['total_amount'])
        ws.cell(row=row_num, column=8).number_format = '#,##0'
        ws.cell(row=row_num, column=9, value=w.get('total_invoices_with_vip0', 0))
        ws.cell(row=row_num, column=10, value=w.get('total_amount_with_vip0', 0))
        ws.cell(row=row_num, column=10).number_format = '#,##0'

    ws.column_dimensions['A'].width = 22  # Week label e.g. "Week 1 (1/1-7/1)"
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 14
    for col in range(9, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _create_week_comparison_sheet(wb, data, header_fill, header_font, header_align):
    """By Week - All Shops comparison. Period = one table, all shops included."""
    ws = wb.create_sheet("By Week - All Shops")

    all_weeks = [(w['week_sort'], w['week_label']) for w in data.get('by_week', [])]
    sorted_shops = sorted(data['by_shop'], key=lambda x: x['shop_name'])

    shop_week_map = {
        shop['shop_name']: {w['week_sort']: w for w in shop.get('by_week', [])}
        for shop in sorted_shops
    }

    headers = ["Shop", "Active", "Returning", "Return Rate", "INV(RET)", "AMT(RET)", "Total INV", "Total Amount"]
    current_row = 1

    for week_sort, week_label in all_weeks:
        ws.cell(row=current_row, column=1, value=f"WEEK: {week_label}").font = Font(bold=True, size=12)
        ws.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
        current_row += 1

        for shop in sorted_shops:
            w = shop_week_map[shop['shop_name']].get(week_sort, {})
            ws.cell(row=current_row, column=1, value=shop['shop_name'])
            ws.cell(row=current_row, column=2, value=w.get('total_customers', 0))
            ws.cell(row=current_row, column=3, value=w.get('returning_customers', 0))
            ws.cell(row=current_row, column=4, value=f"{w.get('return_rate', 0)}%")
            ws.cell(row=current_row, column=5, value=w.get('returning_invoices', 0))
            ws.cell(row=current_row, column=6, value=w.get('returning_amount', 0))
            ws.cell(row=current_row, column=6).number_format = '#,##0'
            ws.cell(row=current_row, column=7, value=w.get('total_invoices_with_vip0', 0))
            ws.cell(row=current_row, column=7).number_format = '#,##0'
            ws.cell(row=current_row, column=8, value=w.get('total_amount_with_vip0', 0))
            ws.cell(row=current_row, column=8).number_format = '#,##0'
            current_row += 1

        current_row += 1

    ws.column_dimensions['A'].width = 32
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 15
    for col in range(9, 15):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _create_details_sheet(wb, data, header_fill, header_font, header_align):
    """Customer Details sheet."""
    ws = wb.create_sheet("Customer Details")
    
    headers = ["VIP ID", "Name", "Grade", "Reg Date", "First Purchase", "Purchases", "Return Visits", "Total Spent"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, c in enumerate(data['customer_details'], 2):
        ws.cell(row=row_num, column=1, value=c['vip_id'])
        ws.cell(row=row_num, column=2, value=c['name'])
        ws.cell(row=row_num, column=3, value=c.get('vip_grade', ''))
        ws.cell(row=row_num, column=4, value=str(c.get('registration_date', '')) if c.get('registration_date') else '')
        ws.cell(row=row_num, column=5, value=str(c['first_purchase_date']))
        ws.cell(row=row_num, column=6, value=c['total_purchases'])
        ws.cell(row=row_num, column=7, value=c['return_visits'])
        ws.cell(row=row_num, column=8, value=c['total_spent'])
        ws.cell(row=row_num, column=8).number_format = '#,##0'
    
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18

def _create_buyer_without_info_sheet(wb, data, header_fill, header_font, header_align):
    """Buyer Without Info (VIP ID = 0) sheet."""
    bwi = data.get('buyer_without_info_stats')
    if not bwi:
        return

    ws = wb.create_sheet("Buyer Without Info")
    row = 1

    # Period summary
    ws.cell(row=row, column=1, value="Period Summary").font = Font(bold=True)
    row += 1
    period = bwi.get('period', {})
    for label, value in [
        ("Total Invoices (Period)", period.get('total_invoices', 0)),
        ("Total Amount (Period)", period.get('total_amount', 0)),
        ("% of All Invoices", f"{period.get('pct_of_all_invoices', 0)}%"),
        ("% of All Amount", f"{period.get('pct_of_all_amount', 0)}%"),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if 'Amount' in label:
            ws.cell(row=row, column=2).number_format = '#,##0'
        row += 1

    row += 1  # spacer

    # All-time summary
    ws.cell(row=row, column=1, value="All-Time Summary").font = Font(bold=True)
    row += 1
    alltime = bwi.get('all_time', {})
    for label, value in [
        ("Total Invoices (All Time)", alltime.get('total_invoices', 0)),
        ("Total Amount (All Time)", alltime.get('total_amount', 0)),
    ]:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        if 'Amount' in label:
            ws.cell(row=row, column=2).number_format = '#,##0'
        row += 1

    row += 1  # spacer

    # By shop breakdown
    ws.cell(row=row, column=1, value="By Shop (Period)").font = Font(bold=True)
    row += 1
    shop_headers = ["Shop", "Invoices", "Amount", "% Inv (All Period)", "% Amt (All Period)"]
    for col_num, header in enumerate(shop_headers, 1):
        cell = ws.cell(row=row, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    row += 1
    for s in bwi.get('by_shop', []):
        ws.cell(row=row, column=1, value=s['shop_name'])
        ws.cell(row=row, column=2, value=s['invoices'])
        ws.cell(row=row, column=3, value=s['amount'])
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4, value=f"{s['pct_of_period_invoices']}%")
        ws.cell(row=row, column=5, value=f"{s['pct_of_period_amount']}%")
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18


"""
ULTIMATE FIX - Uses customer_utils.get_customer_info() for proper lookup
"""

def _create_reconciliation_sheet(wb, data, header_fill, header_font, header_align):
    """
    ULTIMATE FIX: Use customer_utils.get_customer_info() instead of manual lookup
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from collections import defaultdict
    from App.analytics.customer_utils import get_customer_info
    
    customer_purchases = data.get('customer_purchases', {})
    
    if not customer_purchases:
        ws = wb.create_sheet("⚠️ Reconciliation")
        ws['A1'] = "⚠️ Customer purchase data not available"
        ws['A1'].font = Font(bold=True, size=14, color="FF0000")
        return
    
    # ============================================================================
    # Helper Functions
    # ============================================================================
    
    def to_date(d):
        """Convert datetime to date for comparison"""
        if d is None:
            return None
        return d.date() if hasattr(d, 'date') else d
    
    def calculate_return_visits_local(purchases_sorted, reg_date):
        """Calculate return visits with proper date comparison"""
        n = len(purchases_sorted)
        if n == 0:
            return (0, False)
        
        # Convert to dates for comparison
        first_date = to_date(purchases_sorted[0]['date'])
        reg_date_cmp = to_date(reg_date)
        
        # Apply formula
        if reg_date_cmp and first_date == reg_date_cmp:
            return_visits = n - 1
        else:
            return_visits = n
        
        is_returning = (return_visits > 0)
        return (return_visits, is_returning)
    
    # ============================================================================
    # Get Summary Data
    # ============================================================================
    
    ov = data.get('overview', {})
    overview_total = ov.get('returning_invoices', 0)
    shop_sum = sum(s.get('returning_invoices', 0) for s in data.get('by_shop', []))
    season_sum = sum(s.get('returning_invoices', 0) for s in data.get('by_session', []))
    
    shop_diff = overview_total - shop_sum
    season_diff = overview_total - season_sum
    
    # ============================================================================
    # SHEET 1: SHOP RECONCILIATION
    # ============================================================================
    
    ws_shop = wb.create_sheet("Reconciliation - Shops")
    
    # Title
    ws_shop['A1'] = "SHOP RECONCILIATION"
    ws_shop['A1'].font = Font(bold=True, size=14, color="FF0000")
    ws_shop.merge_cells('A1:I1')
    
    row = 3
    
    # Summary
    ws_shop[f'A{row}'] = "Global Returning Invoices:"
    ws_shop[f'B{row}'] = overview_total
    ws_shop[f'B{row}'].font = Font(bold=True, color="0066CC")
    row += 1
    
    ws_shop[f'A{row}'] = "Sum of Shop Returning Invoices:"
    ws_shop[f'B{row}'] = shop_sum
    row += 1
    
    ws_shop[f'A{row}'] = "Difference:"
    ws_shop[f'B{row}'] = shop_diff
    ws_shop[f'B{row}'].font = Font(bold=True, size=12, color="FF0000")
    row += 2
    
    # Single pass: compute both shop_problems and season_problems simultaneously.
    # purchases lists are already sorted by build_customer_purchase_map — no re-sort needed.
    shop_problems = []
    season_problems = []

    for vip_id, purchases in customer_purchases.items():
        if vip_id == '0' or not purchases:
            continue

        # purchases already sorted; use directly
        customer_obj = purchases[0].get('customer')
        grade, reg_date, name = get_customer_info(vip_id, customer_obj)

        global_rv, global_is_ret = calculate_return_visits_local(purchases, reg_date)
        if not global_is_ret:
            continue

        global_ret_inv = len(purchases)

        # Build per-shop and per-season sub-lists in one scan (sub-lists inherit sort order)
        by_shop = defaultdict(list)
        by_season = defaultdict(list)
        for p in purchases:
            by_shop[p.get('shop', 'Unknown')].append(p)
            by_season[p.get('session', 'Unknown')].append(p)

        # ── Shop diff ──────────────────────────────────────────────────────────
        shop_total = 0
        for shop_purch in by_shop.values():
            _, is_ret = calculate_return_visits_local(shop_purch, reg_date)
            if is_ret:
                shop_total += len(shop_purch)

        cust_shop_diff = global_ret_inv - shop_total
        if cust_shop_diff > 0:
            reg_date_cmp = to_date(reg_date)
            if reg_date_cmp:
                reg_day_purch = [p for p in purchases if to_date(p['date']) == reg_date_cmp]
                reg_day_shops = {p.get('shop', 'Unknown') for p in reg_day_purch}
            else:
                reg_day_purch = []
                reg_day_shops = set()

            if len(reg_day_shops) >= 2:
                pattern = "Multi-shop reg day"
            elif len(reg_day_purch) >= 1 and len(by_shop) >= 2:
                pattern = "Reg day → other shops"
            else:
                pattern = "Other"

            # sub-lists already in date order; first element is earliest
            shop_details = []
            by_shop_sorted = sorted(by_shop.items(), key=lambda kv: kv[1][0]['date'])
            for sh, sh_purch in by_shop_sorted[:3]:
                _, sh_is_ret = calculate_return_visits_local(sh_purch, reg_date)
                shop_details.append(f"{sh[:25]}({len(sh_purch)},ret={sh_is_ret})")

            shop_problems.append({
                'vip_id': vip_id, 'name': name, 'reg_date': reg_date,
                'total_purchases': global_ret_inv, 'global_ret_inv': global_ret_inv,
                'shop_total': shop_total, 'difference': cust_shop_diff,
                'pattern': pattern, 'details': "; ".join(shop_details),
            })

        # ── Season diff ────────────────────────────────────────────────────────
        season_total = 0
        for season_purch in by_season.values():
            _, is_ret = calculate_return_visits_local(season_purch, reg_date)
            if is_ret:
                season_total += len(season_purch)

        cust_season_diff = global_ret_inv - season_total
        if cust_season_diff > 0:
            season_details = []
            by_season_sorted = sorted(by_season.items(), key=lambda kv: kv[1][0]['date'])
            for ssn, ssn_purch in by_season_sorted[:3]:
                _, ssn_is_ret = calculate_return_visits_local(ssn_purch, reg_date)
                season_details.append(f"{ssn}({len(ssn_purch)},ret={ssn_is_ret})")

            season_problems.append({
                'vip_id': vip_id, 'name': name, 'reg_date': reg_date,
                'total_purchases': global_ret_inv, 'global_ret_inv': global_ret_inv,
                'season_total': season_total, 'difference': cust_season_diff,
                'num_seasons': len(by_season), 'details': "; ".join(season_details),
            })

    shop_problems.sort(key=lambda x: x['difference'], reverse=True)
    
    # Pattern summary
    pattern1 = [c for c in shop_problems if c['pattern'] == "Multi-shop reg day"]
    pattern2 = [c for c in shop_problems if c['pattern'] == "Reg day → other shops"]
    
    ws_shop[f'A{row}'] = "PATTERNS:"
    ws_shop[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_shop[f'A{row}'] = f"Pattern 1 (Multi-shop reg day): {len(pattern1)} customers, {sum(c['difference'] for c in pattern1)} invoices"
    row += 1
    
    ws_shop[f'A{row}'] = f"Pattern 2 (Reg day → other shops): {len(pattern2)} customers, {sum(c['difference'] for c in pattern2)} invoices"
    row += 2
    
    # Table
    ws_shop[f'A{row}'] = f"ALL {len(shop_problems)} CUSTOMERS WITH SHOP DIFFERENCES:"
    ws_shop[f'A{row}'].font = Font(bold=True, size=11)
    ws_shop.merge_cells(f'A{row}:I{row}')
    row += 1
    
    # Headers
    headers = ["VIP ID", "Name", "Reg Date", "Total Purch", "Global", "Shop Sum", "Diff", "Pattern", "Details"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_shop.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row += 1
    
    # Data rows
    for c in shop_problems:
        ws_shop.cell(row=row, column=1, value=c['vip_id'])
        ws_shop.cell(row=row, column=2, value=c['name'])
        ws_shop.cell(row=row, column=3, value=c['reg_date'].strftime('%Y-%m-%d') if c['reg_date'] else 'NULL')
        ws_shop.cell(row=row, column=4, value=c['total_purchases'])
        ws_shop.cell(row=row, column=5, value=c['global_ret_inv'])
        ws_shop.cell(row=row, column=6, value=c['shop_total'])
        
        diff_cell = ws_shop.cell(row=row, column=7, value=c['difference'])
        diff_cell.font = Font(bold=True, color="FF0000")
        
        pattern_cell = ws_shop.cell(row=row, column=8, value=c['pattern'])
        if c['pattern'] == "Multi-shop reg day":
            pattern_cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        else:
            pattern_cell.fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
        
        ws_shop.cell(row=row, column=9, value=c['details'])
        row += 1
    
    # Column widths
    ws_shop.column_dimensions['A'].width = 12
    ws_shop.column_dimensions['B'].width = 25
    ws_shop.column_dimensions['C'].width = 12
    ws_shop.column_dimensions['D'].width = 12
    ws_shop.column_dimensions['E'].width = 10
    ws_shop.column_dimensions['F'].width = 10
    ws_shop.column_dimensions['G'].width = 8
    ws_shop.column_dimensions['H'].width = 22
    ws_shop.column_dimensions['I'].width = 50
    
    # ============================================================================
    # SHEET 2: SEASON RECONCILIATION
    # ============================================================================
    
    ws_season = wb.create_sheet("Reconciliation - Seasons")
    
    # Title
    ws_season['A1'] = "SEASON RECONCILIATION"
    ws_season['A1'].font = Font(bold=True, size=14, color="FF0000")
    ws_season.merge_cells('A1:I1')
    
    row = 3
    
    # Summary
    ws_season[f'A{row}'] = "Global Returning Invoices:"
    ws_season[f'B{row}'] = overview_total
    ws_season[f'B{row}'].font = Font(bold=True, color="0066CC")
    row += 1
    
    ws_season[f'A{row}'] = "Sum of Season Returning Invoices:"
    ws_season[f'B{row}'] = season_sum
    row += 1
    
    ws_season[f'A{row}'] = "Difference:"
    ws_season[f'B{row}'] = season_diff
    ws_season[f'B{row}'].font = Font(bold=True, size=12, color="FF0000")
    row += 2
    
    season_problems.sort(key=lambda x: x['difference'], reverse=True)
    
    # Pattern explanation
    ws_season[f'A{row}'] = f"Pattern: {len(season_problems)} customers, {sum(c['difference'] for c in season_problems)} invoices"
    ws_season[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws_season[f'A{row}'] = "First purchase on reg day in Season 1 → NOT returning in Season 1"
    row += 1
    ws_season[f'A{row}'] = "Then purchases in other seasons → IS returning in those seasons"
    row += 1
    ws_season[f'A{row}'] = "Each customer loses exactly 1 invoice from first season"
    row += 2
    
    # Table
    ws_season[f'A{row}'] = f"ALL {len(season_problems)} CUSTOMERS WITH SEASON DIFFERENCES:"
    ws_season[f'A{row}'].font = Font(bold=True, size=11)
    ws_season.merge_cells(f'A{row}:I{row}')
    row += 1
    
    # Headers
    headers = ["VIP ID", "Name", "Reg Date", "Total Purch", "Global", "Season Sum", "Diff", "Num Seasons", "Details"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_season.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row += 1
    
    # Data rows (ALL season problems)
    for c in season_problems:
        ws_season.cell(row=row, column=1, value=c['vip_id'])
        ws_season.cell(row=row, column=2, value=c['name'])
        ws_season.cell(row=row, column=3, value=c['reg_date'].strftime('%Y-%m-%d') if c['reg_date'] else 'NULL')
        ws_season.cell(row=row, column=4, value=c['total_purchases'])
        ws_season.cell(row=row, column=5, value=c['global_ret_inv'])
        ws_season.cell(row=row, column=6, value=c['season_total'])
        
        diff_cell = ws_season.cell(row=row, column=7, value=c['difference'])
        diff_cell.font = Font(bold=True, color="FF0000")
        
        ws_season.cell(row=row, column=8, value=c['num_seasons'])
        ws_season.cell(row=row, column=9, value=c['details'])
        row += 1
    
    # Column widths
    ws_season.column_dimensions['A'].width = 12
    ws_season.column_dimensions['B'].width = 25
    ws_season.column_dimensions['C'].width = 12
    ws_season.column_dimensions['D'].width = 12
    ws_season.column_dimensions['E'].width = 10
    ws_season.column_dimensions['F'].width = 10
    ws_season.column_dimensions['G'].width = 8
    ws_season.column_dimensions['H'].width = 12
    ws_season.column_dimensions['I'].width = 50
    


# Add this function to the END of excel_export.py

def export_customer_analytics_to_excel(
    pos_customers,
    cnv_customers,
    date_from=None,
    date_to=None,
    points_mismatch=None,
    total_points_mismatch=None,
    cnv_used_points=None,
    zalo_mini_app_list=None,
    zalo_mini_app_inactive_list=None,
    zalo_oa_list=None,
    zalo_stats=None,
):
    """
    Export Customer Analytics (POS vs CNV comparison) to Excel.
    
    Creates up to 5 sheets:
    1. Overview - Summary metrics and filter info
    2. POS Only - All Time - Customers in POS but not CNV
    3. CNV Only - All Time - Customers in CNV but not POS
    4. POS Only - Period - New POS customers not in CNV (if date filtered)
    5. CNV Only - Period - New CNV customers not in POS (if date filtered)
    
    Args:
        pos_customers: QuerySet of Customer objects from POS system
        cnv_customers: QuerySet of CNVCustomer objects from CNV
        date_from: Start date for period filter (optional)
        date_to: End date for period filter (optional)
    
    Returns:
        openpyxl Workbook object
    """
    wb = Workbook()
    
    # Styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    from django.db.models import Subquery

    # Base querysets with phone
    pos_base = pos_customers.filter(phone__isnull=False).exclude(phone='')
    cnv_base = cnv_customers.filter(phone__isnull=False).exclude(phone='')

    # Count sets (done via DB, no Python set needed for this)
    pos_phones_sq = Subquery(cnv_base.values('phone'))
    cnv_phones_sq = Subquery(pos_base.values('phone'))

    pos_only_count = pos_base.exclude(phone__in=pos_phones_sq).count()
    cnv_only_count = cnv_base.exclude(phone__in=cnv_phones_sq).count()
    total_pos = pos_base.count()
    total_cnv = cnv_base.count()

    # Period comparisons (if filtered)
    new_pos_count = 0
    new_cnv_count = 0
    pos_only_period_count = 0
    cnv_only_period_count = 0

    if date_from and date_to:
        new_pos = pos_base.filter(registration_date__gte=date_from, registration_date__lte=date_to)
        new_pos_count = new_pos.count()
        pos_only_period_count = new_pos.exclude(phone__in=Subquery(cnv_base.values('phone'))).count()

        new_cnv = cnv_base.filter(cnv_created_at__gte=date_from, cnv_created_at__lte=date_to)
        new_cnv_count = new_cnv.count()
        cnv_only_period_count = new_cnv.exclude(phone__in=Subquery(pos_base.values('phone'))).count()
    
    # ========================================================================
    # OVERVIEW SHEET
    # ========================================================================
    ws = wb.active
    ws.title = "Overview"
    
    ws['A1'] = "Customer Analytics - POS vs CNV Comparison"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    row = 3
    
    # Filter info
    if date_from and date_to:
        ws[f'A{row}'] = "Period Filter:"
        ws[f'B{row}'] = f"{date_from} to {date_to}"
        ws[f'B{row}'].font = Font(bold=True, color="0066CC")
        row += 1
    
    row += 1
    
    # All-Time Metrics
    ws[f'A{row}'] = "All-Time Metrics"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    metrics_all = [
        ("Total Customers (POS System)", total_pos),
        ("Total CNV Customers", total_cnv),
        ("POS Only (Not in CNV)", pos_only_count),
        ("CNV Only (Not in POS)", cnv_only_count),
    ]
    
    for label, value in metrics_all:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Period Metrics (if filtered)
    if date_from and date_to:
        row += 1
        ws[f'A{row}'] = "Period Metrics"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        metrics_period = [
            ("New Customers (POS)", new_pos_count),
            ("New CNV Customers", new_cnv_count),
            ("New POS Only (Period)", pos_only_period_count),
            ("New CNV Only (Period)", cnv_only_period_count),
        ]
        
        for label, value in metrics_period:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
    
    # Zalo All-Time Metrics
    if zalo_stats:
        row += 1
        ws[f'A{row}'] = "Zalo Metrics (All-Time)"
        ws[f'A{row}'].font = Font(bold=True, color="0068FF")
        row += 1
        zalo_metrics = [
            ("Active Zalo Mini App", zalo_stats.get('zalo_app_all_count', 0)),
            ("% Active Zalo / CNV", f"{zalo_stats.get('zalo_app_all_pct', 0)}%"),
            ("Follow Zalo OA", zalo_stats.get('zalo_oa_all_count', 0)),
            ("% Follow OA / CNV", f"{zalo_stats.get('zalo_oa_all_pct', 0)}%"),
        ]
        for label, value in zalo_metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        if date_from and date_to:
            row += 1
            ws[f'A{row}'] = "Zalo Metrics (Period)"
            ws[f'A{row}'].font = Font(bold=True, color="0068FF")
            row += 1
            zalo_period_metrics = [
                ("Active Zalo Mini App (Period)", zalo_stats.get('zalo_app_period_count', 0)),
                ("% Zalo App / CNV (Period)", f"{zalo_stats.get('zalo_app_period_pct', 0)}%"),
                ("Follow Zalo OA (Period)", zalo_stats.get('zalo_oa_period_count', 0)),
                ("% Follow OA / CNV (Period)", f"{zalo_stats.get('zalo_oa_period_pct', 0)}%"),
            ]
            for label, value in zalo_period_metrics:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20

    # ========================================================================
    # POS ONLY - ALL TIME SHEET
    # ========================================================================
    ws_pos_all = wb.create_sheet("POS Only - All Time")
    
    # Header
    headers = ['VIP ID', 'Phone', 'Name', 'VIP Grade', 'Email', 'Registration Date', 'Points']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_pos_all.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Data — subquery, no large __in set
    pos_only_customers = pos_base.exclude(
        phone__in=Subquery(cnv_base.values('phone'))
    ).order_by('registration_date')
    row = 2
    for cust in pos_only_customers:
        ws_pos_all.cell(row, 1, cust.vip_id)
        ws_pos_all.cell(row, 2, cust.phone)
        ws_pos_all.cell(row, 3, cust.name)
        ws_pos_all.cell(row, 4, cust.vip_grade or '-')
        ws_pos_all.cell(row, 5, cust.email or '-')
        ws_pos_all.cell(row, 6, str(cust.registration_date) if cust.registration_date else '-')
        ws_pos_all.cell(row, 7, cust.points or 0)
        row += 1
    
    # Column widths
    ws_pos_all.column_dimensions['A'].width = 12
    ws_pos_all.column_dimensions['B'].width = 15
    ws_pos_all.column_dimensions['C'].width = 25
    ws_pos_all.column_dimensions['D'].width = 12
    ws_pos_all.column_dimensions['E'].width = 30
    ws_pos_all.column_dimensions['F'].width = 18
    ws_pos_all.column_dimensions['G'].width = 10
    
    # ========================================================================
    # CNV ONLY - ALL TIME SHEET
    # ========================================================================
    ws_cnv_all = wb.create_sheet("CNV Only - All Time")
    
    # Header (8 columns)
    headers = ['CNV ID', 'Phone', 'Name', 'Level', 'Email', 'Registration Date', 'Points', 'Used Points']
    for col_idx, header in enumerate(headers, 1):
        cell = ws_cnv_all.cell(1, col_idx, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    # Data — subquery
    cnv_only_customers = cnv_base.exclude(
        phone__in=Subquery(pos_base.values('phone'))
    ).order_by('cnv_created_at')
    row = 2
    for cust in cnv_only_customers:
        ws_cnv_all.cell(row, 1, cust.cnv_id)
        ws_cnv_all.cell(row, 2, cust.phone)
        ws_cnv_all.cell(row, 3, cust.full_name or '-')
        ws_cnv_all.cell(row, 4, cust.level_name or '-')
        ws_cnv_all.cell(row, 5, cust.email or '-')
        ws_cnv_all.cell(row, 6, str(cust.cnv_created_at.date()) if cust.cnv_created_at else '-')
        ws_cnv_all.cell(row, 7, float(cust.total_points) if cust.total_points else 0)
        ws_cnv_all.cell(row, 8, float(cust.used_points) if cust.used_points else 0)
        row += 1
    
    # Column widths
    ws_cnv_all.column_dimensions['A'].width = 15
    ws_cnv_all.column_dimensions['B'].width = 15
    ws_cnv_all.column_dimensions['C'].width = 25
    ws_cnv_all.column_dimensions['D'].width = 12
    ws_cnv_all.column_dimensions['E'].width = 30
    ws_cnv_all.column_dimensions['F'].width = 18
    ws_cnv_all.column_dimensions['G'].width = 12
    ws_cnv_all.column_dimensions['H'].width = 12
    
    # ========================================================================
    # PERIOD SHEETS (if filtered)
    # ========================================================================
    if date_from and date_to:
        # POS Only - Period
        pos_period_qs = pos_base.filter(
            registration_date__gte=date_from,
            registration_date__lte=date_to
        ).exclude(phone__in=Subquery(cnv_base.values('phone'))).order_by('registration_date')

        if pos_period_qs.exists():
            ws_pos_period = wb.create_sheet("POS Only - Period")
            headers = ['VIP ID', 'Phone', 'Name', 'VIP Grade', 'Email', 'Registration Date', 'Points']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_pos_period.cell(1, col_idx, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            row = 2
            for cust in pos_period_qs:
                ws_pos_period.cell(row, 1, cust.vip_id)
                ws_pos_period.cell(row, 2, cust.phone)
                ws_pos_period.cell(row, 3, cust.name)
                ws_pos_period.cell(row, 4, cust.vip_grade or '-')
                ws_pos_period.cell(row, 5, cust.email or '-')
                ws_pos_period.cell(row, 6, str(cust.registration_date) if cust.registration_date else '-')
                ws_pos_period.cell(row, 7, cust.points or 0)
                row += 1
            for col, w in zip('ABCDEFG', [12, 15, 25, 12, 30, 18, 10]):
                ws_pos_period.column_dimensions[col].width = w

        # CNV Only - Period
        cnv_period_qs = cnv_base.filter(
            cnv_created_at__gte=date_from,
            cnv_created_at__lte=date_to
        ).exclude(phone__in=Subquery(pos_base.values('phone'))).order_by('cnv_created_at')

        if cnv_period_qs.exists():
            ws_cnv_period = wb.create_sheet("CNV Only - Period")
            headers = ['CNV ID', 'Phone', 'Name', 'Level', 'Email', 'Registration Date', 'Points', 'Used Points']
            for col_idx, header in enumerate(headers, 1):
                cell = ws_cnv_period.cell(1, col_idx, header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
            row = 2
            for cust in cnv_period_qs:
                ws_cnv_period.cell(row, 1, cust.cnv_id)
                ws_cnv_period.cell(row, 2, cust.phone)
                ws_cnv_period.cell(row, 3, cust.full_name or '-')
                ws_cnv_period.cell(row, 4, cust.level_name or '-')
                ws_cnv_period.cell(row, 5, cust.email or '-')
                ws_cnv_period.cell(row, 6, str(cust.cnv_created_at.date()) if cust.cnv_created_at else '-')
                ws_cnv_period.cell(row, 7, float(cust.total_points) if cust.total_points else 0)
                ws_cnv_period.cell(row, 8, float(cust.used_points) if cust.used_points else 0)
                row += 1
            for col, w in zip('ABCDEFGH', [15, 15, 25, 12, 30, 18, 12, 12]):
                ws_cnv_period.column_dimensions[col].width = w

    # ========================================================================
    # POINTS MISMATCH SHEET
    # ========================================================================
    ws_mismatch = wb.create_sheet("Points Mismatch")

    mismatch_headers = [
        'Phone', 'POS VIP ID', 'POS Name', 'POS Grade', 'POS Points', 'POS Used Points',
        'CNV ID', 'CNV Name', 'CNV Level', 'CNV Points', 'CNV Total Points', 'CNV Used Points',
        'Diff Value', 'Diff Note'
    ]
    orange_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    for col_idx, header in enumerate(mismatch_headers, 1):
        cell = ws_mismatch.cell(1, col_idx, header)
        cell.fill = orange_fill
        cell.font = header_font
        cell.alignment = header_align

    row = 2
    for m in (points_mismatch or []):
        ws_mismatch.cell(row, 1, m.get('phone', ''))
        ws_mismatch.cell(row, 2, m.get('pos_vip_id', ''))
        ws_mismatch.cell(row, 3, m.get('pos_name', ''))
        ws_mismatch.cell(row, 4, m.get('pos_grade', ''))
        ws_mismatch.cell(row, 5, m.get('pos_points', 0))
        ws_mismatch.cell(row, 6, m.get('pos_used_points', 0))
        ws_mismatch.cell(row, 7, m.get('cnv_id', ''))
        ws_mismatch.cell(row, 8, m.get('cnv_name', ''))
        ws_mismatch.cell(row, 9, m.get('cnv_level', ''))
        ws_mismatch.cell(row, 10, m.get('cnv_points', 0))
        ws_mismatch.cell(row, 11, m.get('cnv_total_points', 0))
        ws_mismatch.cell(row, 12, m.get('cnv_used_points', 0))
        diff = m.get('diff', 0)
        diff_cell = ws_mismatch.cell(row, 13, diff)
        diff_cell.font = Font(bold=True, color="16A34A" if diff > 0 else "DC2626")
        note = "Run camp to reduce point in CNV" if diff > 0 else "Run camp to increase point in CNV"
        ws_mismatch.cell(row, 14, note)
        row += 1

    for col_letter, width in zip('ABCDEFGHIJKLMN', [15, 12, 25, 12, 10, 12, 12, 25, 12, 10, 14, 12, 10, 35]):
        ws_mismatch.column_dimensions[col_letter].width = width

    # ========================================================================
    # CNV USED POINTS > 0 SHEET
    # ========================================================================
    ws_used = wb.create_sheet("CNV Used Points")

    green_fill = PatternFill(start_color="166534", end_color="166534", fill_type="solid")
    used_headers = ['CNV ID', 'Phone', 'Name', 'Level', 'Email', 'Registration Date', 'Points', 'Used Points', 'Total Points']
    for col_idx, header in enumerate(used_headers, 1):
        cell = ws_used.cell(1, col_idx, header)
        cell.fill = green_fill
        cell.font = header_font
        cell.alignment = header_align

    row = 2
    for cust in (cnv_used_points or []):
        ws_used.cell(row, 1, cust.cnv_id)
        ws_used.cell(row, 2, cust.phone)
        ws_used.cell(row, 3, cust.full_name or '-')
        ws_used.cell(row, 4, cust.level_name or '-')
        ws_used.cell(row, 5, cust.email or '-')
        ws_used.cell(row, 6, str(cust.cnv_created_at.date()) if cust.cnv_created_at else '-')
        ws_used.cell(row, 7, float(cust.points or 0))
        used_cell = ws_used.cell(row, 8, float(cust.used_points or 0))
        used_cell.font = Font(bold=True, color="166534")
        ws_used.cell(row, 9, float(cust.total_points or 0))
        row += 1

    ws_used.column_dimensions['A'].width = 15
    ws_used.column_dimensions['B'].width = 15
    ws_used.column_dimensions['C'].width = 25
    ws_used.column_dimensions['D'].width = 12
    ws_used.column_dimensions['E'].width = 30
    ws_used.column_dimensions['F'].width = 18
    ws_used.column_dimensions['G'].width = 12
    ws_used.column_dimensions['H'].width = 14
    ws_used.column_dimensions['I'].width = 14

    # ========================================================================
    # TOTAL POINTS MISMATCH SHEET  (POS.net_points vs CNV.total_points)
    # ========================================================================
    ws_total_mismatch = wb.create_sheet("Total Points Mismatch")

    total_mismatch_headers = [
        'Phone', 'POS VIP ID', 'POS Name', 'POS Grade', 'POS Points', 'POS Used Points',
        'CNV ID', 'CNV Name', 'CNV Level', 'CNV Points', 'CNV Total Points', 'CNV Used Points',
        'Diff Value', 'Diff Note'
    ]
    purple_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    for col_idx, header in enumerate(total_mismatch_headers, 1):
        cell = ws_total_mismatch.cell(1, col_idx, header)
        cell.fill = purple_fill
        cell.font = header_font
        cell.alignment = header_align

    row = 2
    for m in (total_points_mismatch or []):
        ws_total_mismatch.cell(row, 1, m.get('phone', ''))
        ws_total_mismatch.cell(row, 2, m.get('pos_vip_id', ''))
        ws_total_mismatch.cell(row, 3, m.get('pos_name', ''))
        ws_total_mismatch.cell(row, 4, m.get('pos_grade', ''))
        ws_total_mismatch.cell(row, 5, m.get('pos_points', 0))
        ws_total_mismatch.cell(row, 6, m.get('pos_used_points', 0))
        ws_total_mismatch.cell(row, 7, m.get('cnv_id', ''))
        ws_total_mismatch.cell(row, 8, m.get('cnv_name', ''))
        ws_total_mismatch.cell(row, 9, m.get('cnv_level', ''))
        ws_total_mismatch.cell(row, 10, m.get('cnv_points', 0))
        ws_total_mismatch.cell(row, 11, m.get('cnv_total_points', 0))
        ws_total_mismatch.cell(row, 12, m.get('cnv_used_points', 0))
        diff = m.get('diff', 0)
        diff_cell = ws_total_mismatch.cell(row, 13, diff)
        diff_cell.font = Font(bold=True, color="16A34A" if diff > 0 else "DC2626")
        note = "Run camp to reduce point in CNV" if diff > 0 else "Run camp to increase point in CNV"
        ws_total_mismatch.cell(row, 14, note)
        row += 1

    for col_letter, width in zip('ABCDEFGHIJKLMN', [15, 12, 25, 12, 10, 12, 12, 25, 12, 10, 14, 12, 10, 35]):
        ws_total_mismatch.column_dimensions[col_letter].width = width

    # ========================================================================
    # ZALO MINI APP SHEET
    # ========================================================================
    zalo_blue_fill = PatternFill(start_color="0068FF", end_color="0068FF", fill_type="solid")
    zalo_headers = [
        'CNV ID', 'Phone', 'Full Name', 'Level', 'Email',
        'Reg Date', 'Points', 'Mini App', 'Follow OA', 'In POS'
    ]

    ws_zalo_app = wb.create_sheet("Zalo Mini App")
    for col_idx, header in enumerate(zalo_headers, 1):
        cell = ws_zalo_app.cell(1, col_idx, header)
        cell.fill = zalo_blue_fill
        cell.font = header_font
        cell.alignment = header_align
    for c in (zalo_mini_app_list or []):
        rd = c.get('cnv_created_at')
        ws_zalo_app.append([
            c.get('cnv_id', ''), c.get('phone', ''),
            f"{c.get('last_name') or ''} {c.get('first_name') or ''}".strip() or '-',
            c.get('level_name') or '-', c.get('email') or '-',
            str(rd.date()) if hasattr(rd, 'date') else (str(rd)[:10] if rd else '-'),
            float(c.get('points') or 0),
            'Yes' if c.get('zalo_app_id') else 'No',
            'Yes' if c.get('zalo_oa_id') else 'No',
            'Yes' if c.get('in_pos') else 'No',
        ])
    for col, w in zip('ABCDEFGHIJ', [15, 15, 25, 12, 30, 12, 10, 10, 10, 8]):
        ws_zalo_app.column_dimensions[col].width = w

    # ========================================================================
    # ZALO FOLLOW OA SHEET
    # ========================================================================
    zalo_cyan_fill = PatternFill(start_color="00AAFF", end_color="00AAFF", fill_type="solid")
    ws_zalo_oa = wb.create_sheet("Zalo Follow OA")
    for col_idx, header in enumerate(zalo_headers, 1):
        cell = ws_zalo_oa.cell(1, col_idx, header)
        cell.fill = zalo_cyan_fill
        cell.font = header_font
        cell.alignment = header_align
    for c in (zalo_oa_list or []):
        rd = c.get('cnv_created_at')
        ws_zalo_oa.append([
            c.get('cnv_id', ''), c.get('phone', ''),
            f"{c.get('last_name') or ''} {c.get('first_name') or ''}".strip() or '-',
            c.get('level_name') or '-', c.get('email') or '-',
            str(rd.date()) if hasattr(rd, 'date') else (str(rd)[:10] if rd else '-'),
            float(c.get('points') or 0),
            'Yes' if c.get('zalo_app_id') else 'No',
            'Yes' if c.get('zalo_oa_id') else 'No',
            'Yes' if c.get('in_pos') else 'No',
        ])
    for col, w in zip('ABCDEFGHIJ', [15, 15, 25, 12, 30, 12, 10, 10, 10, 8]):
        ws_zalo_oa.column_dimensions[col].width = w

    # ========================================================================
    # ZALO MINI APP NOT ACTIVE SHEET (download only, not shown on UI)
    # ========================================================================
    zalo_inactive_fill = PatternFill(start_color="888888", end_color="888888", fill_type="solid")
    ws_zalo_inactive = wb.create_sheet("Zalo Not Active")
    for col_idx, header in enumerate(zalo_headers, 1):
        cell = ws_zalo_inactive.cell(1, col_idx, header)
        cell.fill = zalo_inactive_fill
        cell.font = header_font
        cell.alignment = header_align
    for c in (zalo_mini_app_inactive_list or []):
        rd = c.get('cnv_created_at')
        ws_zalo_inactive.append([
            c.get('cnv_id', ''), c.get('phone', ''),
            f"{c.get('last_name') or ''} {c.get('first_name') or ''}".strip() or '-',
            c.get('level_name') or '-', c.get('email') or '-',
            str(rd.date()) if hasattr(rd, 'date') else (str(rd)[:10] if rd else '-'),
            float(c.get('points') or 0),
            'No',
            'Yes' if c.get('zalo_oa_id') else 'No',
            'Yes' if c.get('in_pos') else 'No',
        ])
    for col, w in zip('ABCDEFGHIJ', [15, 15, 25, 12, 30, 12, 10, 10, 10, 8]):
        ws_zalo_inactive.column_dimensions[col].width = w

    # ========================================================================
    # ALL CNV CUSTOMERS SHEET (filtered by period if dates provided)
    # ========================================================================
    cnv_all_fill = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
    cnv_all_headers = [
        'CNV ID', 'Phone', 'Full Name', 'Level', 'Email',
        'Reg Date', 'Points', 'Used Points', 'Total Points',
        'Zalo App ID', 'Zalo OA ID', 'Zalo Created At'
    ]

    sheet_title = f"CNV {date_from} to {date_to}" if (date_from and date_to) else "All CNV Customers"
    if len(sheet_title) > 31:
        sheet_title = sheet_title[:31]

    ws_cnv_all_full = wb.create_sheet(sheet_title)
    for col_idx, header in enumerate(cnv_all_headers, 1):
        cell = ws_cnv_all_full.cell(1, col_idx, header)
        cell.fill = cnv_all_fill
        cell.font = header_font
        cell.alignment = header_align

    cnv_all_qs = cnv_base.order_by('cnv_created_at')
    if date_from and date_to:
        cnv_all_qs = cnv_all_qs.filter(cnv_created_at__gte=date_from, cnv_created_at__lte=date_to)

    row = 2
    for cust in cnv_all_qs:
        full_name = getattr(cust, 'full_name', None) or (
            f"{cust.last_name or ''} {cust.first_name or ''}".strip()
        )
        ws_cnv_all_full.cell(row, 1, cust.cnv_id)
        ws_cnv_all_full.cell(row, 2, cust.phone or '-')
        ws_cnv_all_full.cell(row, 3, full_name or '-')
        ws_cnv_all_full.cell(row, 4, cust.level_name or '-')
        ws_cnv_all_full.cell(row, 5, cust.email or '-')
        ws_cnv_all_full.cell(row, 6, str(cust.cnv_created_at.date()) if cust.cnv_created_at else '-')
        ws_cnv_all_full.cell(row, 7, float(cust.points or 0))
        ws_cnv_all_full.cell(row, 8, float(cust.used_points or 0))
        ws_cnv_all_full.cell(row, 9, float(cust.total_points or 0))
        ws_cnv_all_full.cell(row, 10, cust.zalo_app_id or '')
        ws_cnv_all_full.cell(row, 11, cust.zalo_oa_id or '')
        zalo_dt = cust.zalo_app_created_at
        ws_cnv_all_full.cell(row, 12, str(zalo_dt.date()) if zalo_dt else '')
        row += 1

    for col, w in zip('ABCDEFGHIJKL', [15, 15, 25, 12, 30, 12, 10, 12, 12, 15, 15, 16]):
        ws_cnv_all_full.column_dimensions[col].width = w

    return wb


# ── Per-tab export ─────────────────────────────────────────────────────────────
# Must be defined AFTER all _create_* functions.

_TAB_SHEETS = {
    "grade":      ([_create_grade_sheet],                           "By VIP Grade"),
    "season":     ([_create_season_sheet],                          "By Season"),
    "month":      ([_create_month_sheet],                           "By Month"),
    "week":       ([_create_week_sheet],                            "By Week"),
    "shop":       ([_create_shop_sheet, _create_shop_detail_sheet], "By Shop"),
    "grade_all":      ([_create_grade_comparison_sheet],   "By Grade - All Shops"),
    "season_all":     ([_create_season_comparison_sheet],  "By Season - All Shops"),
    "month_all":      ([_create_month_comparison_sheet],   "By Month - All Shops"),
    "week_all":       ([_create_week_comparison_sheet],    "By Week - All Shops"),
    # Aliases matching the tab AJAX keys used in templates
    "grade_allshops":  ([_create_grade_comparison_sheet],  "By Grade - All Shops"),
    "season_allshops": ([_create_season_comparison_sheet], "By Season - All Shops"),
    "month_allshops":  ([_create_month_comparison_sheet],  "By Month - All Shops"),
    "week_allshops":   ([_create_week_comparison_sheet],   "By Week - All Shops"),
}


def _prepend_filter_info(wb, date_from=None, date_to=None, shop_group=None):
    """Insert a filter-info row at the top of every sheet in the workbook."""
    parts = []
    if date_from and date_to:
        parts.append(f"Period: {date_from} → {date_to}")
    if shop_group:
        parts.append(f"Shop Group: {shop_group}")
    xl_prepend_filter_row(wb, parts)


def export_tab_to_excel(tab, data, date_from=None, date_to=None, shop_group=None):
    """Export a single analytics tab to Excel (tab sheet(s) only, with filter row)."""
    if tab not in _TAB_SHEETS:
        return None

    wb = Workbook()
    wb.remove(wb.active)  # drop default blank sheet
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    creators, _ = _TAB_SHEETS[tab]
    for fn in creators:
        fn(wb, data, header_fill, header_font, header_align)

    _prepend_filter_info(wb, date_from=date_from, date_to=date_to, shop_group=shop_group)
    return wb


# ── CNV per-tab export ────────────────────────────────────────────────────────

def _cnv_hdr(ws, headers, fill, font, align):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = fill; c.font = font; c.alignment = align


def _build_cnv_used_points_ws(wb, data, hf, font, align):
    ws = wb.create_sheet("CNV Used Points")
    _cnv_hdr(ws, ["CNV ID", "Phone", "Full Name", "Level", "Email",
                  "Reg Date", "Points", "Used Pts", "Total Pts", "In POS"], hf, font, align)
    for r, c in enumerate(data.get("cnv_used_points_list") or [], 2):
        ws.cell(r, 1, c.get("cnv_id", ""))
        ws.cell(r, 2, c.get("phone", ""))
        ws.cell(r, 3, f"{c.get('last_name') or ''} {c.get('first_name') or ''}".strip())
        ws.cell(r, 4, c.get("level_name", ""))
        ws.cell(r, 5, c.get("email", "") or "")
        ws.cell(r, 6, str(c["cnv_created_at"].date()) if c.get("cnv_created_at") else "")
        ws.cell(r, 7, c.get("points") or 0)
        ws.cell(r, 8, c.get("used_points") or 0)
        ws.cell(r, 9, c.get("total_points") or 0)
        ws.cell(r, 10, "Yes" if c.get("in_pos") else "No")
    for col, w in zip("ABCDEFGHIJ", [14, 14, 22, 10, 28, 12, 10, 10, 10, 8]):
        ws.column_dimensions[col].width = w


def _build_mismatch_ws(wb, data, key, title, hf, font, align):
    ws = wb.create_sheet(title)
    # Columns: POS Pts + Used + Net, then CNV Pts → Used Pts → Total Pts (order matches UI tables)
    _cnv_hdr(ws, [
        "Phone", "POS VIP ID", "POS Name", "POS Grade",
        "POS Pts", "POS Used Pts", "POS Net Pts",
        "CNV ID", "CNV Name", "CNV Level",
        "CNV Points", "CNV Used Points", "CNV Total Points",
        "Diff",
    ], hf, font, align)
    for r, c in enumerate(data.get(key) or [], 2):
        ws.cell(r,  1, c.get("phone", ""))
        ws.cell(r,  2, c.get("pos_vip_id", ""))
        ws.cell(r,  3, c.get("pos_name", ""))
        ws.cell(r,  4, c.get("pos_grade", ""))
        ws.cell(r,  5, c.get("pos_points") or 0)
        ws.cell(r,  6, c.get("pos_used_points") or 0)
        ws.cell(r,  7, c.get("pos_net_points") or 0)
        ws.cell(r,  8, c.get("cnv_id", ""))
        ws.cell(r,  9, c.get("cnv_name", ""))
        ws.cell(r, 10, c.get("cnv_level", ""))
        ws.cell(r, 11, c.get("cnv_points") or 0)
        ws.cell(r, 12, c.get("cnv_used_points") or 0)
        ws.cell(r, 13, c.get("cnv_total_points") or 0)
        ws.cell(r, 14, c.get("diff") or 0)
    for col, w in zip("ABCDEFGHIJKLMN", [14, 12, 22, 10, 10, 12, 10, 14, 22, 10, 10, 12, 12, 10]):
        ws.column_dimensions[col].width = w


def _build_zalo_ws(wb, data, key, title, hf, font, align):
    ws = wb.create_sheet(title)
    _cnv_hdr(ws, ["CNV ID", "Phone", "Full Name", "Level", "Email",
                  "Reg Date", "Points", "Zalo App ID", "Zalo OA ID", "In POS"], hf, font, align)
    for c in (data.get(key) or []):
        rd = c.get("cnv_created_at")
        ws.append([
            c.get("cnv_id", ""),
            c.get("phone", ""),
            f"{c.get('last_name') or ''} {c.get('first_name') or ''}".strip(),
            c.get("level_name", ""),
            c.get("email", "") or "",
            str(rd.date()) if rd else "",
            c.get("points") or 0,
            c.get("zalo_app_id", "") or "",
            c.get("zalo_oa_id", "") or "",
            "Yes" if c.get("in_pos") else "No",
        ])
    for col, w in zip("ABCDEFGHIJ", [14, 14, 22, 10, 28, 12, 10, 20, 20, 8]):
        ws.column_dimensions[col].width = w


def _build_pos_only_ws(wb, data, key, title, hf, font, align):
    ws = wb.create_sheet(title)
    _cnv_hdr(ws, ["VIP ID", "Phone", "Name", "VIP Grade", "Email", "Reg Date", "Points"],
             hf, font, align)
    for r, c in enumerate(data.get(key) or [], 2):
        ws.cell(r, 1, c.get("vip_id", ""))
        ws.cell(r, 2, c.get("phone", ""))
        ws.cell(r, 3, c.get("name", ""))
        ws.cell(r, 4, c.get("vip_grade", ""))
        ws.cell(r, 5, c.get("email", "") or "")
        ws.cell(r, 6, str(c["registration_date"]) if c.get("registration_date") else "")
        ws.cell(r, 7, c.get("points") or 0)
    for col, w in zip("ABCDEFG", [12, 14, 22, 10, 28, 12, 10]):
        ws.column_dimensions[col].width = w


def _build_cnv_only_ws(wb, data, key, title, hf, font, align):
    ws = wb.create_sheet(title)
    _cnv_hdr(ws, ["CNV ID", "Phone", "Full Name", "Level", "Email",
                  "Reg Date", "Points", "Used Pts", "Total Pts"], hf, font, align)
    for r, c in enumerate(data.get(key) or [], 2):
        ws.cell(r, 1, c.get("cnv_id", ""))
        ws.cell(r, 2, c.get("phone", ""))
        ws.cell(r, 3, f"{c.get('last_name') or ''} {c.get('first_name') or ''}".strip())
        ws.cell(r, 4, c.get("level_name", ""))
        ws.cell(r, 5, c.get("email", "") or "")
        ws.cell(r, 6, str(c["cnv_created_at"].date()) if c.get("cnv_created_at") else "")
        ws.cell(r, 7, c.get("points") or 0)
        ws.cell(r, 8, c.get("used_points") or 0)
        ws.cell(r, 9, c.get("total_points") or 0)
    for col, w in zip("ABCDEFGHI", [14, 14, 22, 10, 28, 12, 10, 10, 10]):
        ws.column_dimensions[col].width = w


def _build_cnv_breakdown_ws(wb, breakdown, key, title, has_shop, hf, font, align):
    """Build one Registration Breakdown sheet.
    has_shop=True → extra 'Shop' column between label and data cols.
    """
    ws = wb.create_sheet(title)
    if has_shop:
        headers = ["Period", "Shop",
                   "New POS (INV)", "New POS (NO INV)", "New POS (TOTAL)", "POS Only",
                   "New CNV", "CNV Only",
                   "Zalo App", "% App/CNV", "Zalo OA", "% OA/CNV"]
    else:
        headers = ["Period",
                   "New POS (INV)", "New POS (NO INV)", "New POS (TOTAL)", "POS Only",
                   "New CNV", "CNV Only",
                   "Zalo App", "% App/CNV", "Zalo OA", "% OA/CNV"]
    _cnv_hdr(ws, headers, hf, font, align)

    rows = breakdown.get(key) or []
    for r_idx, r in enumerate(rows, 2):
        col = 1
        ws.cell(r_idx, col, r.get("label", "")); col += 1
        if has_shop:
            ws.cell(r_idx, col, r.get("shop", "")); col += 1
        ws.cell(r_idx, col, r.get("new_pos_inv", 0));    col += 1
        ws.cell(r_idx, col, r.get("new_pos_no_inv", 0)); col += 1
        ws.cell(r_idx, col, r.get("new_pos", 0));        col += 1
        ws.cell(r_idx, col, r.get("new_pos_only", 0));   col += 1
        ws.cell(r_idx, col, r.get("new_cnv", 0));        col += 1
        ws.cell(r_idx, col, r.get("new_cnv_only", 0));   col += 1
        ws.cell(r_idx, col, r.get("zalo_app", 0));       col += 1
        ws.cell(r_idx, col, r.get("zalo_app_pct", 0.0)); col += 1
        ws.cell(r_idx, col, r.get("zalo_oa", 0));        col += 1
        ws.cell(r_idx, col, r.get("zalo_oa_pct", 0.0))

    if has_shop:
        for col_letter, w in zip("ABCDEFGHIJKL", [16, 22, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]):
            ws.column_dimensions[col_letter].width = w
    else:
        for col_letter, w in zip("ABCDEFGHIJK", [16, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]):
            ws.column_dimensions[col_letter].width = w


def _build_cnv_period_grouped_ws(wb, breakdown, key, title, hf, font, align):
    """Build period-grouped Registration Breakdown sheet matching UI tab format.
    Each period = one table block: bold header row + shop-data header + shop rows.
    breakdown[key] is flat [{label, shop, new_pos_inv, ...}].
    """
    ws = wb.create_sheet(title)
    shop_headers = ["Shop", "POS(INV)", "POS(NO INV)", "POS Total", "POS Only",
                    "CNV", "CNV Only", "Zalo App", "%App", "Zalo OA", "%OA"]
    n_cols = len(shop_headers)

    rows = breakdown.get(key) or []
    # Group by label preserving order
    from collections import OrderedDict
    period_map: dict = OrderedDict()
    for r in rows:
        lbl = r.get("label", "")
        period_map.setdefault(lbl, []).append(r)

    current_row = 1
    for label, shop_rows in period_map.items():
        # Period header
        ws.cell(current_row, 1, label).font = Font(bold=True, size=12)
        if n_cols > 1:
            ws.merge_cells(f'A{current_row}:{get_column_letter(n_cols)}{current_row}')
        current_row += 1

        # Column headers
        for col_num, h in enumerate(shop_headers, 1):
            c = ws.cell(current_row, col_num, h)
            c.fill = hf; c.font = font; c.alignment = align
        current_row += 1

        # Shop data rows
        for r in shop_rows:
            ws.cell(current_row, 1,  r.get("shop", ""))
            ws.cell(current_row, 2,  r.get("new_pos_inv", 0))
            ws.cell(current_row, 3,  r.get("new_pos_no_inv", 0))
            ws.cell(current_row, 4,  r.get("new_pos", 0))
            ws.cell(current_row, 5,  r.get("new_pos_only", 0))
            ws.cell(current_row, 6,  r.get("new_cnv", 0))
            ws.cell(current_row, 7,  r.get("new_cnv_only", 0))
            ws.cell(current_row, 8,  r.get("zalo_app", 0))
            ws.cell(current_row, 9,  r.get("zalo_app_pct", 0.0))
            ws.cell(current_row, 10, r.get("zalo_oa", 0))
            ws.cell(current_row, 11, r.get("zalo_oa_pct", 0.0))
            current_row += 1

        current_row += 1  # blank row between periods

    ws.column_dimensions['A'].width = 32
    for col_letter, w in zip("BCDEFGHIJK", [10, 10, 10, 10, 10, 10, 10, 8, 10, 8]):
        ws.column_dimensions[col_letter].width = w


def _build_cnv_shop_grouped_ws(wb, breakdown, key, title, time_col_name, hf, font, align):
    """Build a By-Shop-then-time Registration Breakdown sheet.
    breakdown[key] is a list of {shop, rows: [{label, new_pos_inv, ...}]}.
    """
    ws = wb.create_sheet(title)
    headers = ["Shop", time_col_name,
               "New POS (INV)", "New POS (NO INV)", "New POS (TOTAL)", "POS Only",
               "New CNV", "CNV Only",
               "Zalo App", "% App/CNV", "Zalo OA", "% OA/CNV"]
    _cnv_hdr(ws, headers, hf, font, align)

    r_idx = 2
    shop_groups = breakdown.get(key) or []
    for sg in shop_groups:
        shop_name = sg.get("shop", "")
        for r in sg.get("rows", []):
            col = 1
            ws.cell(r_idx, col, shop_name);                   col += 1
            ws.cell(r_idx, col, r.get("label", ""));          col += 1
            ws.cell(r_idx, col, r.get("new_pos_inv", 0));     col += 1
            ws.cell(r_idx, col, r.get("new_pos_no_inv", 0));  col += 1
            ws.cell(r_idx, col, r.get("new_pos", 0));         col += 1
            ws.cell(r_idx, col, r.get("new_pos_only", 0));    col += 1
            ws.cell(r_idx, col, r.get("new_cnv", 0));         col += 1
            ws.cell(r_idx, col, r.get("new_cnv_only", 0));    col += 1
            ws.cell(r_idx, col, r.get("zalo_app", 0));        col += 1
            ws.cell(r_idx, col, r.get("zalo_app_pct", 0.0));  col += 1
            ws.cell(r_idx, col, r.get("zalo_oa", 0));         col += 1
            ws.cell(r_idx, col, r.get("zalo_oa_pct", 0.0))
            r_idx += 1

    for col_letter, w in zip("ABCDEFGHIJKL", [22, 16, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]):
        ws.column_dimensions[col_letter].width = w


def _build_cnv_shop_detail_ws(wb, breakdown, hf, font, align):
    """Build a per-shop nested detail sheet.

    Mirrors Sales Analytics _create_shop_detail_sheet concept:
    for each shop → dark header row → summary row → By Season / By Month / By Week sections.

    breakdown["shop_detail"] is a list of:
        {"shop": str, "summary": row_dict|None, "by_season": [...], "by_month": [...], "by_week": [...]}
    """
    ws = wb.create_sheet("By Shop - Detail")

    SHOP_FILL  = PatternFill(start_color="1a252f", end_color="1a252f", fill_type="solid")
    SHOP_FONT  = Font(bold=True, color="FFFFFF", size=11)
    SEC_FILL   = PatternFill(start_color="2980b9", end_color="2980b9", fill_type="solid")
    SEC_FONT   = Font(bold=True, color="FFFFFF", size=9)
    TOTAL_FONT = Font(bold=True)

    DATA_HDRS = ["Period",
                 "New POS (INV)", "New POS (NO INV)", "New POS (TOTAL)", "POS Only",
                 "New CNV", "CNV Only",
                 "Zalo App", "% App/CNV", "Zalo OA", "% OA/CNV"]
    ncols = len(DATA_HDRS)

    def _write_data_headers(row):
        for ci, h in enumerate(DATA_HDRS, 1):
            c = ws.cell(row, ci, h)
            c.fill = hf; c.font = font; c.alignment = align

    def _write_data_row(row, r):
        ws.cell(row, 1,  r.get("label", ""))
        ws.cell(row, 2,  r.get("new_pos_inv",   0))
        ws.cell(row, 3,  r.get("new_pos_no_inv", 0))
        ws.cell(row, 4,  r.get("new_pos",        0))
        ws.cell(row, 5,  r.get("new_pos_only",   0))
        ws.cell(row, 6,  r.get("new_cnv",        0))
        ws.cell(row, 7,  r.get("new_cnv_only",   0))
        ws.cell(row, 8,  r.get("zalo_app",       0))
        ws.cell(row, 9,  r.get("zalo_app_pct",   0.0))
        ws.cell(row, 10, r.get("zalo_oa",        0))
        ws.cell(row, 11, r.get("zalo_oa_pct",    0.0))

    cur = 1
    for sd in (breakdown.get("shop_detail") or []):
        shop_name = sd.get("shop", "")

        # ── Shop header row ──────────────────────────────────────────────
        ws.cell(cur, 1, f"  {shop_name}")
        ws.cell(cur, 1).fill = SHOP_FILL
        ws.cell(cur, 1).font = SHOP_FONT
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
        cur += 1

        # ── Summary totals row ───────────────────────────────────────────
        summary = sd.get("summary")
        if summary:
            ws.cell(cur, 1, "TOTAL")
            ws.cell(cur, 1).font = TOTAL_FONT
            _write_data_row(cur, {**summary, "label": "TOTAL"})
            cur += 1

        # ── Sub-sections: By Season / By Month / By Week ─────────────────
        for sub_key, sub_label in (("by_season", "By Season"),
                                   ("by_month",  "By Month"),
                                   ("by_week",   "By Week")):
            rows = sd.get(sub_key) or []
            if not rows:
                continue

            # Section label
            ws.cell(cur, 1, sub_label)
            ws.cell(cur, 1).fill = SEC_FILL
            ws.cell(cur, 1).font = SEC_FONT
            ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=ncols)
            cur += 1

            # Column headers
            _write_data_headers(cur); cur += 1

            # Data
            for r in rows:
                _write_data_row(cur, r); cur += 1

            cur += 1  # blank spacer after each section

        cur += 1  # blank spacer between shops

    ws.column_dimensions["A"].width = 18
    for col_letter in "BCDEFGHIJK":
        ws.column_dimensions[col_letter].width = 12


_CNV_TAB_SHEETS = {
    "points":       ["cnv_used_points", "points_mismatch", "total_points_mismatch"],
    "zalo":         ["zalo_mini_app", "zalo_mini_app_inactive", "zalo_oa"],
    "pos_cnv":      ["pos_only_all",    "cnv_only_all", "pos_only_period", "cnv_only_period"],
    "breakdown":    ["bd_season", "bd_month", "bd_week", "bd_shop", "bd_shop_detail",
                     "bd_season_shop", "bd_month_shop", "bd_week_shop"],
    # Individual breakdown sub-tabs
    "bd_season":      ["bd_season"],
    "bd_month":       ["bd_month"],
    "bd_week":        ["bd_week"],
    "bd_shop":        ["bd_shop", "bd_shop_detail"],
    "bd_season_shop": ["bd_season_shop"],
    "bd_month_shop":  ["bd_month_shop"],
    "bd_week_shop":   ["bd_week_shop"],
    # Aliases matching template tab AJAX keys — period-grouped format matching UI
    "bd_season_allshops": ["bd_season_shop_grouped"],
    "bd_month_allshops":  ["bd_month_shop_grouped"],
    "bd_week_allshops":   ["bd_week_shop_grouped"],
}


def export_cnv_tab_to_excel(tab, data, date_from=None, date_to=None):
    """Export a single CNV customer-comparison tab (1 sheet per table)."""
    if tab not in _CNV_TAB_SHEETS:
        return None

    wb = Workbook()
    wb.remove(wb.active)
    hf  = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    fnt = Font(bold=True, color="FFFFFF")
    aln = Alignment(horizontal="center", vertical="center")

    bd = data.get("breakdown") or {}

    builders = {
        "cnv_used_points":     lambda: _build_cnv_used_points_ws(wb, data, hf, fnt, aln),
        "points_mismatch":     lambda: _build_mismatch_ws(wb, data, "points_mismatch",       "Points Mismatch",       hf, fnt, aln),
        "total_points_mismatch": lambda: _build_mismatch_ws(wb, data, "total_points_mismatch", "Total Points Mismatch", hf, fnt, aln),
        "zalo_mini_app":          lambda: _build_zalo_ws(wb, data, "zalo_mini_app_list",          "Zalo Mini App",    hf, fnt, aln),
        "zalo_mini_app_inactive": lambda: _build_zalo_ws(wb, data, "zalo_mini_app_inactive_list", "Zalo Not Active",  hf, fnt, aln),
        "zalo_oa":                lambda: _build_zalo_ws(wb, data, "zalo_oa_list",                "Zalo Follow OA",   hf, fnt, aln),
        "pos_only_all":        lambda: _build_pos_only_ws(wb, data, "pos_only_all",    "POS Only - All Time",    hf, fnt, aln),
        "cnv_only_all":        lambda: _build_cnv_only_ws(wb, data, "cnv_only_all",    "CNV Only - All Time",    hf, fnt, aln),
        "pos_only_period":     lambda: _build_pos_only_ws(wb, data, "pos_only_period", "POS Only - Period",      hf, fnt, aln),
        "cnv_only_period":     lambda: _build_cnv_only_ws(wb, data, "cnv_only_period", "CNV Only - Period",      hf, fnt, aln),
        # breakdown sheets
        "bd_season":      lambda: _build_cnv_breakdown_ws(wb, bd, "season",      "By Season",           False, hf, fnt, aln),
        "bd_month":       lambda: _build_cnv_breakdown_ws(wb, bd, "month",       "By Month",            False, hf, fnt, aln),
        "bd_week":        lambda: _build_cnv_breakdown_ws(wb, bd, "week",        "By Week",             False, hf, fnt, aln),
        "bd_shop":        lambda: _build_cnv_breakdown_ws(wb, bd, "shop",        "By Shop",             False, hf, fnt, aln),
        "bd_shop_detail": lambda: _build_cnv_shop_detail_ws(wb, bd, hf, fnt, aln),
        "bd_season_shop": lambda: _build_cnv_breakdown_ws(wb, bd, "season_shop", "By Season - All Shops", True, hf, fnt, aln),
        "bd_month_shop":  lambda: _build_cnv_breakdown_ws(wb, bd, "month_shop",  "By Month - All Shops",  True, hf, fnt, aln),
        "bd_week_shop":   lambda: _build_cnv_breakdown_ws(wb, bd, "week_shop",   "By Week - All Shops",   True, hf, fnt, aln),
        # Period-grouped format (matching UI tab layout)
        "bd_season_shop_grouped": lambda: _build_cnv_period_grouped_ws(wb, bd, "season_shop", "By Season - All Shops", hf, fnt, aln),
        "bd_month_shop_grouped":  lambda: _build_cnv_period_grouped_ws(wb, bd, "month_shop",  "By Month - All Shops",  hf, fnt, aln),
        "bd_week_shop_grouped":   lambda: _build_cnv_period_grouped_ws(wb, bd, "week_shop",   "By Week - All Shops",   hf, fnt, aln),
        "bd_shop_season": lambda: _build_cnv_shop_grouped_ws(wb, bd, "shop_season", "By Shop - By Season", "Season", hf, fnt, aln),
        "bd_shop_month":  lambda: _build_cnv_shop_grouped_ws(wb, bd, "shop_month",  "By Shop - By Month",  "Month",  hf, fnt, aln),
        "bd_shop_week":   lambda: _build_cnv_shop_grouped_ws(wb, bd, "shop_week",   "By Shop - By Week",   "Week",   hf, fnt, aln),
    }

    for key in _CNV_TAB_SHEETS[tab]:
        # skip period sheets if no data (no date filter applied)
        if key.endswith("_period") and not data.get(key):
            continue
        builders[key]()

    # filter info row
    parts = []
    if date_from and date_to:
        parts.append(f"Period: {date_from} → {date_to}")
    if parts:
        line = "  |  ".join(parts)
        for ws in wb.worksheets:
            ws.insert_rows(1)
            ws["A1"] = line
            ws["A1"].font = Font(italic=True, color="888888", size=9)   
    return wb


# ─────────────────────────────────────────────────────────────────────────────
# CHART PAGE EXCEL EXPORTS  (Option A: JS reads UI state → server generates)
# ─────────────────────────────────────────────────────────────────────────────

# ── Shared helpers ────────────────────────────────────────────────────────────

_SALES_METRIC_LABEL = {
    'total_customers': 'Active Customers',
    'returning_customers': 'Returning Customers',
    'return_rate': 'Return Rate (%)',
    'returning_invoices': 'INV(RET)',
    'total_invoices': 'INV(CUS)',
    'total_invoices_with_vip0': 'Total Invoices',
    'returning_amount': 'AMT(RET)',
    'total_amount': 'AMT(CUS)',
    'total_amount_with_vip0': 'Total Amount',
}
_CUSTOMER_METRIC_LABEL = {
    'new_pos_inv': 'NEW POS (INV)', 'new_pos_no_inv': 'NEW POS (NO INV)',
    'new_pos': 'NEW POS (TOTAL)', 'new_pos_only': 'POS ONLY',
    'new_cnv': 'NEW CNV', 'new_cnv_only': 'CNV ONLY',
    'zalo_app': 'ZALO APP', 'zalo_app_pct': '% APP/CNV',
    'zalo_oa': 'ZALO OA', 'zalo_oa_pct': '% OA/CNV',
}
_COUPON_METRIC_LABEL = {
    'used': 'Used Coupons', 'pct_of_used': '% Used / Used All',
    'coupon_amount': 'Coupon Amount (VND)', 'unique_amount': 'Total Amount Unique (VND)',
}
_ALL_METRIC_LABELS = {**_SALES_METRIC_LABEL, **_CUSTOMER_METRIC_LABEL, **_COUPON_METRIC_LABEL}

# Sales: map xaxis name → (by_shop sub-key, period label key in each entry)
# NOTE: by_week entries use 'week_label' (display: "Week 1 (6/1-12/1)") and
#       'week_sort' (key: "2025-W01"). We use week_label for display on chart.
_SALES_AXIS = {
    'season': ('by_session', 'session'),
    'month':  ('by_month',   'month'),
    'week':   ('by_week',    'week_label'),   # was 'week' — no such key in data
    'year':   ('by_year',    'year'),
}
# Customer: shop_stats[shop_name] keys and period label keys
_CUSTOMER_AXIS = {
    'season': ('season', 'label'),
    'month':  ('month',  'month'),
    'week':   ('week',   'week'),
}
# Coupon trend shop/camp series: xaxis → key in shop_series/camp_series
_COUPON_AXIS_LABEL = {
    'season': 'season', 'month': 'month', 'week': 'week', 'year': 'year',
}


def _is_pct_metric(metric):
    return 'rate' in metric or 'pct' in metric


def _is_amount_metric(metric):
    return 'amount' in metric


def _fmt_metric(metric):
    """Format a cell value: percentages as float (0..1), amounts as #,##0."""
    if _is_pct_metric(metric):
        return ('pct', '0.0%')
    if _is_amount_metric(metric):
        return ('num', '#,##0')
    return ('int', '0')


def _build_sales_shop_map(data, xaxis, metric, selected_shops):
    """
    Returns (ordered_periods, {shop_name: {period_key: value}}).
    selected_shops: list of shop names; empty = all shops.
    """
    axis_key, period_lbl_key = _SALES_AXIS.get(xaxis, _SALES_AXIS['month'])
    result = {}
    all_periods = None

    for shop_entry in data.get('by_shop', []):
        sname = shop_entry['shop_name']
        if selected_shops and sname not in selected_shops:
            continue
        period_list = shop_entry.get(axis_key, [])
        if all_periods is None:
            all_periods = [e[period_lbl_key] for e in period_list]
        # return_rate is stored 0-100 from backend; convert to fraction for Excel %
        def _v(e):
            v = e.get(metric, 0) or 0
            return v / 100 if _is_pct_metric(metric) and v > 1 else v
        result[sname] = {e[period_lbl_key]: _v(e) for e in period_list}

    if all_periods is None:
        # fallback: periods from aggregate data
        agg_key = axis_key  # same key name in data top level
        all_periods = [e[period_lbl_key] for e in data.get(agg_key, [])]

    return all_periods or [], result


def _extract_year_period(full_period, xaxis):
    """Returns (year_str, period_within_year_str)."""
    s = str(full_period)
    if xaxis == 'month' and len(s) >= 7 and s[4] == '-':
        return s[:4], s[5:]
    if xaxis == 'week' and '-' in s:
        parts = s.split('-', 1)
        return parts[0], parts[1]
    if xaxis == 'season' and s and s[-4:].isdigit():
        return s[-4:], s[:-5].strip() if len(s) > 4 else s
    return s, s


def _build_yoy_data(shop_map, periods, xaxis, metric):
    """
    Group summed values by year.
    Returns (period_within_year_list, year_list, {year: {period_within: value}}).
    """
    year_period_map = {}
    for full_period in periods:
        value = sum(sm.get(full_period, 0) for sm in shop_map.values())
        year_str, within = _extract_year_period(full_period, xaxis)
        if year_str not in year_period_map:
            year_period_map[year_str] = {}
        year_period_map[year_str][within] = year_period_map[year_str].get(within, 0) + value

    years = sorted(year_period_map.keys())
    all_within = sorted(set(p for y in year_period_map.values() for p in y.keys()))
    return all_within, years, year_period_map


# ── UI colour palettes — match chart.html JS exactly ─────────────────────────
# SHOP_PALETTE: used for line chart series (one colour per shop)
_SHOP_PALETTE = [
    '#e6194B', '#3cb44b', '#4363d8', '#f58231', '#911eb4',
    '#42d4f4', '#f032e6', '#469990', '#9a6324', '#800000',
    '#dcbeff', '#aaffc3', '#808000', '#ffd8b1', '#000075',
    '#a9a9a9', '#ffe119', '#fabebe',
]
# COMP_YEAR_PALETTE: used for YOY bar chart series (one colour per year)
_COMP_YEAR_PALETTE = [
    '#1565C0', '#E53935', '#2E7D32', '#F57C00', '#6A1B9A',
    '#00838F', '#AD1457', '#558B2F', '#4527A0', '#FF6F00',
]
# Single-series bar chart colour (matches BAR_GRADIENT_COLOR in UI)
_BAR_SINGLE_COLOR = '#0047B3'

# Chart figure dimensions for rendering
_CHART_FIG_W_IN  = 14.0   # inches wide at render DPI
_CHART_FIG_H_IN  =  5.5   # inches tall at render DPI
_CHART_DPI       = 150    # render DPI — high quality PNG

# Display pixel size passed to openpyxl Image.
# Excel renders image at native pixel dimensions (no DPI rescaling).
# We tell openpyxl to display at screen-DPI (96) equivalent of our figure size:
#   display_px = fig_in × 96   →  14.0 × 96 = 1344px wide, 5.5 × 96 = 528px tall
# At default Excel row height 15pt (≈ 20px), 528px ≈ 26 rows.
_CHART_DISP_W_PX = int(_CHART_FIG_W_IN * 96)   # 1344
_CHART_DISP_H_PX = int(_CHART_FIG_H_IN * 96)   # 528

# Rows reserved above data table — must be > display height in rows + 2 buffer.
# 528px / 20px/row ≈ 26 rows → use 30 for safety.
_CHART_DATA_OFFSET = 30


def _hex_to_rgb(h):
    """'#RRGGBB' or 'RRGGBB' → (r,g,b) floats 0-1."""
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))


def _fmt_large(val, is_pct=False, is_amount=False):
    """Format y-axis tick label: compact thousands/millions, or percent."""
    if is_pct:
        return f'{val*100:.1f}%' if val < 1.01 else f'{val:.1f}%'
    if is_amount:
        if abs(val) >= 1_000_000_000:
            return f'{val/1_000_000_000:.1f}B'
        if abs(val) >= 1_000_000:
            return f'{val/1_000_000:.1f}M'
        if abs(val) >= 1_000:
            return f'{val/1_000:.0f}K'
    if abs(val) >= 1_000_000:
        return f'{val/1_000_000:.1f}M'
    if abs(val) >= 1_000:
        return f'{val/1_000:.0f}K'
    return f'{val:,.0f}'


def _chart_fig_setup(title, ylabel, n_periods, vertical_grid=False):
    """Create a matplotlib Figure styled to match the UI chart.

    vertical_grid=True  → add subtle vertical gridlines (for line charts).
                          Each vertical line aligns with an x tick, so when it
                          crosses a data line the dot at that intersection is
                          clearly visible — matching the UI tooltip crosshair feel.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    fig, ax = plt.subplots(figsize=(_CHART_FIG_W_IN, _CHART_FIG_H_IN), dpi=_CHART_DPI)

    # Background
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')

    # Horizontal gridlines — always present (light grey, matches UI)
    ax.yaxis.grid(True, color='#dee2e6', linewidth=0.6, linestyle='-', zorder=0)

    # Vertical gridlines — line charts only, even subtler than horizontal
    if vertical_grid:
        ax.xaxis.grid(True, color='#e9ecef', linewidth=0.5, linestyle='-', zorder=0)
    else:
        ax.xaxis.grid(False)

    ax.set_axisbelow(True)

    # Spines — keep left, remove top/right, make bottom subtle
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#dee2e6')
    ax.spines['bottom'].set_color('#dee2e6')

    # Title
    ax.set_title(title, fontsize=11, fontweight='bold', color='#2c3e50', pad=10, loc='left')

    # Y axis label
    ax.set_ylabel(ylabel, fontsize=9, color='#495057', labelpad=8)

    # Tick style
    ax.tick_params(axis='both', labelsize=8, colors='#495057', length=3)
    ax.tick_params(axis='x', rotation=30)

    # X tick label density — skip labels for large datasets
    if n_periods > 26:
        skip = max(2, n_periods // 26)
        ax.xaxis.set_major_locator(mticker.MultipleLocator(skip))

    return fig, ax


def _fig_to_xl_image(fig):
    """Render matplotlib figure to PNG bytes, wrap in openpyxl Image.

    Memory hygiene:
      - plt.close(fig) releases the matplotlib Figure immediately.
      - We copy bytes into a fresh BytesIO so the original buffer can be
        garbage-collected; openpyxl holds only the small fresh buffer until
        the workbook is saved.
    """
    import io
    import matplotlib.pyplot as plt
    from openpyxl.drawing.image import Image as XLImage

    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=_CHART_DPI,
                bbox_inches='tight', facecolor='white')
    plt.close(fig)            # release Figure memory now
    png_bytes = buf.getvalue()
    buf.close()               # release original buffer
    img = XLImage(io.BytesIO(png_bytes))  # fresh buffer owned by Image
    return img


def _write_chart_image_to_sheet(ws, img):
    """Anchor image at A2 with explicit display size so it never overlaps the data table.

    openpyxl passes the raw PNG pixel dimensions to Excel (no DPI correction).
    Setting img.width/img.height explicitly overrides this, telling Excel to
    display at exactly _CHART_DISP_W_PX × _CHART_DISP_H_PX screen pixels —
    equivalent to _CHART_FIG_W_IN × _CHART_FIG_H_IN inches at 96 DPI.
    """
    img.width  = _CHART_DISP_W_PX
    img.height = _CHART_DISP_H_PX
    img.anchor = 'A2'
    ws.add_image(img)


def _write_line_chart_sheet(wb, sheet_title, periods, shop_or_entity_map, metric,
                             x_label='Period', chart_title=''):
    """
    Sheet: matplotlib line chart image at top (B2), data table below (row 33+).
    One line per shop/entity, UI palette colours, legend at bottom.
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mticker

    ws = wb.create_sheet(title=sheet_title[:31])
    entities = list(shop_or_entity_map.keys())
    _, num_fmt = _fmt_metric(metric)
    is_pct    = _is_pct_metric(metric)
    is_amount = _is_amount_metric(metric)
    ylabel    = _ALL_METRIC_LABELS.get(metric, metric)

    # ── Data table (row 33+) ──────────────────────────────────────────────
    header_row = _CHART_DATA_OFFSET + 1
    data_start = header_row + 1
    ws.cell(row=header_row, column=1, value=x_label).font = XL_HDR_FONT
    ws.cell(row=header_row, column=1).fill  = XL_HDR_FILL
    ws.cell(row=header_row, column=1).alignment = XL_HDR_ALIGN
    for ci, name in enumerate(entities, 2):
        c = ws.cell(row=header_row, column=ci, value=name)
        c.font = XL_HDR_FONT; c.fill = XL_HDR_FILL; c.alignment = XL_HDR_ALIGN
    for ri, period in enumerate(periods, data_start):
        ws.cell(row=ri, column=1, value=str(period))
        for ci, name in enumerate(entities, 2):
            v = shop_or_entity_map[name].get(period, 0)
            c = ws.cell(row=ri, column=ci, value=v)
            c.number_format = num_fmt
    ws.column_dimensions['A'].width = 20
    for ci in range(2, 2 + len(entities)):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    if not periods or not entities:
        return ws

    # ── Matplotlib line chart ─────────────────────────────────────────────
    n = len(periods)
    # vertical_grid=True: subtle vertical lines at each x tick; dots at
    # intersections make each data point clearly visible (matches UI crosshair)
    fig, ax = _chart_fig_setup(chart_title or sheet_title, ylabel, n,
                               vertical_grid=True)
    x_idx = list(range(n))

    # Dot size scales down for large datasets to avoid clutter
    dot_size = 5 if n <= 26 else (4 if n <= 52 else 3)

    for i, name in enumerate(entities):
        color  = _SHOP_PALETTE[i % len(_SHOP_PALETTE)]
        values = [shop_or_entity_map[name].get(p, 0) for p in periods]
        ax.plot(x_idx, values,
                color=color, linewidth=2, label=name,
                # Filled dot with thin white edge — visible at vertical grid intersections
                marker='o', markersize=dot_size,
                markerfacecolor=color,
                markeredgecolor='white', markeredgewidth=0.8,
                zorder=4)   # above gridlines

    # X-axis ticks & labels — must align with vertical gridlines
    ax.set_xticks(x_idx)
    ax.set_xticklabels([str(p) for p in periods], rotation=35, ha='right', fontsize=7.5)
    if n > 26:
        skip = max(2, n // 26)
        ax.set_xticks(x_idx[::skip])
        ax.set_xticklabels([str(periods[j]) for j in x_idx[::skip]], rotation=35, ha='right', fontsize=7.5)

    # Y-axis formatter
    ax.yaxis.set_major_formatter(
        plt.FuncFormatter(lambda v, _: _fmt_large(v, is_pct=is_pct, is_amount=is_amount))
    )

    # Legend — bottom outside plot, up to 6 per row
    if entities:
        ncol = min(len(entities), 6)
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.20),
                  ncol=ncol, fontsize=8, frameon=True,
                  fancybox=False, edgecolor='#dee2e6',
                  handlelength=1.5, handleheight=0.8,
                  borderpad=0.6, columnspacing=1.0)

    fig.tight_layout(rect=[0, 0.10, 1, 1])

    img = _fig_to_xl_image(fig)
    _write_chart_image_to_sheet(ws, img)
    return ws


def _write_bar_chart_sheet(wb, sheet_title, periods, data_map, metric,
                            x_label='Period', chart_title=''):
    """
    Sheet: matplotlib bar chart image at top (B2), data table below (row 33+).
    data_map: {series_name: {period: value}} — multi-series (YOY)
           or {period: value}               — single-series (Period Totals)
    """
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import numpy as np

    ws = wb.create_sheet(title=sheet_title[:31])
    _, num_fmt = _fmt_metric(metric)
    is_pct    = _is_pct_metric(metric)
    is_amount = _is_amount_metric(metric)
    ylabel    = _ALL_METRIC_LABELS.get(metric, metric)

    if not isinstance(data_map, dict):
        return ws

    first_val  = next(iter(data_map.values()), None)
    multi_series = isinstance(first_val, dict)

    header_row = _CHART_DATA_OFFSET + 1
    data_start = header_row + 1

    # ── Data table ────────────────────────────────────────────────────────
    if multi_series:
        series_names = list(data_map.keys())
        ws.cell(row=header_row, column=1, value=x_label).font = XL_HDR_FONT
        ws.cell(row=header_row, column=1).fill = XL_HDR_FILL
        ws.cell(row=header_row, column=1).alignment = XL_HDR_ALIGN
        for ci, sname in enumerate(series_names, 2):
            c = ws.cell(row=header_row, column=ci, value=sname)
            c.font = XL_HDR_FONT; c.fill = XL_HDR_FILL; c.alignment = XL_HDR_ALIGN
        for ri, period in enumerate(periods, data_start):
            ws.cell(row=ri, column=1, value=str(period))
            for ci, sname in enumerate(series_names, 2):
                v = data_map[sname].get(period, 0)
                c = ws.cell(row=ri, column=ci, value=v)
                c.number_format = num_fmt
        num_series = len(series_names)
    else:
        series_names = None
        ws.cell(row=header_row, column=1, value=x_label).font = XL_HDR_FONT
        ws.cell(row=header_row, column=1).fill = XL_HDR_FILL
        ws.cell(row=header_row, column=1).alignment = XL_HDR_ALIGN
        c = ws.cell(row=header_row, column=2, value=_ALL_METRIC_LABELS.get(metric, metric))
        c.font = XL_HDR_FONT; c.fill = XL_HDR_FILL; c.alignment = XL_HDR_ALIGN
        for ri, period in enumerate(periods, data_start):
            ws.cell(row=ri, column=1, value=str(period))
            c = ws.cell(row=ri, column=2, value=data_map.get(period, 0))
            c.number_format = num_fmt
        num_series = 1

    ws.column_dimensions['A'].width = 20
    for ci in range(2, 2 + num_series):
        ws.column_dimensions[get_column_letter(ci)].width = 15

    if not periods:
        return ws

    # ── Matplotlib bar chart ──────────────────────────────────────────────
    n = len(periods)
    fig, ax = _chart_fig_setup(chart_title or sheet_title, ylabel, n)
    x_idx = np.arange(n)

    if multi_series:
        ns = len(series_names)
        bar_w = max(0.12, min(0.75 / ns, 0.3))
        offsets = np.linspace(-(ns-1)/2 * bar_w, (ns-1)/2 * bar_w, ns)
        for i, sname in enumerate(series_names):
            color  = _COMP_YEAR_PALETTE[i % len(_COMP_YEAR_PALETTE)]
            values = [data_map[sname].get(p, 0) for p in periods]
            ax.bar(x_idx + offsets[i], values, bar_w,
                   color=color, label=str(sname),
                   edgecolor='white', linewidth=0.4,
                   alpha=0.85, zorder=3)
        # Legend below chart
        ncol = min(ns, 6)
        ax.legend(loc='upper center', bbox_to_anchor=(0.5, -0.22),
                  ncol=ncol, fontsize=8.5, frameon=True,
                  fancybox=False, edgecolor='#dee2e6',
                  handlelength=1.2, borderpad=0.6)
        fig.tight_layout(rect=[0, 0.08, 1, 1])
    else:
        values = [data_map.get(p, 0) for p in periods]
        bar_w  = max(0.3, min(0.7, 12 / max(n, 1)))
        ax.bar(x_idx, values, bar_w,
               color=_BAR_SINGLE_COLOR, alpha=0.85,
               edgecolor='white', linewidth=0.3, zorder=3)
        fig.tight_layout()

    # X-axis
    ax.set_xticks(x_idx)
    ax.set_xticklabels([str(p) for p in periods], rotation=35, ha='right', fontsize=7.5)
    if n > 26:
        skip = max(2, n // 26)
        ax.set_xticks(x_idx[::skip])
        ax.set_xticklabels([str(periods[j]) for j in range(0, n, skip)],
                           rotation=35, ha='right', fontsize=7.5)
    ax.set_xlim(-0.7, n - 0.3)

    # Y-axis formatter
    ax.yaxis.set_major_formatter(
        plt.FuncFormatter(lambda v, _: _fmt_large(v, is_pct=is_pct, is_amount=is_amount))
    )

    img = _fig_to_xl_image(fig)
    _write_chart_image_to_sheet(ws, img)
    return ws


# ─────────────────────────────────────────────────────────────────────────────
# Sales Chart Excel Export
# ─────────────────────────────────────────────────────────────────────────────

def export_sales_chart_to_excel(
    data, date_from=None, date_to=None, shop_group=None,
    trend_xaxis='month', trend_metric='return_rate', trend_shops=None,
    bar_xaxis='month', bar_metric='total_customers', bar_shops=None,
    yoy_xaxis='month', yoy_metric='total_customers', yoy_shops=None,
):
    """
    Build Excel workbook for the Sales Analytics Chart page.

    Sheet 1 — Overview: summary metrics
    Sheet 2 — Shop Trends: LineChart (one series per shop), metric & xaxis from UI
    Sheet 3 — Period Totals: BarChart (total across selected shops per period)
    Sheet 4 — Year-over-Year: BarChart (one series per year, sum of selected shops)
    """
    wb = Workbook()

    # ── Overview sheet ────────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov["A1"] = "Sales Analytics Chart — Overview"
    ws_ov["A1"].font = XL_TITLE_FONT

    row = 2
    if date_from and date_to:
        ws_ov[f"A{row}"] = "Period:"
        ws_ov[f"B{row}"] = f"{date_from} → {date_to}"
        ws_ov[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1
    if shop_group:
        ws_ov[f"A{row}"] = "Shop Group:"
        ws_ov[f"B{row}"] = shop_group
        ws_ov[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1
    row += 1

    ov = data.get("overview", {})
    xl_write_header(ws_ov, ["Metric", "Value"], row=row)
    row += 1
    for label, value in [
        ("Active Customers (Period)", ov.get("active_customers", 0)),
        ("Returning Customers (Period)", ov.get("returning_customers", 0)),
        ("Return Rate (Period %)", ov.get("return_rate", 0)),
        ("Return Rate All-Time (%)", ov.get("return_rate_all_time", 0)),
        ("New Members in Period", ov.get("new_members_in_period", 0)),
        ("Total Customers in DB", ov.get("total_customers_in_db", 0)),
        ("Member Active All Time", ov.get("member_active_all_time", 0)),
        (None, None),
        ("Total Invoices (excl. VIP0)", ov.get("total_invoices_without_vip0", 0)),
        ("Total Amount (excl. VIP0)", ov.get("total_amount_without_vip0", ov.get("total_amount_period", 0))),
        ("Total Invoices (incl. VIP0)", ov.get("total_invoices_with_vip0", 0)),
        ("Total Amount (incl. VIP0)", ov.get("total_amount_with_vip0", 0)),
    ]:
        if label is None:
            row += 1
            continue
        ws_ov.cell(row=row, column=1, value=label)
        c = ws_ov.cell(row=row, column=2, value=value)
        if "Amount" in label:
            c.number_format = "#,##0"
        row += 1
    ws_ov.column_dimensions["A"].width = 38
    ws_ov.column_dimensions["B"].width = 20

    # ── Section 2: Shop Trends (Line chart) ───────────────────────────────────
    trend_periods, trend_shop_map = _build_sales_shop_map(data, trend_xaxis, trend_metric, trend_shops or [])
    metric_lbl = _SALES_METRIC_LABEL.get(trend_metric, trend_metric)
    _write_line_chart_sheet(
        wb, "Shop Trends", trend_periods, trend_shop_map, trend_metric,
        x_label=trend_xaxis.capitalize(),
        chart_title=f"Shop Trends — {metric_lbl} by {trend_xaxis.capitalize()}",
    )

    # ── Section 3: Period Totals (Bar chart) ──────────────────────────────────
    bar_periods, bar_shop_map = _build_sales_shop_map(data, bar_xaxis, bar_metric, bar_shops or [])
    _, bar_num_fmt = _fmt_metric(bar_metric)
    bar_totals = {p: sum(sm.get(p, 0) for sm in bar_shop_map.values()) for p in bar_periods}
    metric_lbl2 = _SALES_METRIC_LABEL.get(bar_metric, bar_metric)
    _write_bar_chart_sheet(
        wb, "Period Totals", bar_periods, bar_totals, bar_metric,
        x_label=bar_xaxis.capitalize(),
        chart_title=f"Period Totals — {metric_lbl2} by {bar_xaxis.capitalize()}",
    )

    # ── Section 4: Year-over-Year (Bar chart) ─────────────────────────────────
    yoy_periods, yoy_shop_map = _build_sales_shop_map(data, yoy_xaxis, yoy_metric, yoy_shops or [])
    within_periods, years, yoy_year_map = _build_yoy_data(yoy_shop_map, yoy_periods, yoy_xaxis, yoy_metric)
    metric_lbl3 = _SALES_METRIC_LABEL.get(yoy_metric, yoy_metric)
    _write_bar_chart_sheet(
        wb, "Year-over-Year", within_periods, yoy_year_map, yoy_metric,
        x_label=f"Period ({yoy_xaxis})",
        chart_title=f"Year-over-Year — {metric_lbl3}",
    )

    return wb


def export_customer_chart_to_excel(
    data, start_date="", end_date="",
    trend_xaxis='month', trend_metric='new_pos', trend_shops=None,
    bar_xaxis='month', bar_metric='new_pos', bar_shops=None,
    yoy_xaxis='month', yoy_metric='new_pos',
):
    """
    Build Excel workbook for the Customer Analytics Chart page (CNV).

    Sheet 1 — Overview: summary metrics
    Sheet 2 — Shop Trends: LineChart (one series per shop), metric & xaxis from UI
    Sheet 3 — Period Totals: BarChart (total across selected shops per period)
    Sheet 4 — Year-over-Year: BarChart (one series per year, all-time data)
    """
    wb = Workbook()

    # ── Overview sheet ────────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov["A1"] = "Customer Analytics Chart — Overview"
    ws_ov["A1"].font = XL_TITLE_FONT

    row = 2
    if start_date and end_date:
        ws_ov[f"A{row}"] = "Period:"
        ws_ov[f"B{row}"] = f"{start_date} → {end_date}"
        ws_ov[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1
    row += 1

    ov = data.get("overview", {})
    xl_write_header(ws_ov, ["Metric", "Value"], row=row)
    row += 1
    for label, value in [
        ("Total CNV Customers", ov.get("total_cnv", 0)),
        ("Total POS Customers", ov.get("total_pos", 0)),
        ("Active Zalo (Mini App)", ov.get("active_zalo", 0)),
        ("Follow OA", ov.get("follow_oa", 0)),
        ("CNV Only", ov.get("cnv_only", 0)),
        ("POS Only", ov.get("pos_only", 0)),
    ]:
        ws_ov.cell(row=row, column=1, value=label)
        ws_ov.cell(row=row, column=2, value=value)
        row += 1
    ws_ov.column_dimensions["A"].width = 35
    ws_ov.column_dimensions["B"].width = 20

    # Helper: build {shop_name: {period: value}} from shop_stats
    # shop_stats can be list [{shop_name, by_month, ...}] or dict {name: {...}}
    _shop_stats_raw = data.get("shop_stats", [])
    if isinstance(_shop_stats_raw, list):
        shop_stats = {s['shop_name']: s for s in _shop_stats_raw}
    else:
        shop_stats = _shop_stats_raw

    def _cust_shop_map(xaxis, metric, selected_shops):
        axis_key, period_key = _CUSTOMER_AXIS.get(xaxis, ('month', 'month'))
        result = {}
        for sname, sdata in shop_stats.items():
            if selected_shops and sname not in selected_shops:
                continue
            period_list = sdata.get(axis_key, [])
            def _v(e, _m=metric):
                v = e.get(_m, 0) or 0
                return v / 100 if _is_pct_metric(_m) and v > 1 else v
            result[sname] = {e[period_key]: _v(e) for e in period_list}
        return result

    def _cust_periods(xaxis, selected_shops):
        axis_key, period_key = _CUSTOMER_AXIS.get(xaxis, ('month', 'month'))
        # Get from first matching shop
        for sname, sdata in shop_stats.items():
            if selected_shops and sname not in selected_shops:
                continue
            return [e[period_key] for e in sdata.get(axis_key, [])]
        # Fallback from aggregate stats key
        agg_key = {'season': 'season_stats', 'month': 'month_stats', 'week': 'week_stats'}.get(xaxis, 'month_stats')
        pk = {'season': 'label', 'month': 'month', 'week': 'week'}.get(xaxis, 'month')
        return [e[pk] for e in data.get(agg_key, [])]

    # ── Section 2: Shop Trends (Line chart) ───────────────────────────────────
    t_shops_sel = trend_shops or []
    trend_periods = _cust_periods(trend_xaxis, t_shops_sel)
    trend_shop_map = _cust_shop_map(trend_xaxis, trend_metric, t_shops_sel)
    metric_lbl = _CUSTOMER_METRIC_LABEL.get(trend_metric, trend_metric)
    _write_line_chart_sheet(
        wb, "Shop Trends", trend_periods, trend_shop_map, trend_metric,
        x_label=trend_xaxis.capitalize(),
        chart_title=f"Shop Trends — {metric_lbl} by {trend_xaxis.capitalize()}",
    )

    # ── Section 3: Period Totals (Bar chart) ──────────────────────────────────
    b_shops_sel = bar_shops or []
    bar_periods = _cust_periods(bar_xaxis, b_shops_sel)
    bar_shop_map = _cust_shop_map(bar_xaxis, bar_metric, b_shops_sel)
    bar_totals = {p: sum(sm.get(p, 0) for sm in bar_shop_map.values()) for p in bar_periods}
    metric_lbl2 = _CUSTOMER_METRIC_LABEL.get(bar_metric, bar_metric)
    _write_bar_chart_sheet(
        wb, "Period Totals", bar_periods, bar_totals, bar_metric,
        x_label=bar_xaxis.capitalize(),
        chart_title=f"Period Totals — {metric_lbl2} by {bar_xaxis.capitalize()}",
    )

    # ── Section 4: Year-over-Year (Bar chart, all-time data) ──────────────────
    # Use all_month_stats / all_season_stats / all_week_stats (all-time, no period filter)
    yoy_all_key = {'season': 'all_season_stats', 'month': 'all_month_stats', 'week': 'all_week_stats'}.get(yoy_xaxis, 'all_month_stats')
    yoy_pk = {'season': 'label', 'month': 'month', 'week': 'week'}.get(yoy_xaxis, 'month')
    yoy_all_periods = [e[yoy_pk] for e in data.get(yoy_all_key, [])]

    def _v_yoy(e):
        v = e.get(yoy_metric, 0) or 0
        return v / 100 if _is_pct_metric(yoy_metric) and v > 1 else v

    yoy_agg_map = {e[yoy_pk]: _v_yoy(e) for e in data.get(yoy_all_key, [])}
    # Group by year
    yoy_within_periods, yoy_years, yoy_year_map = _build_yoy_data(
        {'__agg__': yoy_agg_map}, yoy_all_periods, yoy_xaxis, yoy_metric
    )
    metric_lbl3 = _CUSTOMER_METRIC_LABEL.get(yoy_metric, yoy_metric)
    _write_bar_chart_sheet(
        wb, "Year-over-Year", yoy_within_periods, yoy_year_map, yoy_metric,
        x_label=f"Period ({yoy_xaxis})",
        chart_title=f"Year-over-Year — {metric_lbl3} (All Time)",
    )

    return wb


def export_coupon_chart_to_excel(
    summary, trend, date_from=None, date_to=None,
    coupon_id_prefix=None, shop_group=None,
    shop_xaxis='month', shop_metric='used', shop_shops=None,
    camp_xaxis='month', camp_metric='used', camp_campaigns=None,
):
    """
    Build Excel workbook for the Coupon Analytics Chart page.

    Sheet 1 — Overview: all_time + period summary metrics
    Sheet 2 — Shop Trends: LineChart (one series per shop), metric & xaxis from UI
    Sheet 3 — Campaign Trends: LineChart (one series per campaign)

    summary: result of get_coupon_summary() → {all_time: {...}, period: {...}}
    trend:   result of calculate_coupon_trend_data() → {time_labels, shops, shop_series, camp_series, ...}
    """
    wb = Workbook()

    # ── Overview sheet ────────────────────────────────────────────────────────
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov["A1"] = "Coupon Analytics Chart — Overview"
    ws_ov["A1"].font = XL_TITLE_FONT

    row = 2
    if date_from and date_to:
        ws_ov[f"A{row}"] = "Period:"
        ws_ov[f"B{row}"] = f"{date_from} → {date_to}"
        ws_ov[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1
    if coupon_id_prefix:
        ws_ov[f"A{row}"] = "Coupon Prefix:"
        ws_ov[f"B{row}"] = coupon_id_prefix
        ws_ov[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1
    if shop_group:
        ws_ov[f"A{row}"] = "Shop Group:"
        ws_ov[f"B{row}"] = shop_group
        ws_ov[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1
    row += 1

    xl_write_header(ws_ov, ["Metric", "Value"], row=row)
    row += 1
    for section_label, section_key in [("All-Time", "all_time"), ("Period", "period")]:
        sec = summary.get(section_key, {})
        ws_ov.cell(row=row, column=1, value=f"── {section_label} ──")
        ws_ov.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for label, key, fmt in [
            ("Total Coupons", "total", "0"),
            ("Used Count", "used_count", "0"),
            ("Unused", "unused", "0"),
            ("Usage Rate (%)", "usage_rate", "0.00"),
            ("Total Coupon Amount", "total_coupon_amount", "#,##0"),
            ("Total Invoice Amount", "total_invoice_amount", "#,##0"),
        ]:
            ws_ov.cell(row=row, column=1, value=label)
            c = ws_ov.cell(row=row, column=2, value=sec.get(key, 0))
            c.number_format = fmt
            row += 1
        row += 1
    ws_ov.column_dimensions["A"].width = 35
    ws_ov.column_dimensions["B"].width = 20

    # ── Helper: build {entity_name: {period: value}} from trend series ────────
    def _build_entity_map(series_dict, xaxis, metric, selected_entities,
                          time_labels_key="time_labels"):
        """
        series_dict: {entity_name: {period_key: {metric: value, ...}}}
        time_labels_key: 'time_labels' (shop chart) or 'time_labels_camp' (campaign chart)
        Returns (periods_list, {entity_name: {period: value}})

        For weekly xaxis: time_labels contains week_sort keys (e.g. '2025-W01').
        We translate these to display labels using week_label_map before returning.
        """
        axis_key    = _COUPON_AXIS_LABEL.get(xaxis, 'month')
        entity_series = series_dict.get(axis_key, {})
        raw_labels  = trend.get(time_labels_key, {}).get(axis_key, [])

        # Translate week_sort → week_label for display
        week_lbl_map = trend.get("week_label_map", {}) if xaxis == 'week' else {}
        display_labels = [week_lbl_map.get(str(p), str(p)) for p in raw_labels]

        result = {}
        for ename, period_map in entity_series.items():
            if selected_entities and ename not in selected_entities:
                continue
            entry = {}
            for raw_pk, disp_pk in zip(raw_labels, display_labels):
                raw_str  = str(raw_pk)
                row_data = period_map.get(raw_str, period_map.get(raw_pk, {}))
                v = row_data.get(metric, 0) or 0
                if _is_pct_metric(metric) and v > 1:
                    v /= 100
                entry[disp_pk] = v
            result[ename] = entry

        return display_labels, result

    # ── Section 2: Shop Trends (Line chart) ───────────────────────────────────
    shop_periods, shop_entity_map = _build_entity_map(
        trend.get("shop_series", {}), shop_xaxis, shop_metric, shop_shops or [],
        time_labels_key="time_labels",
    )
    metric_lbl = _COUPON_METRIC_LABEL.get(shop_metric, shop_metric)
    _write_line_chart_sheet(
        wb, "Shop Trends", shop_periods, shop_entity_map, shop_metric,
        x_label=shop_xaxis.capitalize(),
        chart_title=f"Shop Trends — {metric_lbl} by {shop_xaxis.capitalize()}",
    )

    # ── Section 3: Campaign Trends (Line chart) ───────────────────────────────
    # Key is 'campaign_series' (not 'camp_series') in calculate_coupon_trend_data().
    # Campaign chart uses its own time_labels_camp (all-time, unfiltered).
    camp_periods, camp_entity_map = _build_entity_map(
        trend.get("campaign_series", {}), camp_xaxis, camp_metric, camp_campaigns or [],
        time_labels_key="time_labels_camp",
    )
    metric_lbl2 = _COUPON_METRIC_LABEL.get(camp_metric, camp_metric)
    _write_line_chart_sheet(
        wb, "Campaign Trends", camp_periods, camp_entity_map, camp_metric,
        x_label=camp_xaxis.capitalize(),
        chart_title=f"Campaign Trends — {metric_lbl2} by {camp_xaxis.capitalize()}",
    )

    return wb


# ── Product Analytics Excel Export ───────────────────────────────────────────

def export_product_analytics_to_excel(tabs_data: dict, date_from=None, date_to=None, shop_group=None):
    """Export Sales & Product Analytics tabs to Excel workbook."""
    from App.analytics.category_translations import translate_category
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    # Overview sheet
    overview = (tabs_data.get('month') or tabs_data.get('brand') or {}).get('overview', {})
    ws.append(["Sales & Product Analytics Export"])
    ws['A1'].font = XL_TITLE_FONT
    ws.append([f"Period: {date_from} — {date_to}" if date_from else "Period: All Time"])
    ws.append([f"Shop Group: {shop_group or 'All'}"])
    ws.append([])
    if overview:
        ws.append(["KPI", "Value"])
        xl_write_header(ws, ["KPI", "Value"], row=ws.max_row)
        for k, v in [
            ("Total Lines", overview.get('total_lines', 0)),
            ("Total Qty (pcs)", overview.get('total_qty', 0)),
            ("Tag Amount (VND)", overview.get('total_tag_amount', 0)),
            ("Sales Amount (VND)", overview.get('total_amount', 0)),
            ("Settlement (VND)", overview.get('total_settlement', 0)),
            ("Avg Disc %", f"{overview.get('disc_pct', '—')}%"),
        ]:
            ws.append([k, v])
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20

    _HEADERS = ["Period/Group", "Qty (pcs)", "Tag Amt (VND)", "Sales Amt (VND)",
                "Settlement (VND)", "Disc %", "Lines"]
    _CAT_HEADERS = ["", "Category L1", "Category L2", "Category L3", "Qty (pcs)",
                    "Tag Amt (VND)", "Sales Amt (VND)", "Settlement (VND)", "Disc %", "Lines"]

    def _write_period_tab(ws_t, rows, period_key, label_key='label'):
        xl_write_header(ws_t, _HEADERS, row=1)
        for r in rows:
            ws_t.append([
                r.get(label_key) or r.get(period_key, '—'),
                r.get('qty', 0),
                round(float(r.get('tag_amount') or 0)),
                round(float(r.get('amount') or 0)),
                round(float(r.get('settlement') or 0)),
                f"{r.get('disc_pct', '—')}%" if r.get('disc_pct') is not None else '—',
                r.get('lines', 0),
            ])
            # cat_groups: L1 → l2_groups → L3 rows
            for grp in r.get('cat_groups', []):
                for l2g in grp.get('l2_groups', []):
                    for cat in l2g.get('rows', []):
                        ws_t.append([
                            '',
                            translate_category(cat.get('category_l1', ''), 'VI'),
                            translate_category(cat.get('category_l2', ''), 'VI'),
                            translate_category(cat.get('category_l3', ''), 'VI'),
                            cat.get('qty', 0),
                            round(float(cat.get('tag_amount') or 0)),
                            round(float(cat.get('amount') or 0)),
                            round(float(cat.get('settlement') or 0)),
                            f"{cat.get('disc_pct', '—')}%" if cat.get('disc_pct') is not None else '—',
                            cat.get('lines', 0),
                        ])
        for col in range(1, 9):
            ws_t.column_dimensions[get_column_letter(col)].width = 18

    # By Month sheet
    if 'month' in tabs_data:
        ws_m = wb.create_sheet("By Month")
        _write_period_tab(ws_m, tabs_data['month'].get('by_month', []), 'month_trunc')

    # By Brand sheet
    if 'brand' in tabs_data:
        ws_b = wb.create_sheet("By Brand")
        _write_period_tab(ws_b, tabs_data['brand'].get('by_brand', []), 'brand', label_key='brand')

    # By Category sheet — Brand → L1 → L2 → L3
    if 'category' in tabs_data:
        ws_c = wb.create_sheet("By Category")
        headers_c = ["Brand", "Category L1", "Category L2", "Category L3", "Qty (pcs)",
                     "Tag Amt (VND)", "Sales Amt (VND)", "Settlement (VND)", "Disc %", "Lines"]
        xl_write_header(ws_c, headers_c, row=1)
        for br_grp in tabs_data['category'].get('by_category', []):
            brand = br_grp.get('brand', '')
            for grp in br_grp.get('l1_groups', []):
                l1 = translate_category(grp.get('l1', ''), 'VI')
                for l2g in grp.get('l2_groups', []):
                    l2 = translate_category(l2g.get('l2', ''), 'VI')
                    for r in l2g.get('rows', []):
                        ws_c.append([
                            brand,
                            l1,
                            l2,
                            translate_category(r.get('category_l3', ''), 'VI'),
                            r.get('qty', 0),
                            round(float(r.get('tag_amount') or 0)),
                            round(float(r.get('amount') or 0)),
                            round(float(r.get('settlement') or 0)),
                            f"{r.get('disc_pct', '—')}%" if r.get('disc_pct') is not None else '—',
                            r.get('lines', 0),
                        ])
        for col, w in zip(range(1, 11), [16, 22, 26, 22, 10, 16, 16, 16, 8, 8]):
            ws_c.column_dimensions[get_column_letter(col)].width = w

    # By Shop sheet
    if 'shop' in tabs_data:
        ws_s = wb.create_sheet("By Shop")
        sh_headers = ["Shop", "Qty (pcs)", "Tag Amt (VND)", "Sales Amt (VND)",
                      "Settlement (VND)", "Disc %", "Lines"]
        xl_write_header(ws_s, sh_headers, row=1)
        for r in tabs_data['shop'].get('by_shop', []):
            ws_s.append([
                r.get('shop_name', '—'),
                r.get('qty', 0),
                round(float(r.get('tag_amount') or 0)),
                round(float(r.get('amount') or 0)),
                round(float(r.get('settlement') or 0)),
                f"{r.get('disc_pct', '—')}%" if r.get('disc_pct') is not None else '—',
                r.get('lines', 0),
            ])
        for col in range(1, 8):
            ws_s.column_dimensions[get_column_letter(col)].width = 22

    # By Year sheet
    if 'year' in tabs_data:
        ws_y = wb.create_sheet("By Year")
        _write_period_tab(ws_y, tabs_data['year'].get('by_year', []), 'year_trunc')

    # By Week sheet
    if 'week' in tabs_data:
        ws_w = wb.create_sheet("By Week")
        _write_period_tab(ws_w, tabs_data['week'].get('by_week', []), 'week_label')

    # By Sales Season sheet
    if 'sales_season' in tabs_data:
        ws_ss = wb.create_sheet("By Sales Season")
        _write_period_tab(ws_ss, tabs_data['sales_season'].get('by_sales_season', []), 'label')

    # By Product Season sheet
    if 'product_season' in tabs_data:
        ws_ps = wb.create_sheet("By Product Season")
        _write_period_tab(ws_ps, tabs_data['product_season'].get('by_product_season', []), 'label')

    # By VIP Grade sheet
    if 'vip_grade' in tabs_data:
        ws_vg = wb.create_sheet("By VIP Grade")
        _write_period_tab(ws_vg, tabs_data['vip_grade'].get('by_vip_grade', []), 'vip_grade', label_key='grade')

    # Top Products sheet
    if 'product' in tabs_data:
        ws_p = wb.create_sheet("Top Products")
        p_headers = ["#", "Product Code", "Product Name", "Brand",
                     "Cat L1", "Cat L2", "Cat L3",
                     "Year", "Season", "Qty (pcs)", "Tag Amt (VND)",
                     "Sales Amt (VND)", "Settlement (VND)", "Disc %"]
        xl_write_header(ws_p, p_headers, row=1)
        for i, r in enumerate(tabs_data['product'].get('by_product', []), 1):
            ws_p.append([
                i,
                r.get('product_code', ''),
                r.get('product_name', ''),
                r.get('brand', ''),
                translate_category(r.get('category_l1', ''), 'VI'),
                translate_category(r.get('category_l2', ''), 'VI'),
                translate_category(r.get('category_l3', ''), 'VI'),
                r.get('year', ''),
                r.get('season', ''),
                r.get('qty', 0),
                round(float(r.get('tag_amount') or 0)),
                round(float(r.get('amount') or 0)),
                round(float(r.get('settlement') or 0)),
                f"{r.get('disc_pct', '—')}%" if r.get('disc_pct') is not None else '—',
            ])
        for col, w in zip(range(1, 15), [4, 16, 32, 14, 20, 24, 20, 8, 8, 10, 14, 14, 14, 8]):
            ws_p.column_dimensions[get_column_letter(col)].width = w

    return wb