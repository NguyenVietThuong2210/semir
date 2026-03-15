"""
App/analytics/coupon_analytics.py

Coupon usage analytics and Excel export.
Separate domain from customer analytics.

Version: 3.4 - FIXED: Total amount now uses invoice amounts, not face_value
All field names verified against Coupon model
"""

from collections import defaultdict
from decimal import Decimal

from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from App.models import Coupon, Customer, SalesTransaction
from App.models_cnv import CNVCustomer


def calc_coupon_amount(face_value, invoice_amount):
    """
    Calculate the actual discount value of a coupon given the invoice amount.

    face_value > 1  → cash VND  → coupon_amount = face_value
    0 < face_value <= 1 → percentage (e.g. 0.9 = 90% pay → 10% discount)
        → coupon_amount = (discount_pct / face_value_pct) * invoice_amount
        → special case face_value = 1.0 (100% pay) → coupon_amount = 0
    """
    from decimal import Decimal

    if not face_value or face_value <= 0:
        return Decimal(0)
    if face_value > 1:
        return Decimal(str(face_value))
    if face_value >= 1:
        return Decimal(0)
    discount = Decimal(1) - Decimal(str(face_value))
    base = Decimal(str(face_value))
    return (discount / base) * Decimal(str(invoice_amount or 0))


def format_face_value(face_value):
    """
    Format face value for display.

    Rules:
    - If face_value > 1: Display as VND amount (e.g., "50,000 VND")
    - If 0 < face_value <= 1: Display as percentage (e.g., 0.95 → "95%")
    - If face_value = 0 or None: Display as "0"

    Args:
        face_value: Decimal or float value

    Returns:
        Formatted string
    """
    if not face_value or face_value == 0:
        return "0"

    if face_value > 1:
        # VND amount
        return f"{face_value:,.0f} VND"
    else:
        # Percentage (0.95 → 95%)
        percentage = face_value * 100
        return f"{percentage:.0f}%"


def calculate_coupon_analytics(
    date_from=None, date_to=None, coupon_id_prefix=None, shop_group=None
):
    """
    Calculate coupon usage analytics.

    IMPORTANT: Total amounts are calculated from INVOICE AMOUNTS, not face_value.

    Args:
        date_from: Start date for period filter
        date_to: End date for period filter
        coupon_id_prefix: Filter by coupon ID prefix (case-insensitive)
        shop_group: Filter by shop group (Bala Group, Semir Group, Others Group)

    Returns:
        Dict with all_time, period, by_shop, and details data
    """
    qs = Coupon.objects.all()

    # Apply shop group filter using icontains
    if shop_group:
        if shop_group == "Bala Group":
            # Filter shops containing "Bala" or "巴拉"
            qs = qs.filter(
                Q(using_shop__icontains="Bala") | Q(using_shop__icontains="巴拉")
            )
        elif shop_group == "Semir Group":
            # Filter shops containing "Semir" or "森马"
            qs = qs.filter(
                Q(using_shop__icontains="Semir") | Q(using_shop__icontains="森马")
            )
        elif shop_group == "Others Group":
            # Exclude Bala and Semir shops
            qs = qs.exclude(
                Q(using_shop__icontains="Bala")
                | Q(using_shop__icontains="巴拉")
                | Q(using_shop__icontains="Semir")
                | Q(using_shop__icontains="森马")
            )

    # Filter by usage date if specified
    if date_from or date_to:
        usage_filter = Q(using_date__isnull=False)
        if date_from:
            usage_filter &= Q(using_date__gte=date_from)
        if date_to:
            usage_filter &= Q(using_date__lte=date_to)
        period_qs = qs.filter(usage_filter)
    else:
        period_qs = qs

    # Filter by coupon ID prefix if specified (supports comma-separated multi-prefix OR logic)
    if coupon_id_prefix:
        _prefixes = [p.strip() for p in coupon_id_prefix.split(",") if p.strip()]
        if len(_prefixes) == 1:
            qs = qs.filter(coupon_id__istartswith=_prefixes[0])
            period_qs = period_qs.filter(coupon_id__istartswith=_prefixes[0])
        elif _prefixes:
            _pq = Q()
            for _p in _prefixes:
                _pq |= Q(coupon_id__istartswith=_p)
            qs = qs.filter(_pq)
            period_qs = period_qs.filter(_pq)

    # All-time stats - counts only (amounts calculated below)
    all_time_total = qs.count()
    all_time_used = qs.filter(using_date__isnull=False).count()
    all_time_unused = all_time_total - all_time_used
    all_time_usage_rate = round(
        all_time_used / all_time_total * 100 if all_time_total else 0, 2
    )

    # Period stats - counts only (amounts calculated below)
    period_total = period_qs.count()
    period_used = period_qs.filter(using_date__isnull=False).count()
    period_unused = period_total - period_used
    period_usage_rate = round(
        period_used / period_total * 100 if period_total else 0, 2
    )

    # ===========================================================================
    # CALCULATE AMOUNTS FROM INVOICE AMOUNTS (NOT face_value)
    # ===========================================================================

    all_time_amount = Decimal(0)
    all_time_coupon_amount = Decimal(0)
    all_time_unique_amount = Decimal(0)
    period_amount = Decimal(0)
    period_coupon_amount = Decimal(0)
    period_unique_amount = Decimal(0)
    shop_data = {}
    coupon_details = []

    # Helper function to get invoice amount for a coupon
    def get_invoice_amount(coupon):
        """Get invoice amount for a coupon, fallback to face_value if no invoice found."""
        if coupon.docket_number:
            try:
                txn = SalesTransaction.objects.get(invoice_number=coupon.docket_number)
                return txn.sales_amount or coupon.face_value or Decimal(0)
            except SalesTransaction.DoesNotExist:
                pass
        return coupon.face_value or Decimal(0)

    # ===========================================================================
    # DUPLICATE INVOICE DETECTION
    # ===========================================================================
    from django.db.models import Count as _Count

    # All-time duplicates
    _dup_dockets_all = set(
        Coupon.objects.filter(using_date__isnull=False, docket_number__isnull=False)
        .exclude(docket_number="")
        .values("docket_number")
        .annotate(_c=_Count("id"))
        .filter(_c__gt=1)
        .values_list("docket_number", flat=True)
    )

    # Period duplicates (within filtered qs)
    _dup_dockets_period = set(
        period_qs.filter(using_date__isnull=False, docket_number__isnull=False)
        .exclude(docket_number="")
        .values("docket_number")
        .annotate(_c=_Count("id"))
        .filter(_c__gt=1)
        .values_list("docket_number", flat=True)
    )

    # Build duplicate_invoices list (all coupons sharing a docket_number, period scope)
    duplicate_invoices = []
    if _dup_dockets_period:
        for docket in sorted(_dup_dockets_period):
            coupons_for_docket = list(
                period_qs.filter(
                    docket_number=docket, using_date__isnull=False
                ).order_by("coupon_id")
            )
            try:
                txn = SalesTransaction.objects.get(invoice_number=docket)
                inv_amount = txn.sales_amount or Decimal(0)
                shop_name = txn.shop_name or ""
                sales_date = txn.sales_date
            except SalesTransaction.DoesNotExist:
                inv_amount = Decimal(0)
                shop_name = ""
                sales_date = None
            for c in coupons_for_docket:
                duplicate_invoices.append(
                    {
                        "docket_number": docket,
                        "coupon_id": c.coupon_id,
                        "face_value_display": format_face_value(c.face_value),
                        "coupon_amount": float(
                            calc_coupon_amount(c.face_value, inv_amount)
                        ),
                        "using_date": c.using_date,
                        "using_shop": c.using_shop or "",
                        "inv_amount": float(inv_amount),
                        "shop_name": shop_name,
                        "sales_date": sales_date,
                        "member_id": c.member_id or "",
                        "member_name": c.member_name or "",
                        "member_phone": c.member_phone or "",
                    }
                )

    # Calculate all-time amount (process ALL used coupons)
    _seen_dockets_all = set()
    for coupon in qs.filter(using_date__isnull=False):
        inv_amount = get_invoice_amount(coupon)
        all_time_amount += inv_amount
        all_time_coupon_amount += calc_coupon_amount(coupon.face_value, inv_amount)
        # Unique: count invoice amount only once per docket
        dk = coupon.docket_number or f"__no_docket_{coupon.pk}"
        if dk not in _seen_dockets_all:
            _seen_dockets_all.add(dk)
            all_time_unique_amount += inv_amount

    # ===========================================================================
    # PROCESS PERIOD COUPONS - Build details and calculate period amounts
    # ===========================================================================
    _seen_dockets_period = set()

    for coupon in period_qs.filter(using_date__isnull=False).order_by("-using_date"):
        # Get customer info from coupon
        vip_id = coupon.member_id or None
        vip_name = coupon.member_name or None
        phone = coupon.member_phone or None

        # Initialize transaction fields
        sales_date = None
        inv_shop = None
        inv_amount = None
        note = None

        # Try to get invoice/transaction info
        if coupon.docket_number:
            try:
                txn = SalesTransaction.objects.get(invoice_number=coupon.docket_number)

                # Get customer info from transaction if not in coupon
                if not vip_id:
                    vip_id = txn.vip_id
                if not vip_name and txn.vip_id and txn.vip_id != "0":
                    try:
                        cust = Customer.objects.get(vip_id=txn.vip_id)
                        vip_name = cust.name
                        phone = cust.phone
                    except Customer.DoesNotExist:
                        vip_name = txn.vip_name

                # Get transaction details
                sales_date = txn.sales_date
                inv_shop = txn.shop_name
                inv_amount = txn.sales_amount

                # Validate shop match
                if coupon.using_shop and inv_shop and coupon.using_shop != inv_shop:
                    note = f"Shop mismatch: Coupon@{coupon.using_shop} vs Invoice@{inv_shop}"

            except SalesTransaction.DoesNotExist:
                note = f"Invoice {coupon.docket_number} not found"

        # Final amount (invoice amount or fallback to face_value)
        final_amount = inv_amount or coupon.face_value or Decimal(0)
        coupon_amt = calc_coupon_amount(coupon.face_value, final_amount)

        # Accumulate period amount
        period_amount += final_amount
        period_coupon_amount += coupon_amt

        # Unique invoice tracking
        dk = coupon.docket_number or f"__no_docket_{coupon.pk}"
        is_duplicate = dk in _dup_dockets_period and bool(coupon.docket_number)
        if dk not in _seen_dockets_period:
            _seen_dockets_period = _seen_dockets_period | {dk}
            period_unique_amount += final_amount

        # Accumulate by shop
        shop = coupon.using_shop or "Unknown"
        if shop not in shop_data:
            shop_data[shop] = {
                "total": 0,
                "used": 0,
                "amount": Decimal(0),
                "coupon_amount": Decimal(0),
                "unique_amount": Decimal(0),
                "seen_dockets": set(),
            }
        shop_data[shop]["used"] += 1
        shop_data[shop]["amount"] += final_amount
        shop_data[shop]["coupon_amount"] += coupon_amt
        # Unique per shop: count invoice amount once per docket per shop
        if dk not in shop_data[shop]["seen_dockets"]:
            shop_data[shop]["seen_dockets"].add(dk)
            shop_data[shop]["unique_amount"] += final_amount

        # Build detail record
        coupon_details.append(
            {
                "coupon_id": coupon.coupon_id,
                "creator": coupon.creator or "",
                "face_value": coupon.face_value or 0,
                "face_value_display": format_face_value(coupon.face_value),
                "using_shop": coupon.using_shop or "Unknown",
                "using_date": coupon.using_date,
                "docket_number": coupon.docket_number or "",
                "vip_id": vip_id or "",
                "customer_name": vip_name or "-",
                "customer_phone": phone or "-",
                "sales_day": sales_date,
                "inv_shop": inv_shop or "-",
                "amount": float(final_amount),
                "coupon_amount": float(coupon_amt),
                "is_duplicate": is_duplicate,
                "note": note or "",
            }
        )

    # Add unused coupons to shop totals
    for coupon in period_qs.filter(using_date__isnull=True):
        shop = coupon.using_shop or "Unknown"
        if shop not in shop_data:
            shop_data[shop] = {
                "total": 0,
                "used": 0,
                "amount": Decimal(0),
                "coupon_amount": Decimal(0),
                "unique_amount": Decimal(0),
                "seen_dockets": set(),
            }
        shop_data[shop]["total"] += 1

    # Update shop totals (add used to total for each shop)
    for shop_name in shop_data:
        shop_data[shop_name]["total"] += shop_data[shop_name]["used"]

    # ===========================================================================
    # BUILD SHOP STATS
    # ===========================================================================

    total_used_all_shops = period_used

    shop_stats = []
    for shop_name, data in shop_data.items():
        total = data["total"]
        used = data["used"]
        pct_of_used = round(
            used / total_used_all_shops * 100 if total_used_all_shops else 0, 2
        )
        usage_rate = round(used / total * 100 if total else 0, 2)

        shop_stats.append(
            {
                "shop_name": shop_name,
                "total": total,
                "used": used,
                "unused": total - used,
                "used_pct_of_used": pct_of_used,
                "usage_rate": usage_rate,
                "total_amount": float(data["unique_amount"]),  # unique invoice amount
                "coupon_amount": float(data["coupon_amount"]),
            }
        )

    shop_stats.sort(key=lambda x: x["used"], reverse=True)

    # Calculate percentages for template
    all_time_used_pct = (
        round(all_time_used / all_time_total * 100, 2) if all_time_total else 0
    )
    all_time_unused_pct = (
        round(all_time_unused / all_time_total * 100, 2) if all_time_total else 0
    )

    period_used_pct = round(period_used / period_total * 100, 2) if period_total else 0
    period_unused_pct = (
        round(period_unused / period_total * 100, 2) if period_total else 0
    )

    # ===========================================================================
    # ENRICH DETAILS WITH CNV DATA (match by phone)
    # ===========================================================================
    phones = {
        d["customer_phone"]
        for d in coupon_details
        if d["customer_phone"] and d["customer_phone"] != "-"
    }
    cnv_map = {
        c["phone"]: c
        for c in CNVCustomer.objects.filter(phone__in=phones).values(
            "phone", "cnv_id", "points", "total_points"
        )
    }
    for d in coupon_details:
        cnv = cnv_map.get(d["customer_phone"])
        if cnv:
            d["cnv_id"] = cnv["cnv_id"]
            d["cnv_points"] = cnv["points"]
            d["cnv_total_points"] = cnv["total_points"]
        else:
            d["cnv_id"] = ""
            d["cnv_points"] = ""
            d["cnv_total_points"] = ""

    return {
        "all_time": {
            "total": all_time_total,
            "used": all_time_used,
            "unused": all_time_unused,
            "used_pct": all_time_used_pct,
            "unused_pct": all_time_unused_pct,
            "usage_rate": all_time_usage_rate,
            "total_amount": float(all_time_amount),
            "total_coupon_amount": float(all_time_coupon_amount),
            "unique_invoice_amount": float(all_time_unique_amount),
            "duplicate_invoice_count": len(_dup_dockets_all),
        },
        "period": {
            "total": period_total,
            "used": period_used,
            "unused": period_unused,
            "used_pct": period_used_pct,
            "unused_pct": period_unused_pct,
            "usage_rate": period_usage_rate,
            "total_amount": float(period_amount),
            "total_coupon_amount": float(period_coupon_amount),
            "unique_invoice_amount": float(period_unique_amount),
            "duplicate_invoice_count": len(_dup_dockets_period),
        },
        "by_shop": shop_stats,
        "details": coupon_details,
        "duplicate_invoices": duplicate_invoices,
    }


def calculate_coupon_trend_data(date_from=None, date_to=None, shop_group=None, coupon_id_prefix=None):
    """
    Compute time-series trend data for the coupon chart page.

    Groups used coupons (filtered by prefix if given) by shop×time and campaign×time.
    Returns serialisable plain dicts — safe to cache in Redis.

    Bucket types: week (YYYY-Www), month (YYYY-MM), season (M2-4 2025 …), year (YYYY)
    """
    from collections import defaultdict
    from App.analytics.season_utils import (
        get_session_key, get_week_info,
        week_sort_key, session_sort_key,
    )
    from App.models import CouponCampaign

    # ── Base queryset: date + shop_group only (used for campaign series) ─────
    qs_base = Coupon.objects.all()
    if shop_group:
        if shop_group == "Bala Group":
            qs_base = qs_base.filter(
                Q(using_shop__icontains="Bala") | Q(using_shop__icontains="巴拉")
            )
        elif shop_group == "Semir Group":
            qs_base = qs_base.filter(
                Q(using_shop__icontains="Semir") | Q(using_shop__icontains="森马")
            )
        elif shop_group == "Others Group":
            qs_base = qs_base.exclude(
                Q(using_shop__icontains="Bala")
                | Q(using_shop__icontains="巴拉")
                | Q(using_shop__icontains="Semir")
                | Q(using_shop__icontains="森马")
            )

    usage_filter = Q(using_date__isnull=False)
    if date_from:
        usage_filter &= Q(using_date__gte=date_from)
    if date_to:
        usage_filter &= Q(using_date__lte=date_to)
    period_qs_all = qs_base.filter(usage_filter)  # for campaign series (no prefix)

    # ── Shop queryset: additionally filtered by prefix (chart 1 only) ────────
    qs_shop = qs_base
    if coupon_id_prefix:
        _prefixes = [p.strip() for p in coupon_id_prefix.split(",") if p.strip()]
        if len(_prefixes) == 1:
            qs_shop = qs_shop.filter(coupon_id__istartswith=_prefixes[0])
        elif _prefixes:
            _pq = Q()
            for _p in _prefixes:
                _pq |= Q(coupon_id__istartswith=_p)
            qs_shop = qs_shop.filter(_pq)
    period_qs_shop = qs_shop.filter(usage_filter)  # for shop series (with prefix)

    # ── Bulk fetch coupon rows ────────────────────────────────────────────────
    _fields = ("coupon_id", "using_shop", "using_date", "face_value", "docket_number")
    coupon_rows_all = list(period_qs_all.values(*_fields))   # campaign data
    coupon_rows_shop = list(period_qs_shop.values(*_fields))  # shop data

    if not coupon_rows_all and not coupon_rows_shop:
        return {
            "time_labels": {"week": [], "month": [], "season": [], "year": []},
            "week_label_map": {},
            "total_by_time": {"week": {}, "month": {}, "season": {}, "year": {}},
            "total_by_time_shop": {"week": {}, "month": {}, "season": {}, "year": {}},
            "shops": [],
            "shop_series": {},
            "campaigns": [],
            "campaign_series": {},
            "campaign_list": [],
        }

    # Bulk-fetch invoice amounts for both row sets
    all_dockets = {r["docket_number"] for r in coupon_rows_all if r["docket_number"]} | \
                  {r["docket_number"] for r in coupon_rows_shop if r["docket_number"]}
    txn_map = {
        t["invoice_number"]: float(t["sales_amount"] or 0)
        for t in SalesTransaction.objects.filter(
            invoice_number__in=list(all_dockets)
        ).values("invoice_number", "sales_amount")
    }

    # Campaigns
    campaigns = list(CouponCampaign.objects.values("name", "prefix"))

    # ── Bucket key helpers ────────────────────────────────────────────────────
    # week: sort key = "YYYY-W01"; display label collected in week_label_map
    week_label_map = {}  # sort_key → "Week N (d/m-d/m)"

    def _week(d):
        sort_key, display_label = get_week_info(d)
        week_label_map[sort_key] = display_label
        return sort_key

    def _month(d):
        return f"{d.year}-{d.month:02d}"

    def _season(d):
        return get_session_key(d)

    def _year(d):
        return str(d.year)

    bucket_fns = {"week": _week, "month": _month, "season": _season, "year": _year}

    # Nested defaultdicts: [btype][entity][time_key] = {used, coupon_amount, unique_amount, _seen}
    def _new_bucket():
        return {"used": 0, "coupon_amount": 0.0, "unique_amount": 0.0, "_seen": set()}

    shop_agg = {b: defaultdict(lambda: defaultdict(_new_bucket)) for b in bucket_fns}
    campaign_agg = {
        b: defaultdict(lambda: defaultdict(_new_bucket)) for b in bucket_fns
    }
    # separate totals: shop uses prefix-filtered rows, campaign uses all rows
    total_agg_shop = {b: defaultdict(int) for b in bucket_fns}
    total_agg_camp = {b: defaultdict(int) for b in bucket_fns}

    def _row_amounts(row):
        inv_amt = txn_map.get(row["docket_number"], 0.0) if row["docket_number"] else 0.0
        if not inv_amt and row["face_value"] and float(row["face_value"]) > 1:
            inv_amt = float(row["face_value"])
        return inv_amt, float(calc_coupon_amount(row["face_value"], inv_amt))

    # ── Chart 1: shop series (prefix-filtered) ────────────────────────────────
    for row in coupon_rows_shop:
        d = row["using_date"]
        if not d:
            continue
        shop = row["using_shop"] or "Unknown"
        dk = row["docket_number"] or f'__no_{row["coupon_id"]}'
        inv_amt, cpn_amt = _row_amounts(row)
        for btype, fn in bucket_fns.items():
            tkey = fn(d)
            total_agg_shop[btype][tkey] += 1
            sb = shop_agg[btype][shop][tkey]
            sb["used"] += 1
            sb["coupon_amount"] += cpn_amt
            if dk not in sb["_seen"]:
                sb["_seen"].add(dk)
                sb["unique_amount"] += inv_amt

    # ── Chart 2: campaign series (all coupons, no prefix filter) ─────────────
    for row in coupon_rows_all:
        d = row["using_date"]
        if not d:
            continue
        dk = row["docket_number"] or f'__no_{row["coupon_id"]}'
        inv_amt, cpn_amt = _row_amounts(row)
        coupon_id_upper = row["coupon_id"].upper()
        matched = [
            c["name"]
            for c in campaigns
            if any(
                coupon_id_upper.startswith(p.strip().upper())
                for p in c["prefix"].split(",")
                if p.strip()
            )
        ]
        camp_targets = matched if matched else ["(No Campaign)"]
        for btype, fn in bucket_fns.items():
            tkey = fn(d)
            total_agg_camp[btype][tkey] += 1
            for camp in camp_targets:
                cb = campaign_agg[btype][camp][tkey]
                cb["used"] += 1
                cb["coupon_amount"] += cpn_amt
                if dk not in cb["_seen"]:
                    cb["_seen"].add(dk)
                    cb["unique_amount"] += inv_amt

    # ── Collect sorted time labels (correct chronological order per type) ─────
    time_labels = {}
    for btype in bucket_fns:
        all_keys = set()
        for entity_d in shop_agg[btype].values():
            all_keys.update(entity_d.keys())
        if btype == "week":
            time_labels[btype] = sorted(all_keys, key=week_sort_key)
        elif btype == "season":
            time_labels[btype] = sorted(all_keys, key=session_sort_key)
        else:
            time_labels[btype] = sorted(all_keys)

    # ── Serialise (remove _seen sets) ─────────────────────────────────────────
    def _serialise(agg):
        return {
            btype: {
                entity: {
                    tkey: {k: v for k, v in b.items() if k != "_seen"}
                    for tkey, b in tkeys.items()
                }
                for entity, tkeys in entity_d.items()
            }
            for btype, entity_d in agg.items()
        }

    shop_series = _serialise(shop_agg)
    campaign_series = _serialise(campaign_agg)

    all_shops = sorted(shop_series.get("month", {}).keys())
    all_campaigns = sorted(campaign_series.get("month", {}).keys())

    total_by_time = {btype: dict(total_agg_camp[btype]) for btype in bucket_fns}
    total_by_time_shop = {btype: dict(total_agg_shop[btype]) for btype in bucket_fns}

    return {
        "time_labels": time_labels,
        "week_label_map": week_label_map,
        "total_by_time": total_by_time,
        "total_by_time_shop": total_by_time_shop,
        "shops": all_shops,
        "shop_series": shop_series,
        "campaigns": all_campaigns,
        "campaign_series": campaign_series,
        "campaign_list": campaigns,
    }


def export_coupon_to_excel(
    data, date_from=None, date_to=None, coupon_id_prefix=None, shop_group=None
):
    """
    Export coupon analytics to Excel workbook.

    Args:
        data: Coupon analytics dict from calculate_coupon_analytics()
        date_from: Period start date (for display)
        date_to: Period end date (for display)
        coupon_id_prefix: Coupon ID prefix filter (for display)
        shop_group: Shop group filter (for display)

    Returns:
        openpyxl Workbook object
    """
    wb = Workbook()

    header_fill = PatternFill(
        start_color="6F42C1", end_color="6F42C1", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    # Summary sheet
    ws = wb.active
    ws.title = "Summary"

    ws["A1"] = "Coupon Analytics Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:B1")

    # Display active filters
    row = 3
    if date_from and date_to:
        ws[f"A{row}"] = "Period Filter:"
        ws[f"B{row}"] = f"{date_from} to {date_to}"
        ws[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1

    if coupon_id_prefix:
        ws[f"A{row}"] = "Coupon ID Prefix:"
        ws[f"B{row}"] = coupon_id_prefix
        ws[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1

    if shop_group:
        ws[f"A{row}"] = "Shop Group:"
        ws[f"B{row}"] = shop_group
        ws[f"B{row}"].font = Font(bold=True, color="0066CC")
        row += 1

    row += 1
    ws[f"A{row}"] = "All-Time Statistics"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    at = data["all_time"]
    for label, value in [
        ("Total Coupons", at["total"]),
        ("Used", at["used"]),
        ("Unused", at["unused"]),
        ("Usage Rate", f"{at['usage_rate']}%"),
        ("Total Amount (Invoice)", f"{at['total_amount']:,.0f} VND"),
        ("Unique Invoice Amount", f"{at['unique_invoice_amount']:,.0f} VND"),
        ("Total Coupon Amount", f"{at['total_coupon_amount']:,.0f} VND"),
        ("Duplicate Invoice Count", at["duplicate_invoice_count"]),
    ]:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        row += 1

    row += 1
    ws[f"A{row}"] = "Period Statistics"
    ws[f"A{row}"].font = Font(bold=True)
    row += 1

    pd = data["period"]
    for label, value in [
        ("Total Coupons", pd["total"]),
        ("Used", pd["used"]),
        ("Unused", pd["unused"]),
        ("Usage Rate", f"{pd['usage_rate']}%"),
        ("Total Amount (Invoice)", f"{pd['total_amount']:,.0f} VND"),
        ("Unique Invoice Amount", f"{pd['unique_invoice_amount']:,.0f} VND"),
        ("Total Coupon Amount", f"{pd['total_coupon_amount']:,.0f} VND"),
        ("Duplicate Invoice Count", pd["duplicate_invoice_count"]),
    ]:
        ws[f"A{row}"] = label
        ws[f"B{row}"] = value
        row += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20

    # By Using Shop sheet
    ws_shop = wb.create_sheet("By Using Shop")

    headers = [
        "Using Shop",
        "Total",
        "Used",
        "Unused",
        "Usage Rate",
        "% of Used (All Shops)",
        "Coupon Amount",
        "Total Amount (Unique Invoice)",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws_shop.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, shop in enumerate(data["by_shop"], 2):
        ws_shop.cell(row=row_num, column=1, value=shop["shop_name"])
        ws_shop.cell(row=row_num, column=2, value=shop["total"])
        ws_shop.cell(row=row_num, column=3, value=shop["used"])
        ws_shop.cell(row=row_num, column=4, value=shop["unused"])
        ws_shop.cell(row=row_num, column=5, value=f"{shop['usage_rate']}%")
        ws_shop.cell(row=row_num, column=6, value=f"{shop['used_pct_of_used']}%")
        ws_shop.cell(row=row_num, column=7, value=f"{shop['coupon_amount']:,.0f}")
        ws_shop.cell(row=row_num, column=8, value=f"{shop['total_amount']:,.0f}")

    for col in range(1, 9):
        ws_shop.column_dimensions[get_column_letter(col)].width = 22

    # Details sheet
    ws_detail = wb.create_sheet("Coupon Details")

    headers = [
        "Coupon ID",
        "Creator",
        "Face Value",
        "Using Shop",
        "Using Date",
        "Docket Number",
        "VIP ID",
        "Name",
        "Phone",
        "Sales Date",
        "Invoice Shop",
        "Amount (Invoice)",
        "Coupon Amount",
        "Note",
        "CNV ID",
        "CNV Points",
        "CNV Total Points",
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_num, detail in enumerate(data["details"], 2):
        ws_detail.cell(row=row_num, column=1, value=detail["coupon_id"])
        ws_detail.cell(row=row_num, column=2, value=detail["creator"])
        ws_detail.cell(row=row_num, column=3, value=detail["face_value_display"])
        ws_detail.cell(row=row_num, column=4, value=detail["using_shop"])
        ws_detail.cell(
            row=row_num,
            column=5,
            value=str(detail["using_date"]) if detail["using_date"] else "",
        )
        ws_detail.cell(row=row_num, column=6, value=detail["docket_number"])
        ws_detail.cell(row=row_num, column=7, value=detail["vip_id"])
        ws_detail.cell(row=row_num, column=8, value=detail["customer_name"])
        ws_detail.cell(row=row_num, column=9, value=detail["customer_phone"])
        ws_detail.cell(
            row=row_num,
            column=10,
            value=str(detail["sales_day"]) if detail["sales_day"] else "",
        )
        ws_detail.cell(row=row_num, column=11, value=detail["inv_shop"])
        ws_detail.cell(row=row_num, column=12, value=f"{detail['amount']:,.0f}")
        ws_detail.cell(row=row_num, column=13, value=f"{detail['coupon_amount']:,.0f}")
        ws_detail.cell(row=row_num, column=14, value=detail["note"])
        ws_detail.cell(row=row_num, column=15, value=detail.get("cnv_id") or "")
        ws_detail.cell(
            row=row_num,
            column=16,
            value=float(detail["cnv_points"]) if detail.get("cnv_points") != "" else "",
        )
        ws_detail.cell(
            row=row_num,
            column=17,
            value=(
                float(detail["cnv_total_points"])
                if detail.get("cnv_total_points") != ""
                else ""
            ),
        )

    for col in range(1, 18):
        ws_detail.column_dimensions[get_column_letter(col)].width = 18

    return wb
