"""
analytics.py  --  v3.1 - FINAL VERSION WITH UNIFIED FUNCTIONS

✅ ALL BUGS FIXED:
1. Customer Details: Unified customer lookup function
2. By Season: Cross-season returns calculated correctly
3. Shop → By Season: Cross-shop-season returns calculated correctly  
4. Shop → By Grade: Now uses unified customer lookup (fixes "No Grade" issue)

RETURN VISIT FORMULA (user-confirmed):
  Registration Date = First Purchase Date:
    - First invoice of that day    = first visit (NOT a return)
    - Subsequent invoices same day = return visits
    - Example: Reg 1/1, buy 3x on 1/1 -> total_purchases=3, return_visits=2
  
  Registration Date != First Purchase Date (or no reg date):
    - Customer already visited to register, so first purchase = already returning
    - First invoice = return visit
    - Example: Reg 1/1, buy 1x on 5/1 -> total_purchases=1, return_visits=1
  
  is_returning = return_visits > 0

KEY IMPROVEMENTS:
  - Unified get_customer_info() function used everywhere
  - Prevents "No Grade" proliferation from None foreign keys
  - Consistent customer data across all breakdowns
"""
import logging
from collections import defaultdict
from decimal import Decimal

from django.db.models import Count, Min, Max

from .models import Customer, SalesTransaction, Coupon
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger('customer_analytics')

GRADE_ORDER = {'No Grade': 0, 'Member': 1, 'Silver': 2, 'Gold': 3, 'Diamond': 4}

SEASON_DEFS = [
    ('M2-4',  [2, 3, 4]),
    ('M5-7',  [5, 6, 7]),
    ('M8-10', [8, 9, 10]),
    ('M11-1', [11, 12, 1]),
]

# ============================================================================
# CACHE for customer lookups to avoid repeated DB queries
# ============================================================================
_customer_cache = {}


def _normalize_grade(raw):
    if not raw:
        return 'No Grade'
    raw = raw.strip()
    if raw.lower() in ('olden', 'gold', 'golden'):
        return 'Gold'
    return raw


def _get_session_key(d):
    """Year-aware season label."""
    if not d:
        return 'Unknown'
    m, y = d.month, d.year
    for prefix, months in SEASON_DEFS:
        if m in months:
            if prefix == 'M11-1':
                return f"M11-1 {y-1}-{y}" if m == 1 else f"M11-1 {y}-{y+1}"
            return f"{prefix} {y}"
    return 'Unknown'


def _session_sort_key(label):
    parts = label.rsplit(' ', 1)
    if len(parts) != 2:
        return (9999, 99)
    prefix, years = parts[0], parts[1]
    try:
        first_year = int(years.split('-')[0])
    except ValueError:
        return (9999, 99)
    order_map = {'M2-4': 0, 'M5-7': 1, 'M8-10': 2, 'M11-1': 3}
    return (first_year, order_map.get(prefix, 99))


def _get_session_for_range(date_from, date_to):
    if not date_from or not date_to:
        return None
    months, cur = set(), date_from.replace(day=1)
    while cur <= date_to:
        months.add(cur.month)
        cur = cur.replace(year=cur.year + 1, month=1) if cur.month == 12 \
            else cur.replace(month=cur.month + 1)
    for label, m_list in SEASON_DEFS:
        if months.issubset(set(m_list)):
            return label
    return None


# ============================================================================
# UNIFIED CUSTOMER INFO FUNCTION - Used everywhere for consistency
# ============================================================================
def get_customer_info(vip_id, customer_obj=None):
    """
    Get customer grade and registration date.
    
    This function ensures consistent customer data across ALL calculations:
    - Period-level
    - By Grade
    - By Season  
    - Shop-level
    - Shop → By Grade
    - Shop → By Season
    
    Args:
        vip_id: Customer VIP ID
        customer_obj: Customer object from foreign key (may be None)
    
    Returns:
        (grade, reg_date, customer_name)
    """
    global _customer_cache
    
    # Special case: VIP ID = 0 (no customer info)
    if vip_id == '0':
        return ('No Grade', None, 'Unknown (No VIP)')
    
    # Try to use provided customer object first
    cust = customer_obj
    
    # If customer object is None, lookup from database (with caching)
    if not cust:
        if vip_id in _customer_cache:
            cust = _customer_cache[vip_id]
        else:
            try:
                cust = Customer.objects.get(vip_id=vip_id)
                _customer_cache[vip_id] = cust
            except Customer.DoesNotExist:
                _customer_cache[vip_id] = None
                cust = None
    
    # Extract info
    if cust:
        grade = _normalize_grade(cust.vip_grade)
        reg_date = cust.registration_date
        name = cust.name or 'Unknown'
    else:
        # Customer not found in database
        grade = 'No Grade'
        reg_date = None
        name = 'Unknown'
    
    return (grade, reg_date, name)


def _calculate_return_visits(purchases_sorted, reg_date):
    """
    Return (return_visits, is_returning).
    
    Rule:
      - If reg_date == first_purchase_date: first invoice NOT a return, rest are
        -> return_visits = total_purchases - 1
      - If reg_date != first_purchase_date (or no reg): first invoice IS a return
        -> return_visits = total_purchases
      - is_returning = return_visits > 0
    """
    n = len(purchases_sorted)
    if n == 0:
        return (0, False)
    
    first_date = purchases_sorted[0]['date']
    
    if reg_date and first_date == reg_date:
        return_visits = n - 1
    else:
        return_visits = n
    
    return (return_visits, return_visits > 0)


def _empty_bucket():
    return {'active': 0, 'returning': 0, 'invoices': 0, 'amount': Decimal(0)}


def calculate_return_rate_analytics(date_from=None, date_to=None):
    """
    Calculate return visit analytics with ALL BUGS FIXED.
    
    Uses unified get_customer_info() function throughout to ensure:
    - Consistent customer data
    - Proper grade classification
    - Minimal "No Grade" customers (only VIP ID = 0)
    """
    global _customer_cache
    _customer_cache = {}  # Reset cache for each calculation
    
    logger.info("START date_from=%s date_to=%s", date_from, date_to)

    total_customers_in_db    = Customer.objects.count()
    member_active_all_time   = Customer.objects.filter(points__gt=0).count()
    member_inactive_all_time = Customer.objects.filter(points=0).count()

    qs = SalesTransaction.objects.select_related('customer').order_by('sales_date')
    if date_from:
        qs = qs.filter(sales_date__gte=date_from)
    if date_to:
        qs = qs.filter(sales_date__lte=date_to)
    if not qs.exists():
        return None

    sales_list = list(qs)
    date_stats = qs.aggregate(start_date=Min('sales_date'), end_date=Max('sales_date'))
    logger.info("Transactions: %d", len(sales_list))

    # Group by customer (VIP ID blank/0 -> key='0')
    customer_purchases = defaultdict(list)
    for s in sales_list:
        vid = (s.vip_id or '').strip()
        key = '0' if vid in ('', '0', 'None') else vid
        customer_purchases[key].append({
            'date':     s.sales_date,
            'invoice':  s.invoice_number,
            'amount':   s.sales_amount or Decimal(0),
            'shop':     s.shop_name or 'Unknown Shop',
            'customer': s.customer if key != '0' else None,
            'session':  _get_session_key(s.sales_date),
        })

    buyer_no_info_invoices = len(customer_purchases.get('0', []))
    logger.info("Customers: %d  buyer_no_info_invoices: %d",
                len(customer_purchases), buyer_no_info_invoices)

    # ========================================================================
    # PERIOD-LEVEL METRICS
    # ========================================================================
    returning_customers   = set()
    new_members_in_period = set()
    customer_details      = []
    total_amount_period   = Decimal(0)

    for vip_id, purchases in customer_purchases.items():
        purchases_sorted = sorted(purchases, key=lambda x: x['date'])
        n = len(purchases_sorted)

        # UNIFIED: Get customer info
        grade, reg_date, name = get_customer_info(vip_id, purchases_sorted[0]['customer'])

        rc, is_ret = _calculate_return_visits(purchases_sorted, reg_date)
        if is_ret:
            returning_customers.add(vip_id)

        if reg_date and vip_id != '0':
            lo = date_from or date_stats['start_date']
            hi = date_to   or date_stats['end_date']
            if lo <= reg_date <= hi:
                new_members_in_period.add(vip_id)

        amt = sum(p['amount'] for p in purchases_sorted)
        total_amount_period += amt

        customer_details.append({
            'vip_id':              vip_id,
            'name':                name,
            'vip_grade':           grade,
            'registration_date':   reg_date,
            'first_purchase_date': purchases_sorted[0]['date'],
            'total_purchases':     n,
            'return_visits':       rc,
            'total_spent':         float(amt),
        })

    total_active    = len(customer_purchases)
    total_returning = len(returning_customers)
    return_rate_p   = round(total_returning / total_active          * 100, 2) if total_active          else 0
    return_rate_at  = round(total_returning / total_customers_in_db * 100, 2) if total_customers_in_db else 0

    # All-time grade counts
    grade_db = {}
    for row in Customer.objects.values('vip_grade').annotate(cnt=Count('id')):
        g = _normalize_grade(row['vip_grade'])
        grade_db[g] = grade_db.get(g, 0) + row['cnt']

    # ========================================================================
    # BY VIP GRADE (ALL PERIOD)
    # ========================================================================
    grade_buckets = defaultdict(_empty_bucket)
    for d in customer_details:
        g = d['vip_grade']
        grade_buckets[g]['active']   += 1
        grade_buckets[g]['amount']   += Decimal(str(d['total_spent']))
        grade_buckets[g]['invoices'] += d['total_purchases']
        if d['return_visits'] > 0:
            grade_buckets[g]['returning'] += 1

    for g in GRADE_ORDER:
        if g not in grade_buckets:
            _ = grade_buckets[g]

    grade_stats = []
    for grade in sorted(grade_buckets, key=lambda x: GRADE_ORDER.get(x, 99)):
        s   = grade_buckets[grade]
        tdb = grade_db.get(grade, 0)
        grade_stats.append({
            'grade':                grade,
            'total_customers':      s['active'],
            'total_in_db':          tdb,
            'returning_customers':  s['returning'],
            'return_rate':          round(s['returning'] / s['active'] * 100 if s['active'] else 0, 2),
            'return_rate_all_time': round(s['returning'] / tdb         * 100 if tdb          else 0, 2),
            'total_invoices':       s['invoices'],
            'total_amount':         float(s['amount']),
        })

    # ========================================================================
    # BY SEASON (ALL PERIOD) - FIXED: Cross-season returns
    # ========================================================================
    session_buckets = defaultdict(_empty_bucket)

    for vip_id, purchases in customer_purchases.items():
        # UNIFIED: Get customer info
        grade, reg_date, name = get_customer_info(vip_id, purchases[0]['customer'])

        # Sort ALL purchases by date (needed for cross-season return check)
        all_purchases_sorted = sorted(purchases, key=lambda x: x['date'])
        
        # Group by session
        by_sess = defaultdict(list)
        for p in purchases:
            by_sess[p['session']].append(p)

        for sk, sp in by_sess.items():
            sp_sorted = sorted(sp, key=lambda x: x['date'])
            session_first_date = sp_sorted[0]['date']
            
            # Customer is ACTIVE in this session
            session_buckets[sk]['active']   += 1
            session_buckets[sk]['invoices'] += len(sp)
            session_buckets[sk]['amount']   += sum(p['amount'] for p in sp)
            
            # FIX: Customer is RETURNING in this session if they have ANY prior purchases
            # (not just within this season)
            has_prior_purchases = any(p['date'] < session_first_date for p in all_purchases_sorted)
            reg_before_session = (reg_date and reg_date < session_first_date)
            
            if has_prior_purchases or reg_before_session:
                session_buckets[sk]['returning'] += 1

    session_stats = []
    for key in sorted(session_buckets, key=_session_sort_key):
        s = session_buckets[key]
        ac = s['active']; rc = s['returning']
        session_stats.append({
            'session':             key,
            'total_customers':     ac,
            'returning_customers': rc,
            'return_rate':         round(rc / ac * 100 if ac else 0, 2),
            'total_invoices':      s['invoices'],
            'total_amount':        float(s['amount']),
        })
    logger.info("session_stats: %d rows", len(session_stats))

    # ========================================================================
    # SHOP STATS WITH SUB-BREAKDOWNS - ALL FIXED
    # ========================================================================
    shop_customers  = defaultdict(set)
    shop_invoices   = defaultdict(int)
    shop_amount     = defaultdict(Decimal)
    shop_returning  = defaultdict(set)
    shop_grade      = defaultdict(lambda: defaultdict(lambda: {
        'active': set(), 'returning': set(), 'invoices': 0, 'amount': Decimal(0)
    }))
    shop_sess       = defaultdict(lambda: defaultdict(_empty_bucket))

    for vip_id, purchases in customer_purchases.items():
        # UNIFIED: Get customer info (CRITICAL FIX for shop → by grade)
        grade, reg_date, name = get_customer_info(vip_id, purchases[0]['customer'])

        by_shop_visits = defaultdict(list)
        for p in purchases:
            by_shop_visits[p['shop']].append(p)

        for sh, sh_p in by_shop_visits.items():
            shop_customers[sh].add(vip_id)
            shop_invoices[sh] += len(sh_p)
            for p in sh_p:
                shop_amount[sh] += p['amount']

            # Shop-level returning calculation
            _, ret = _calculate_return_visits(sorted(sh_p, key=lambda x: x['date']), reg_date)
            if ret:
                shop_returning[sh].add(vip_id)

            # ================================================================
            # Shop → By Grade - FIXED: Uses unified customer lookup
            # ================================================================
            shop_grade[sh][grade]['active'].add(vip_id)
            if ret:
                shop_grade[sh][grade]['returning'].add(vip_id)
            for p in sh_p:
                shop_grade[sh][grade]['invoices'] += 1
                shop_grade[sh][grade]['amount']   += p['amount']

            # ================================================================
            # Shop → By Session - FIXED: Cross-season returns
            # ================================================================
            # Sort all shop purchases (needed for cross-season return check)
            sh_p_sorted = sorted(sh_p, key=lambda x: x['date'])
            
            # Group shop purchases by session
            by_sess_shop = defaultdict(list)
            for p in sh_p:
                by_sess_shop[p['session']].append(p)
            
            for sk, sp in by_sess_shop.items():
                sp_sorted = sorted(sp, key=lambda x: x['date'])
                session_first_date = sp_sorted[0]['date']
                
                shop_sess[sh][sk]['active']   += 1
                shop_sess[sh][sk]['invoices'] += len(sp)
                shop_sess[sh][sk]['amount']   += sum(p['amount'] for p in sp)
                
                # FIX: Check if customer has SHOP purchases before this season
                has_prior_shop_purchases = any(p['date'] < session_first_date for p in sh_p_sorted)
                reg_before_session = (reg_date and reg_date < session_first_date)
                
                if has_prior_shop_purchases or reg_before_session:
                    shop_sess[sh][sk]['returning'] += 1

    all_session_keys = sorted(session_buckets.keys(), key=_session_sort_key)

    shop_stats = []
    for sh in shop_customers:
        ac = len(shop_customers[sh])
        rc = len(shop_returning[sh])

        by_grade_list = []
        for grade in sorted(shop_grade[sh], key=lambda x: GRADE_ORDER.get(x, 99)):
            gd = shop_grade[sh][grade]
            a  = len(gd['active']); r = len(gd['returning'])
            by_grade_list.append({
                'grade':               grade,
                'total_customers':     a,
                'returning_customers': r,
                'return_rate':         round(r / a * 100 if a else 0, 2),
                'total_invoices':      gd['invoices'],
                'total_amount':        float(gd['amount']),
            })

        by_session_list = []
        for key in all_session_keys:
            sd = shop_sess[sh].get(key)
            if sd is None:
                sd = _empty_bucket()
            a = sd['active']; r = sd['returning']
            by_session_list.append({
                'session':             key,
                'total_customers':     a,
                'returning_customers': r,
                'return_rate':         round(r / a * 100 if a else 0, 2),
                'total_invoices':      sd['invoices'],
                'total_amount':        float(sd['amount']),
            })

        shop_stats.append({
            'shop_name':           sh,
            'total_customers':     ac,
            'returning_customers': rc,
            'return_rate':         round(rc / ac * 100 if ac else 0, 2),
            'total_invoices':      shop_invoices[sh],
            'total_amount':        float(shop_amount[sh]),
            'by_grade':            by_grade_list,
            'by_session':          by_session_list,
        })

    shop_stats.sort(key=lambda x: x['total_customers'], reverse=True)
    customer_details.sort(key=lambda x: x['return_visits'], reverse=True)
    logger.info("DONE  total_amount=%.2f  shops=%d  sessions=%d",
                float(total_amount_period), len(shop_stats), len(session_stats))

    return {
        'date_range':    {'start': date_stats['start_date'], 'end': date_stats['end_date']},
        'session_label': _get_session_for_range(date_from, date_to),
        'overview': {
            'active_customers':        total_active,
            'returning_customers':     total_returning,
            'return_rate':             return_rate_p,
            'return_rate_all_time':    return_rate_at,
            'total_amount_period':     float(total_amount_period),
            'buyer_without_info':      buyer_no_info_invoices,
            'new_members_in_period':   len(new_members_in_period),
            'total_customers_in_db':   total_customers_in_db,
            'member_active_all_time':  member_active_all_time,
            'member_inactive_all_time':member_inactive_all_time,
        },
        'by_grade':       grade_stats,
        'by_session':     session_stats,
        'by_shop':        shop_stats,
        'customer_details': customer_details,
    }


def export_analytics_to_excel(data):
    """Export analytics data to Excel workbook."""
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    sub_header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    sub_header_font = Font(bold=True, color="1F3864")
    
    center_align = Alignment(horizontal="center")
    
    # ========================================================================
    # OVERVIEW SHEET
    # ========================================================================
    ws = wb.active
    ws.title = "Overview"
    
    ov = data['overview']
    dr = data['date_range']
    
    ws['A1'] = "Customer Analytics - Overview"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    ws['A3'] = "Date Range:"
    ws['B3'] = f"{dr['start']} to {dr['end']}"
    
    row = 5
    metrics = [
        ("Total Customers (All Time)", ov['total_customers_in_db']),
        ("Active Customers (Period)", ov['active_customers']),
        ("Returning Customers (Period)", ov['returning_customers']),
        ("Return Visit Rate (Period)", f"{ov['return_rate']}%"),
        ("Return Visit Rate (All Time)", f"{ov['return_rate_all_time']}%"),
        ("New Members (Period)", ov['new_members_in_period']),
        ("Buyer Without Info (Period)", ov['buyer_without_info']),
        ("Total Amount (Period)", f"{ov['total_amount_period']:,.0f} VND"),
    ]
    
    for label, value in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    # ========================================================================
    # BY GRADE SHEET
    # ========================================================================
    ws_grade = wb.create_sheet("By VIP Grade")
    
    headers = ["Grade", "Active", "Returning", "Return Rate", "Rate (All Time)", "Invoices", "Amount (VND)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_grade.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, grade_stat in enumerate(data['by_grade'], 2):
        ws_grade.cell(row=row_num, column=1, value=grade_stat['grade'])
        ws_grade.cell(row=row_num, column=2, value=grade_stat['total_customers'])
        ws_grade.cell(row=row_num, column=3, value=grade_stat['returning_customers'])
        ws_grade.cell(row=row_num, column=4, value=f"{grade_stat['return_rate']}%")
        ws_grade.cell(row=row_num, column=5, value=f"{grade_stat['return_rate_all_time']}%")
        ws_grade.cell(row=row_num, column=6, value=grade_stat['total_invoices'])
        ws_grade.cell(row=row_num, column=7, value=f"{grade_stat['total_amount']:,.0f}")
    
    for col in range(1, 8):
        ws_grade.column_dimensions[get_column_letter(col)].width = 18
    
    # ========================================================================
    # BY SEASON SHEET
    # ========================================================================
    ws_session = wb.create_sheet("By Season")
    
    headers = ["Season", "Active", "Returning", "Return Rate", "Invoices", "Amount (VND)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_session.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, session_stat in enumerate(data['by_session'], 2):
        ws_session.cell(row=row_num, column=1, value=session_stat['session'])
        ws_session.cell(row=row_num, column=2, value=session_stat['total_customers'])
        ws_session.cell(row=row_num, column=3, value=session_stat['returning_customers'])
        ws_session.cell(row=row_num, column=4, value=f"{session_stat['return_rate']}%")
        ws_session.cell(row=row_num, column=5, value=session_stat['total_invoices'])
        ws_session.cell(row=row_num, column=6, value=f"{session_stat['total_amount']:,.0f}")
    
    for col in range(1, 7):
        ws_session.column_dimensions[get_column_letter(col)].width = 18
    
    # ========================================================================
    # BY SHOP SHEET
    # ========================================================================
    ws_shop = wb.create_sheet("By Shop")
    
    headers = ["Shop", "Active", "Returning", "Return Rate", "Invoices", "Amount (VND)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_shop.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, shop_stat in enumerate(data['by_shop'], 2):
        ws_shop.cell(row=row_num, column=1, value=shop_stat['shop_name'])
        ws_shop.cell(row=row_num, column=2, value=shop_stat['total_customers'])
        ws_shop.cell(row=row_num, column=3, value=shop_stat['returning_customers'])
        ws_shop.cell(row=row_num, column=4, value=f"{shop_stat['return_rate']}%")
        ws_shop.cell(row=row_num, column=5, value=shop_stat['total_invoices'])
        ws_shop.cell(row=row_num, column=6, value=f"{shop_stat['total_amount']:,.0f}")
    
    ws_shop.column_dimensions['A'].width = 30
    for col in range(2, 7):
        ws_shop.column_dimensions[get_column_letter(col)].width = 16
    
    # ========================================================================
    # CUSTOMER DETAILS SHEET
    # ========================================================================
    ws_detail = wb.create_sheet("Customer Details")
    
    headers = ["VIP ID", "Name", "Grade", "Reg Date", "First Purchase", "Purchases", "Return Visits", "Total Spent"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, customer in enumerate(data['customer_details'], 2):
        ws_detail.cell(row=row_num, column=1, value=customer['vip_id'])
        ws_detail.cell(row=row_num, column=2, value=customer['name'])
        ws_detail.cell(row=row_num, column=3, value=customer['vip_grade'])
        ws_detail.cell(row=row_num, column=4, value=str(customer['registration_date']) if customer['registration_date'] else '')
        ws_detail.cell(row=row_num, column=5, value=str(customer['first_purchase_date']))
        ws_detail.cell(row=row_num, column=6, value=customer['total_purchases'])
        ws_detail.cell(row=row_num, column=7, value=customer['return_visits'])
        ws_detail.cell(row=row_num, column=8, value=f"{customer['total_spent']:,.0f}")
    
    for col in range(1, 9):
        ws_detail.column_dimensions[get_column_letter(col)].width = 18
    
    return wb


# ============================================================================
# COUPON ANALYTICS FUNCTIONS (Unchanged from original)
# ============================================================================

def calculate_coupon_analytics(date_from=None, date_to=None, coupon_id_prefix=None):
    """Calculate coupon usage analytics."""
    from django.db.models import Q, Count
    
    qs = Coupon.objects.all()
    
    # Filter by usage date if specified
    if date_from or date_to:
        usage_filter = Q(using_date__isnull=False)
        if date_from:
            usage_filter &= Q(using_date__gte=date_from)
        if date_to:
            usage_filter &= Q(using_date__lte=date_to)
        period_qs = qs.filter(usage_filter)
    else:
        period_qs = qs
    
    # Filter by coupon ID prefix if specified (case-insensitive)
    if coupon_id_prefix:
        qs = qs.filter(coupon_id__istartswith=coupon_id_prefix)
        period_qs = period_qs.filter(coupon_id__istartswith=coupon_id_prefix)
    
    # All-time stats
    all_time_total = qs.count()
    all_time_used = qs.filter(using_date__isnull=False).count()
    all_time_unused = all_time_total - all_time_used
    all_time_usage_rate = round(all_time_used / all_time_total * 100 if all_time_total else 0, 2)
    all_time_amount = sum(c.face_value or 0 for c in qs.filter(using_date__isnull=False))
    
    # Period stats
    period_total = period_qs.count()
    period_used = period_qs.filter(using_date__isnull=False).count()
    period_unused = period_total - period_used
    period_usage_rate = round(period_used / period_total * 100 if period_total else 0, 2)
    period_amount = sum(c.face_value or 0 for c in period_qs.filter(using_date__isnull=False))
    
    # By shop
    shop_stats = []
    shop_data = {}
    
    for coupon in period_qs.filter(using_date__isnull=False):
        shop = coupon.shop_name or 'Unknown'
        if shop not in shop_data:
            shop_data[shop] = {'total': 0, 'used': 0, 'amount': Decimal(0)}
        shop_data[shop]['used'] += 1
        shop_data[shop]['amount'] += coupon.face_value or Decimal(0)
    
    # Add unused coupons to shop totals
    for coupon in period_qs.filter(using_date__isnull=True):
        shop = coupon.shop_name or 'Unknown'
        if shop not in shop_data:
            shop_data[shop] = {'total': 0, 'used': 0, 'amount': Decimal(0)}
        shop_data[shop]['total'] += 1
    
    for shop in period_qs.values('shop_name').distinct():
        shop_name = shop['shop_name'] or 'Unknown'
        if shop_name in shop_data:
            shop_data[shop_name]['total'] += shop_data[shop_name]['used']
    
    total_used_all_shops = period_used
    
    for shop_name, data in shop_data.items():
        total = data['total']
        used = data['used']
        pct_of_total = round(used / period_total * 100 if period_total else 0, 2)
        pct_of_used = round(used / total_used_all_shops * 100 if total_used_all_shops else 0, 2)
        usage_rate = round(used / total * 100 if total else 0, 2)
        
        shop_stats.append({
            'shop_name': shop_name,
            'total_coupons': total,
            'used_coupons': used,
            'unused_coupons': total - used,
            'usage_rate': usage_rate,
            'pct_of_total': pct_of_total,
            'pct_of_used': pct_of_used,
            'total_amount': float(data['amount']),
        })
    
    shop_stats.sort(key=lambda x: x['used_coupons'], reverse=True)
    
    # Coupon details
    coupon_details = []
    for coupon in period_qs.filter(using_date__isnull=False).order_by('-using_date'):
        # Try to get customer info from transaction
        vip_id = None
        vip_name = None
        phone = None
        
        if coupon.invoice_number:
            try:
                txn = SalesTransaction.objects.get(invoice_number=coupon.invoice_number)
                vip_id = txn.vip_id
                if vip_id and vip_id != '0':
                    try:
                        cust = Customer.objects.get(vip_id=vip_id)
                        vip_name = cust.name
                        phone = cust.phone_number
                    except Customer.DoesNotExist:
                        pass
            except SalesTransaction.DoesNotExist:
                pass
        
        coupon_details.append({
            'coupon_id': coupon.coupon_id,
            'face_value': coupon.face_value or 0,
            'using_date': coupon.using_date,
            'shop_name': coupon.shop_name or 'Unknown',
            'invoice_number': coupon.invoice_number or '',
            'vip_id': vip_id or '',
            'vip_name': vip_name or '-',
            'phone': phone or '-',
            'amount': coupon.face_value or 0,
        })
    
    return {
        'all_time': {
            'total_coupons': all_time_total,
            'used_coupons': all_time_used,
            'unused_coupons': all_time_unused,
            'usage_rate': all_time_usage_rate,
            'total_amount': float(all_time_amount),
        },
        'period': {
            'total_coupons': period_total,
            'used_coupons': period_used,
            'unused_coupons': period_unused,
            'usage_rate': period_usage_rate,
            'total_amount': float(period_amount),
        },
        'by_shop': shop_stats,
        'details': coupon_details,
    }


def export_coupon_to_excel(data, date_from=None, date_to=None):
    """Export coupon analytics to Excel."""
    wb = Workbook()
    
    header_fill = PatternFill(start_color="6F42C1", end_color="6F42C1", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    
    # Summary sheet
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = "Coupon Analytics Summary"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    if date_from and date_to:
        ws['A3'] = "Period:"
        ws['B3'] = f"{date_from} to {date_to}"
    
    row = 5
    ws[f'A{row}'] = "All-Time Statistics"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    at = data['all_time']
    for label, value in [
        ("Total Coupons", at['total_coupons']),
        ("Used", at['used_coupons']),
        ("Unused", at['unused_coupons']),
        ("Usage Rate", f"{at['usage_rate']}%"),
        ("Total Amount", f"{at['total_amount']:,.0f} VND"),
    ]:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    row += 1
    ws[f'A{row}'] = "Period Statistics"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    pd = data['period']
    for label, value in [
        ("Total Coupons", pd['total_coupons']),
        ("Used", pd['used_coupons']),
        ("Unused", pd['unused_coupons']),
        ("Usage Rate", f"{pd['usage_rate']}%"),
        ("Total Amount", f"{pd['total_amount']:,.0f} VND"),
    ]:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # By Shop sheet
    ws_shop = wb.create_sheet("By Shop")
    
    headers = ["Shop", "Total", "Used", "Unused", "Usage Rate", "% of Total", "% of Used", "Amount (VND)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_shop.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, shop in enumerate(data['by_shop'], 2):
        ws_shop.cell(row=row_num, column=1, value=shop['shop_name'])
        ws_shop.cell(row=row_num, column=2, value=shop['total_coupons'])
        ws_shop.cell(row=row_num, column=3, value=shop['used_coupons'])
        ws_shop.cell(row=row_num, column=4, value=shop['unused_coupons'])
        ws_shop.cell(row=row_num, column=5, value=f"{shop['usage_rate']}%")
        ws_shop.cell(row=row_num, column=6, value=f"{shop['pct_of_total']}%")
        ws_shop.cell(row=row_num, column=7, value=f"{shop['pct_of_used']}%")
        ws_shop.cell(row=row_num, column=8, value=f"{shop['total_amount']:,.0f}")
    
    for col in range(1, 9):
        ws_shop.column_dimensions[get_column_letter(col)].width = 16
    
    # Details sheet
    ws_detail = wb.create_sheet("Coupon Details")
    
    headers = ["Coupon ID", "Face Value", "Using Date", "Shop", "Invoice", "VIP ID", "Name", "Phone"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    
    for row_num, detail in enumerate(data['details'], 2):
        ws_detail.cell(row=row_num, column=1, value=detail['coupon_id'])
        ws_detail.cell(row=row_num, column=2, value=f"{detail['face_value']:,.0f}")
        ws_detail.cell(row=row_num, column=3, value=str(detail['using_date']) if detail['using_date'] else '')
        ws_detail.cell(row=row_num, column=4, value=detail['shop_name'])
        ws_detail.cell(row=row_num, column=5, value=detail['invoice_number'])
        ws_detail.cell(row=row_num, column=6, value=detail['vip_id'])
        ws_detail.cell(row=row_num, column=7, value=detail['vip_name'])
        ws_detail.cell(row=row_num, column=8, value=detail['phone'])
    
    for col in range(1, 9):
        ws_detail.column_dimensions[get_column_letter(col)].width = 18
    
    return wb