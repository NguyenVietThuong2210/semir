"""App/views/membership.py — Customer Membership KPI snapshot page."""
import logging
from datetime import datetime

from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render

from App.analytics.membership import (
    DISPLAY_GRADES,
    compare_batches,
    get_all_batch_grade_series,
    get_grade_breakdown_by_store_comparison,
    get_grade_changes,
    get_grade_changes_overview_by_store,
    get_grade_changes_store_transitions,
    get_live_customer_tier_table,
    get_snapshot_registration_stores,
    list_batches,
)
from App.forms import MembershipBackfillForm
from App.permissions import requires_perm, user_has_perm
from App.upload_jobs import acquire_type_lock, create_job, is_type_running
from App.views.shop_detail import _ajax_perm_check, _get_dropdown_options
from App.views.upload import _pre_upload_checks, _start_thread

logger = logging.getLogger(__name__)


def _default_batch_ids(batches):
    """Default From/To selection: the two most recent batches (or the single
    batch if only one exists)."""
    if not batches:
        return None, None
    if len(batches) == 1:
        return None, batches[0].id
    # batches is ordered -snapshot_date,-created_at (newest first)
    return batches[1].id, batches[0].id


@requires_perm("membership.view")
def membership_dashboard(request):
    batches = list_batches()
    default_from, default_to = _default_batch_ids(batches)

    def _parse_id(raw, default):
        if not raw:
            return default
        try:
            return int(raw)
        except (TypeError, ValueError):
            return default

    from_batch = _parse_id(request.GET.get("from_batch"), default_from)
    to_batch = _parse_id(request.GET.get("to_batch"), default_to)

    comparison = compare_batches(from_batch, to_batch)
    chart_series = get_all_batch_grade_series()
    # Two DIFFERENT store lists, on purpose (PO feedback 2026-08-31): the
    # live Customer table for the LIVE-data "Customer Tier Progress" section,
    # vs. whatever actually appears in snapshot data for the two
    # snapshot-scoped sections below. Conflating them silently offered store
    # names absent from a given snapshot batch, producing a misleading
    # all-zero result.
    _, registration_stores, _, _, _, _ = _get_dropdown_options()
    snapshot_stores = get_snapshot_registration_stores()

    return render(
        request,
        "membership.html",
        {
            "batches": batches,
            "from_batch": from_batch,
            "to_batch": to_batch,
            "comparison": comparison,
            "chart_series": chart_series,
            "registration_stores": registration_stores,
            "snapshot_stores": snapshot_stores,
            "backfill_form": MembershipBackfillForm(),
            "can_import": user_has_perm(request.user, "membership.import"),
            "can_delete": user_has_perm(request.user, "membership.delete"),
        },
    )


def membership_table_partial(request):
    """Customer Tier Progress — reads the LIVE Customer table, not a snapshot
    (PO feedback 2026-08-31: "không liên quan gì đến snapshot"/has nothing to
    do with snapshot). Works even if zero snapshot batches exist yet."""
    err = _ajax_perm_check(request, "membership.view")
    if err:
        return err

    grade_filter = request.GET.get("grade") or None
    shop_filter = request.GET.get("shop") or None
    sort = request.GET.get("sort") or "amount_to_next_tier"

    rows, total_count = get_live_customer_tier_table(grade_filter=grade_filter, shop_filter=shop_filter, sort=sort)
    return render(request, "membership/_table_partial.html", {"rows": rows, "total_count": total_count})


def membership_store_breakdown_partial(request):
    """
    "Members per Grade — by Registration Store" — a From/To matrix, one row
    per store, columns = Member/Silver/Gold/Diamond each split into From/To
    (text green if increased, red if decreased, in the template). Redesigned
    2026-09-01, PO feedback: previously "To"-only with a separate per-store
    drill-down mode; the drill-down is now redundant (this matrix already
    shows From/To per store, and membership_movers_partial below shows the
    actual individual customers who moved) — removed rather than kept
    alongside, per independent UI-design review.
    """
    err = _ajax_perm_check(request, "membership.view")
    if err:
        return err

    def _parse_id(raw):
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    to_batch_id = _parse_id(request.GET.get("batch"))
    if not to_batch_id:
        batches = list_batches()
        to_batch_id = batches[0].id if batches else None  # newest first (Meta.ordering)
    from_batch_id = _parse_id(request.GET.get("from_batch"))

    if not to_batch_id:
        return HttpResponse('<div class="alert alert-warning m-0">No snapshot selected.</div>')

    rows = get_grade_breakdown_by_store_comparison(from_batch_id, to_batch_id)
    return render(request, "membership/_store_breakdown_partial.html", {
        "rows": rows, "grades": DISPLAY_GRADES, "from_batch": from_batch_id,
    })


def membership_movers_partial(request):
    """
    "Comparison" section — customers whose grade changed between the From/To
    snapshots. Added 2026-09-01, PO feedback (this is the whole reason
    MembershipSnapshotBatch.grade_members stores vip_id lists instead of just
    counts). Auto-loads like the by-Store matrix — cheap: bounded by how many
    customers actually changed grade between two existing snapshots, not by
    total customer count (see get_grade_changes() docstring).
    """
    err = _ajax_perm_check(request, "membership.view")
    if err:
        return err

    def _parse_id(raw):
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    from_batch_id = _parse_id(request.GET.get("from_batch"))
    to_batch_id = _parse_id(request.GET.get("to_batch"))
    if not to_batch_id:
        batches = list_batches()
        to_batch_id = batches[0].id if batches else None

    if not from_batch_id or not to_batch_id:
        return HttpResponse('<div class="alert alert-warning m-0">Select both a "From" and "To" snapshot to see grade changes.</div>')

    store = request.GET.get("store") or None
    grade = request.GET.get("grade") or None
    direction = request.GET.get("direction") or None

    rows, total_count = get_grade_changes(from_batch_id, to_batch_id, store=store, grade=grade, direction=direction, limit=500)
    return render(request, "membership/_movers_partial.html", {"rows": rows, "total_count": total_count})


def membership_movers_overview_partial(request):
    """
    "Comparison — Members Who Changed Grade" section's aggregate overview —
    a by-Registration-Store x Grade matrix of Downgrade/Upgrade counts,
    sitting above membership_movers_partial's individual-customer list.
    Mirrors membership_store_breakdown_partial exactly (same param parsing,
    same "select both" gating as membership_movers_partial since this also
    needs both ends of the comparison). Added 2026-09-02, PO feedback.

    Merged into ONE table (2026-09-02, PO feedback: "đừng làm nó
    complicated") — get_grade_changes_overview_by_store()'s per-store rows
    (customer's store matched on both sides) and
    get_grade_changes_store_transitions()'s itemized (from_store, to_store)
    pairs (store-name-drift cases the per-store rows can't attribute) are
    combined into a single row list here at the view layer; the analytics
    functions themselves are untouched. A transition row's `store` label is
    rendered as "From → To" so the template needs no awareness of the
    two-source origin — it just loops one `rows` list, same as before this
    change. The trailing 'All Stores' total row (always the LAST element of
    get_grade_changes_overview_by_store()'s return) is re-appended last so
    it stays the final row after the transition rows are spliced in.
    """
    err = _ajax_perm_check(request, "membership.view")
    if err:
        return err

    def _parse_id(raw):
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    from_batch_id = _parse_id(request.GET.get("from_batch"))
    to_batch_id = _parse_id(request.GET.get("to_batch"))

    if not from_batch_id or not to_batch_id:
        return HttpResponse('<div class="alert alert-warning m-0">Select both a "From" and "To" snapshot to see grade changes.</div>')

    store_rows = get_grade_changes_overview_by_store(from_batch_id, to_batch_id)
    transitions = get_grade_changes_store_transitions(from_batch_id, to_batch_id)

    if store_rows:
        *per_store_rows, all_stores_row = store_rows  # 'All Stores' is always last
    else:
        per_store_rows, all_stores_row = [], None

    transition_rows = [
        {
            "store": f"{t['from_store']} → {t['to_store']}",
            "is_total": False,
            "is_transition": True,
            "counts": t["counts"],
            "total_downgrade": t["total_downgrade"],
            "total_upgrade": t["total_upgrade"],
        }
        for t in transitions
    ]

    rows = per_store_rows + transition_rows
    if all_stores_row is not None:
        rows.append(all_stores_row)

    return render(request, "membership/_movers_overview_partial.html", {
        "rows": rows, "grades": DISPLAY_GRADES,
    })


def membership_trend_partial(request):
    """JSON (not HTML) — the trend chart is built entirely client-side by
    Chart.js from a JSON array, so an HTML fragment would be the wrong shape.
    Added 2026-08-31, PO feedback: chart needs a store filter. Bounded by
    batch count, not store count — see get_all_batch_grade_series() docstring."""
    err = _ajax_perm_check(request, "membership.view")
    if err:
        return err
    store = request.GET.get("store") or None
    return JsonResponse({"series": get_all_batch_grade_series(store=store)})


@requires_perm("membership.delete")
def membership_delete_batch(request, batch_id):
    if request.method != "POST":
        return redirect("membership_dashboard")

    from App.models.membership import MembershipSnapshotBatch

    try:
        batch = MembershipSnapshotBatch.objects.get(pk=batch_id)
    except MembershipSnapshotBatch.DoesNotExist:
        messages.error(request, "Snapshot not found.")
        return redirect("membership_dashboard")

    snapshot_date, source_display, row_count = batch.snapshot_date, batch.get_source_display(), batch.row_count
    batch.delete()  # grade_counts/grade_members JSON fields go with the row — no child model/cascade involved
    logger.info(
        "membership_delete_batch batch=%s date=%s source=%s rows=%s user=%s",
        batch_id, snapshot_date, source_display, row_count, request.user, extra={"step": "membership_snapshot"},
    )
    messages.info(request, f"Deleted snapshot {snapshot_date} ({source_display}, {row_count} rows).")
    return redirect("membership_dashboard")


@requires_perm("membership.import")
def membership_backfill_import(request):
    if request.method != "POST":
        return redirect("membership_dashboard")

    form = MembershipBackfillForm(request.POST, request.FILES)
    if not form.is_valid():
        messages.error(request, "Invalid form submission.")
        return redirect("membership_dashboard")

    if is_type_running("membership_backfill"):
        messages.warning(request, "A backfill import is already in progress. Please wait for it to finish.")
        return redirect("membership_dashboard")

    f = request.FILES["file"]
    pre = _pre_upload_checks(request, f, "customers")
    if pre is None:
        return redirect("membership_dashboard")
    file_bytes, file_hash, df = pre

    if not acquire_type_lock("membership_backfill"):
        messages.warning(request, "A backfill import is already in progress. Please wait.")
        return redirect("membership_dashboard")

    job_id = create_job("membership_backfill", f.name, file_hash=file_hash)
    logger.info(
        "membership_backfill_import queued job=%s file=%s user=%s",
        job_id, f.name, request.user, extra={"step": "membership_backfill"},
    )

    from functools import partial

    from App.services.membership_snapshot import create_backfill_snapshot

    fn = partial(
        create_backfill_snapshot,
        snapshot_date=form.cleaned_data["snapshot_date"],
        uploaded_by=request.user if request.user.is_authenticated else None,
        note=form.cleaned_data.get("note", ""),
    )
    _start_thread(job_id, fn, file_bytes, f.name, None, df)
    messages.info(request, f"Backfill snapshot import started — tracking job {job_id[:8]}…")
    return redirect("membership_dashboard")


@requires_perm("membership.compute")
def compute_grade_progress(request):
    """Kick off a full-DB recompute of CustomerGradeProgress (the per-customer
    grade-change DATE simulation, App/analytics/grade_simulation.py) as a
    background job — same async-job pattern as membership_backfill_import
    above, minus the file upload (this computes purely from existing
    SalesTransaction + Customer data, see
    App/services/grade_progress_calc.py::compute_all_grade_progress)."""
    if request.method != "POST":
        return redirect("membership_dashboard")

    if is_type_running("grade_progress_calc"):
        messages.warning(request, "A grade change date calculation is already in progress. Please wait for it to finish.")
        return redirect("membership_dashboard")

    if not acquire_type_lock("grade_progress_calc"):
        messages.warning(request, "A grade change date calculation is already in progress. Please wait.")
        return redirect("membership_dashboard")

    job_id = create_job("grade_progress_calc", filename="")
    logger.info(
        "compute_grade_progress queued job=%s user=%s",
        job_id, request.user, extra={"step": "grade_progress_calc"},
    )

    from App.services.grade_progress_calc import compute_all_grade_progress

    _start_thread(job_id, compute_all_grade_progress, b"", "grade_progress_calc", None, None)
    messages.info(request, f"Grade change date calculation started — tracking job {job_id[:8]}…")
    return redirect("membership_dashboard")


@requires_perm("membership.export")
def export_membership_excel(request):
    """Export Customer Membership section data to Excel — one section per
    request via ?section=, mirroring App/views/shop_detail.py::
    export_shop_detail_excel's single-view-branches-per-section pattern.

    section values: comparison | store | movers | trend | tier — these slugs
    are the contract the frontend download buttons must use.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    section = request.GET.get("section", "").strip()

    def _parse_id(raw):
        try:
            return int(raw)
        except (TypeError, ValueError):
            return None

    from_batch_id = _parse_id(request.GET.get("from_batch"))
    to_batch_id = _parse_id(request.GET.get("to_batch"))
    store = request.GET.get("store") or None
    grade = request.GET.get("grade") or None
    direction = request.GET.get("direction") or None
    shop = request.GET.get("shop") or None

    wb = Workbook()
    ws = wb.active

    HDR_FILL = PatternFill("solid", fgColor="366092")
    HDR_FONT = Font(color="FFFFFF", bold=True)

    def _hdr(ws, row_data, row=1):
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")

    if section == "comparison":
        from App.models.membership import MembershipSnapshotBatch
        if not to_batch_id or not MembershipSnapshotBatch.objects.filter(pk=to_batch_id).exists():
            # compare_batches() returns an all-zero row per grade for a
            # missing batch rather than an empty list (unlike the other 4
            # sections' get_*() helpers) — checked explicitly here so a
            # stale download link (snapshot deleted after page load) redirects
            # with an error instead of silently downloading an all-zero file.
            messages.error(request, "Select a snapshot to export.")
            return redirect("membership_dashboard")
        ws.title = "Grade Comparison"
        rows = compare_batches(from_batch_id, to_batch_id, store=store)
        ws['A1'] = "Members per Grade — Comparison"
        ws['A1'].font = Font(bold=True, size=13)
        _hdr(ws, ["Grade", "From", "To", "Diff", "% Change"], row=3)
        r = 4
        for row in rows:
            pct = f"{row['delta_pct']}%" if row['delta_pct'] is not None else "N/A"
            ws.cell(row=r, column=1, value=row['grade'])
            ws.cell(row=r, column=2, value=row['from_count'])
            ws.cell(row=r, column=3, value=row['to_count'])
            ws.cell(row=r, column=4, value=row['delta'])
            ws.cell(row=r, column=5, value=pct)
            r += 1

    elif section == "store":
        if not to_batch_id:
            messages.error(request, "Select a snapshot to export.")
            return redirect("membership_dashboard")
        ws.title = "By Registration Store"
        rows = get_grade_breakdown_by_store_comparison(from_batch_id, to_batch_id)
        if not rows:
            messages.error(request, "No data to export.")
            return redirect("membership_dashboard")
        header = ["Store"]
        for g in DISPLAY_GRADES:
            header += [f"{g} From", f"{g} To"]
        header += ["Total From", "Total To"]
        _hdr(ws, header, row=1)
        r = 2
        for row in rows:
            counts_by_grade = {c['grade']: c for c in row['counts']}
            vals = [row['store']]
            for g in DISPLAY_GRADES:
                c = counts_by_grade.get(g, {'from': 0, 'to': 0})
                vals += [c['from'], c['to']]
            vals += [row['total_from'], row['total_to']]
            for col, v in enumerate(vals, 1):
                ws.cell(row=r, column=col, value=v)
            r += 1

    elif section == "movers":
        if not from_batch_id or not to_batch_id:
            messages.error(request, "Select both a From and To snapshot to export.")
            return redirect("membership_dashboard")
        ws.title = "Grade Changes"
        # limit=None — the export must contain the FULL filtered result set,
        # not the on-screen partial's 500-row cap.
        rows, _total_count = get_grade_changes(
            from_batch_id, to_batch_id, store=store, grade=grade, direction=direction, limit=None,
        )
        if not rows:
            messages.error(request, "No grade changes to export for this selection.")
            return redirect("membership_dashboard")
        _hdr(ws, ["VIP ID", "Name", "Phone", "Store", "From Store", "To Store",
                  "From Grade", "To Grade", "Direction"], row=1)
        r = 2
        for row in rows:
            ws.cell(row=r, column=1, value=row['vip_id'])
            ws.cell(row=r, column=2, value=row.get('name') or '')
            ws.cell(row=r, column=3, value=row.get('phone') or '')
            ws.cell(row=r, column=4, value=row.get('registration_store') or '')
            ws.cell(row=r, column=5, value=row.get('from_store') or '')
            ws.cell(row=r, column=6, value=row.get('to_store') or '')
            ws.cell(row=r, column=7, value=row['from_grade'])
            ws.cell(row=r, column=8, value=row['to_grade'])
            ws.cell(row=r, column=9, value=row['direction'].capitalize())
            r += 1

    elif section == "trend":
        ws.title = "Grade Trend"
        series = get_all_batch_grade_series(store=store)
        if not series:
            messages.error(request, "No snapshot data to export.")
            return redirect("membership_dashboard")
        header = ["Snapshot Date", "Source"] + DISPLAY_GRADES
        _hdr(ws, header, row=1)
        r = 2
        for point in series:
            vals = [point['snapshot_date'], point['source']] + [
                point['counts'].get(g, 0) for g in DISPLAY_GRADES
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=r, column=col, value=v)
            r += 1

    elif section == "tier":
        ws.title = "Customer Tier Progress"
        # limit=None — the export must contain the FULL live-table result
        # set, not the on-screen partial's 500-row cap.
        rows, _total_count = get_live_customer_tier_table(
            grade_filter=grade, shop_filter=shop, sort='amount_to_next_tier', limit=None,
        )
        if not rows:
            messages.error(request, "No customers to export for this selection.")
            return redirect("membership_dashboard")
        _hdr(ws, ["VIP ID", "Name", "Phone", "Grade", "Annual Spend", "Annual Purchase Count",
                  "Points", "Next Grade", "Amount to Next Tier",
                  "Last Grade Change", "Direction", "Next Review Date",
                  "Purchases Needed to Avoid Downgrade"], row=1)
        r = 2
        for row in rows:
            is_ok = row.get('grade_progress_status') == 'ok'
            ws.cell(row=r, column=1, value=row['vip_id'])
            ws.cell(row=r, column=2, value=row.get('name') or '')
            ws.cell(row=r, column=3, value=row.get('phone') or '')
            ws.cell(row=r, column=4, value=row['grade'])
            ws.cell(row=r, column=5, value=float(row['annual_spend']))
            ws.cell(row=r, column=6, value=row['annual_purchase_count'])
            ws.cell(row=r, column=7, value=row.get('points') or 0)
            ws.cell(row=r, column=8, value=row.get('next_grade') or 'Max Tier')
            ws.cell(row=r, column=9, value=float(row['amount_to_next_tier']))
            ws.cell(row=r, column=10, value=row['last_grade_change_date'].isoformat() if is_ok and row.get('last_grade_change_date') else '—')
            ws.cell(row=r, column=11, value=row['change_direction'].capitalize() if is_ok and row.get('change_direction') else '—')
            ws.cell(row=r, column=12, value=row['next_check_date'].isoformat() if is_ok and row.get('next_check_date') else '—')
            ws.cell(row=r, column=13, value=row['purchases_needed_to_avoid_downgrade'] if is_ok and row.get('purchases_needed_to_avoid_downgrade') is not None else '—')
            r += 1

    else:
        messages.error(request, "No data selected to export.")
        return redirect("membership_dashboard")

    ts = datetime.now().strftime('%H%M%S')
    date_stamp = datetime.now().strftime('%Y%m%d')
    fn = f"membership_{section}_{date_stamp}_{ts}.xlsx"

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    wb.save(resp)
    return resp
