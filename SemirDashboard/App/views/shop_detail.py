"""App/views/shop_detail.py — Shop Detail page: Sales + Customer + Coupon analytics by shop."""
import json
import logging
from datetime import datetime

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.core.cache import cache

from App.permissions import requires_perm, user_has_perm
from App.analytics.tab_functions import (
    get_shop_detail_sales_data,
    get_shop_detail_customer_data,
    get_shop_detail_coupon_data,
)
from App.models import SalesTransaction, Customer, Coupon, CouponCampaign
from App.views.view_utils import parse_date

logger = logging.getLogger(__name__)

_DROPDOWN_TTL = 300  # 5 min cache for dropdown lists


def _ajax_perm_check(request, codename):
    """Return an error HttpResponse for AJAX partial views, or None if OK."""
    if not request.user.is_authenticated:
        return HttpResponse(
            '<div class="alert alert-warning m-0">Session expired. '
            '<a href="/login/">Log in again</a>.</div>',
            status=401,
        )
    if not user_has_perm(request.user, codename):
        return HttpResponse(
            '<div class="alert alert-danger m-0">Permission denied.</div>',
            status=403,
        )
    return None


def _get_dropdown_options():
    """Return (sales_shops, customer_shops, coupon_shops, campaigns) — cached 5 min."""
    key = "shop_detail_dropdowns"
    cached = cache.get(key)
    if cached:
        return cached

    sales_shops = sorted(
        SalesTransaction.objects
        .exclude(shop_name__isnull=True).exclude(shop_name='')
        .values_list('shop_name', flat=True).distinct()
    )
    customer_shops = sorted(
        Customer.objects
        .exclude(registration_store__isnull=True).exclude(registration_store='')
        .values_list('registration_store', flat=True).distinct()
    )
    coupon_shops = sorted(
        Coupon.objects
        .exclude(using_shop__isnull=True).exclude(using_shop='')
        .values_list('using_shop', flat=True).distinct()
    )
    campaigns = list(CouponCampaign.objects.values("id", "name", "prefix"))
    for c in campaigns:
        c["prefix_list"] = [p.strip() for p in (c["prefix"] or "").split(",") if p.strip()]

    result = (sales_shops, customer_shops, coupon_shops, campaigns)
    cache.set(key, result, _DROPDOWN_TTL)
    return result

QUICK_BTNS = [
    ("Last 7 Days", 7),
    ("Last 30 Days", 30),
    ("Last 90 Days", 90),
    ("Last Year", 365),
]
YEAR_BTNS = [2024, 2025, 2026]


@requires_perm("page_shop_detail")
def shop_detail(request):
    """
    Shop Detail page: 3 sections each filtered by a specific shop.
    - Sales analytics: filter by shop_name from SalesTransaction (direct DB query)
    - Customer analytics: filter by registration_store from Customer (cached BD raw, store_filter)
    - Coupon analytics: filter by using_shop from Coupon (direct DB query + prefix/campaign)
    """
    start_date = request.GET.get("start_date", "")
    end_date   = request.GET.get("end_date", "")

    # ── Dropdown options (cached 5 min) ───────────────────────────────────────
    sales_shops, customer_shops, coupon_shops, campaigns = _get_dropdown_options()

    return render(request, "shop_detail.html", {
        "start_date":        start_date,
        "end_date":          end_date,
        "quick_btns":        QUICK_BTNS,
        "year_btns":         YEAR_BTNS,
        "sales_shops":       sales_shops,
        "customer_shops":    customer_shops,
        "coupon_shops":      coupon_shops,
        "campaigns_json":    json.dumps(campaigns),
        "currency":          "VND",
    })


def _safe_date(val):
    """Parse date string silently (no messages framework)."""
    if not val:
        return None
    try:
        return datetime.strptime(val.strip(), "%Y-%m-%d").date()
    except ValueError:
        return None


def shop_detail_sales_partial(request):
    """AJAX: render sales section data for one shop."""
    err = _ajax_perm_check(request, "page_shop_detail")
    if err:
        return err
    shop_name  = request.GET.get("shop", "").strip()
    start_date = request.GET.get("start_date", "")
    end_date   = request.GET.get("end_date", "")
    if not shop_name:
        return HttpResponse('<div class="no-data-msg text-muted py-3">Select a shop to view sales analytics.</div>')
    date_from = _safe_date(start_date)
    date_to   = _safe_date(end_date)
    logger.info("partial sales: shop=%s from=%s to=%s user=%s", shop_name, date_from, date_to, request.user)
    data = get_shop_detail_sales_data(shop_name, date_from=date_from, date_to=date_to)
    return render(request, "shop_detail/_sales_partial.html", {
        "data": data, "shop_name": shop_name,
        "start_date": start_date, "end_date": end_date,
        "currency": "VND",
    })


def shop_detail_customer_partial(request):
    """AJAX: render customer section data for one store."""
    err = _ajax_perm_check(request, "page_shop_detail")
    if err:
        return err
    store_name = request.GET.get("shop", "").strip()
    start_date = request.GET.get("start_date", "")
    end_date   = request.GET.get("end_date", "")
    if not store_name:
        return HttpResponse('<div class="no-data-msg text-muted py-3">Select a store to view customer analytics.</div>')
    logger.info("partial customer: store=%s from=%s to=%s user=%s", store_name, start_date, end_date, request.user)
    data = get_shop_detail_customer_data(store_name, start_date=start_date, end_date=end_date)
    return render(request, "shop_detail/_customer_partial.html", {
        "data": data, "store_name": store_name,
        "start_date": start_date, "end_date": end_date,
        "currency": "VND",
    })


def shop_detail_coupon_partial(request):
    """AJAX: render coupon section data for one shop."""
    err = _ajax_perm_check(request, "page_shop_detail")
    if err:
        return err
    shop_name        = request.GET.get("shop", "").strip()
    start_date       = request.GET.get("start_date", "")
    end_date         = request.GET.get("end_date", "")
    coupon_prefix    = request.GET.get("coupon_prefix", "").strip()
    coupon_campaign_id = request.GET.get("coupon_campaign", "").strip()
    if not shop_name:
        return HttpResponse('<div class="no-data-msg text-muted py-3">Select a shop to view coupon analytics.</div>')
    date_from = _safe_date(start_date)
    date_to   = _safe_date(end_date)
    # Resolve campaign → prefix
    effective_prefix = coupon_prefix
    if coupon_campaign_id:
        try:
            camp = CouponCampaign.objects.get(pk=int(coupon_campaign_id))
            effective_prefix = camp.prefix
        except (CouponCampaign.DoesNotExist, ValueError):
            pass
    logger.info("partial coupon: shop=%s prefix=%s from=%s to=%s user=%s",
                shop_name, effective_prefix or '', date_from, date_to, request.user)
    data = get_shop_detail_coupon_data(
        shop_name, date_from=date_from, date_to=date_to,
        coupon_id_prefix=effective_prefix or None,
    )
    return render(request, "shop_detail/_coupon_partial.html", {
        "data": data, "shop_name": shop_name,
        "start_date": start_date, "end_date": end_date,
        "coupon_prefix": coupon_prefix,
        "coupon_campaign_id": coupon_campaign_id,
        "currency": "VND",
    })


@requires_perm("download_shop_detail")
def export_shop_detail_excel(request):
    """Export Shop Detail data to Excel."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    start_date         = request.GET.get("start_date", "")
    end_date           = request.GET.get("end_date", "")
    section            = request.GET.get("section", "").strip()
    sales_shop         = request.GET.get("sales_shop", "").strip()
    customer_shop      = request.GET.get("customer_shop", "").strip()
    coupon_shop        = request.GET.get("coupon_shop", "").strip()
    coupon_prefix      = request.GET.get("coupon_prefix", "").strip()
    coupon_campaign_id = request.GET.get("coupon_campaign", "").strip()

    date_from = parse_date(start_date, "start date", request)
    date_to   = parse_date(end_date,   "end date",   request)

    # Resolve campaign prefix
    effective_prefix = coupon_prefix
    if coupon_campaign_id:
        try:
            camp = CouponCampaign.objects.get(pk=int(coupon_campaign_id))
            effective_prefix = camp.prefix
        except (CouponCampaign.DoesNotExist, ValueError):
            pass

    wb = Workbook()
    ws = wb.active

    HDR_FILL = PatternFill("solid", fgColor="366092")
    HDR_FONT = Font(color="FFFFFF", bold=True)

    def _hdr(ws, row_data, row=1):
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = HDR_FILL
            c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center")

    if section == "sales" and sales_shop:
        ws.title = "Sales by Shop"
        sh = get_shop_detail_sales_data(sales_shop, date_from=date_from, date_to=date_to)

        ws['A1'] = f"Sales Analytics — {sales_shop}"
        ws['A1'].font = Font(bold=True, size=13)
        if date_from or date_to:
            ws['A2'] = f"Period: {date_from or 'all'} → {date_to or 'all'}"

        if sh:
            _hdr(ws, ["Shop","Active","Returning","Return Rate","INV(RET)","AMT(RET)","Total INV","Total Amount"], row=4)
            ws.append([
                sh['shop_name'], sh['total_customers'], sh['returning_customers'],
                f"{sh['return_rate']}%", sh['returning_invoices'],
                float(sh['returning_amount']), sh['total_invoices_with_vip0'],
                float(sh['total_amount_with_vip0']),
            ])
            r = 7
            for period_label, rows, key in [
                ("By Season", sh.get('by_session', []), 'session'),
                ("By Month",  sh.get('by_month', []),   'month'),
            ]:
                ws.cell(row=r, column=1, value=period_label).font = Font(bold=True)
                r += 1
                _hdr(ws, [period_label.split()[-1],"Active","Returning","Return Rate",
                          "INV(RET)","AMT(RET)","Total INV","Total Amount"], row=r)
                r += 1
                for row in rows:
                    ws.append([row[key], row['total_customers'], row['returning_customers'],
                               f"{row['return_rate']}%", row['returning_invoices'],
                               float(row['returning_amount']), row['total_invoices_with_vip0'],
                               float(row['total_amount_with_vip0'])])
                    r += 1
                r += 1
            ws.cell(row=r, column=1, value="By Week").font = Font(bold=True)
            r += 1
            _hdr(ws, ["Week","Active","Returning","Return Rate","INV(RET)","AMT(RET)","Total INV","Total Amount"], row=r)
            r += 1
            for w in sh.get('by_week', []):
                ws.append([w['week_label'], w['total_customers'], w['returning_customers'],
                           f"{w['return_rate']}%", w['returning_invoices'],
                           float(w['returning_amount']), w['total_invoices_with_vip0'],
                           float(w['total_amount_with_vip0'])])

    elif section == "customer" and customer_shop:
        ws.title = "Customer by Shop"
        cd = get_shop_detail_customer_data(customer_shop, start_date=start_date, end_date=end_date)
        summary = cd.get('summary') if cd else None

        ws['A1'] = f"Customer Analytics — {customer_shop}"
        ws['A1'].font = Font(bold=True, size=13)

        if summary:
            _hdr(ws, ["Shop","POS(INV)","POS(NO INV)","POS Total","POS Only",
                      "New CNV","CNV Only","Zalo App","%App","Zalo OA","%OA"], row=3)
            ws.append([summary['label'], summary.get('new_pos_inv'), summary.get('new_pos_no_inv'),
                       summary.get('new_pos'), summary.get('new_pos_only'), summary.get('new_cnv'),
                       summary.get('new_cnv_only'), summary.get('zalo_app'),
                       f"{summary.get('zalo_app_pct',0)}%",
                       summary.get('zalo_oa'), f"{summary.get('zalo_oa_pct',0)}%"])

        if cd:
            r = 7
            for period_label, rows in [
                ("By Season", cd.get('by_season', [])),
                ("By Month",  cd.get('by_month',  [])),
                ("By Week",   cd.get('by_week',   [])),
            ]:
                ws.cell(row=r, column=1, value=period_label).font = Font(bold=True)
                r += 1
                _hdr(ws, ["Period","POS(INV)","POS(NO INV)","POS Total","POS Only",
                          "New CNV","CNV Only","Zalo App","%App","Zalo OA","%OA"], row=r)
                r += 1
                for row in rows:
                    ws.append([row['label'], row.get('new_pos_inv'), row.get('new_pos_no_inv'),
                               row.get('new_pos'), row.get('new_pos_only'), row.get('new_cnv'),
                               row.get('new_cnv_only'), row.get('zalo_app'),
                               f"{row.get('zalo_app_pct',0)}%",
                               row.get('zalo_oa'), f"{row.get('zalo_oa_pct',0)}%"])
                    r += 1
                r += 1

    elif section == "coupon" and coupon_shop:
        ws.title = "Coupon by Shop"
        cd = get_shop_detail_coupon_data(coupon_shop, date_from=date_from, date_to=date_to,
                                         coupon_id_prefix=effective_prefix or None)
        ws['A1'] = f"Coupon Analytics — {coupon_shop}"
        ws['A1'].font = Font(bold=True, size=13)
        if effective_prefix:
            ws['A2'] = f"Prefix filter: {effective_prefix}"

        _hdr(ws, ["Metric","All-Time","Period"], row=4)
        at, pd = cd.get('all_time', {}), cd.get('period', {})
        for label, key in [
            ("Total Coupons", "total"), ("Used", "used"), ("Unused", "unused"),
            ("Usage Rate %", "usage_rate"),
            ("Total Amount (VND)", "total_amount"),
            ("Coupon Amount (VND)", "total_coupon_amount"),
        ]:
            ws.append([label, at.get(key, 0), pd.get(key, 0)])

        ws2 = wb.create_sheet("Detail")
        _hdr(ws2, ["Coupon ID","Creator","Face Value","Using Shop","Using Date",
                   "VIP ID","Customer Name","Phone","Sales Date","Invoice Shop",
                   "Amount (VND)","Coupon Amt (VND)","Note",
                   "CNV ID","CNV Points","CNV Total Pts"], row=1)
        for d in cd.get('details', []):
            ws2.append([
                d['coupon_id'], d['creator'], d['face_value_display'],
                d['using_shop'], str(d['using_date']) if d['using_date'] else '',
                d['vip_id'], d['customer_name'], d['customer_phone'],
                str(d['sales_day']) if d['sales_day'] else '',
                d['inv_shop'], float(d['amount']), float(d['coupon_amount']),
                d['note'] or '', d['cnv_id'] or '',
                d['cnv_points'] if d['cnv_points'] != '' else '',
                d['cnv_total_points'] if d['cnv_total_points'] != '' else '',
            ])
    else:
        messages.error(request, "No data selected to export.")
        return redirect("shop_detail")

    ts = datetime.now().strftime('%H%M%S')
    period = f"{date_from}_{date_to}" if date_from and date_to else datetime.now().strftime('%Y%m%d')
    shop_slug = (sales_shop or customer_shop or coupon_shop).replace(' ', '_')[:30]
    fn = f"shop_detail_{section}_{shop_slug}_{period}_{ts}.xlsx"

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    wb.save(resp)
    return resp
