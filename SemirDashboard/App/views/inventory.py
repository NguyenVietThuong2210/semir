"""App/views/inventory.py — Global Inventory Analytics page."""
import csv
import io
import logging
from datetime import datetime

from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse

from App.permissions import requires_perm
from App.analytics.inventory_functions import get_inventory_overview, get_product_prefix_detail
from App.models import InventorySnapshot

logger = logging.getLogger(__name__)

SHOP_GROUPS = [
    ("",        "All Shops"),
    ("semir",   "Semir"),
    ("bala",    "Bala Bala"),
    ("others",  "Others"),
]


def _get_filter_options():
    """Return distinct years and seasons available in the inventory, for dropdowns."""
    years = list(
        InventorySnapshot.objects
        .order_by('-year').values_list('year', flat=True).distinct()
        .exclude(year__isnull=True)
    )
    seasons = list(
        InventorySnapshot.objects
        .order_by('season').values_list('season', flat=True).distinct()
        .exclude(season='')
    )
    return years, seasons


@requires_perm("inventory.view")
def inventory_dashboard(request):
    shop_group     = request.GET.get("shop_group", "")
    year_str       = request.GET.get("year", "")
    season         = request.GET.get("season", "")
    product_prefix = request.GET.get("product_prefix", "").strip()

    year = int(year_str) if year_str.isdigit() else None

    data = get_inventory_overview(
        shop_group=shop_group or None,
        year=year,
        season=season or None,
    )

    if not data:
        messages.info(request, "No inventory data found. Please upload an inventory file first.")
        return redirect("upload_inventory")

    prefix_data = {}
    if product_prefix:
        prefix_data = get_product_prefix_detail(
            prefix=product_prefix,
            shop_group=shop_group or None,
            year=year,
            season=season or None,
        )

    years_available, seasons_available = _get_filter_options()

    return render(request, "inventory/dashboard.html", {
        "data":               data,
        "for_sale":           data.get("for_sale", {}),
        "gifts":              data.get("gifts", {}),
        "dead_year":          data.get("dead_year_threshold"),
        "shop_group":         shop_group,
        "shop_groups":        SHOP_GROUPS,
        "year_filter":        year,
        "season_filter":      season,
        "years_available":    years_available,
        "seasons_available":  seasons_available,
        "product_prefix":     product_prefix,
        "prefix_data":        prefix_data,
    })


@requires_perm("inventory.export")
def export_inventory_excel(request):
    """Export full inventory overview to Excel (for_sale by-shop summary + dead stock list)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    shop_group = request.GET.get("shop_group", "")
    year_str   = request.GET.get("year", "")
    season     = request.GET.get("season", "")
    year = int(year_str) if year_str.isdigit() else None

    data = get_inventory_overview(shop_group=shop_group or None, year=year, season=season or None)
    if not data:
        messages.error(request, "No inventory data to export.")
        return redirect("inventory_dashboard")

    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", fgColor="1a3a5c")
    hdr_font = Font(bold=True, color="FFFFFF")

    def _write_sheet(ws, rows, columns):
        ws.append([c[0] for c in columns])
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row.get(c[1], "") for c in columns])
        for i, col in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = col[2]

    # Sheet 1: By Shop (for sale)
    ws1 = wb.active
    ws1.title = "By Shop (For Sale)"
    by_shop = data.get("for_sale", {}).get("by_shop", [])
    _write_sheet(ws1, by_shop, [
        ("Shop", "shop_name", 30),
        ("SKU Lines", "sku_lines", 12),
        ("On Hand Qty", "on_hand_qty", 14),
        ("In Transit", "in_transit_qty", 14),
        ("Total Qty", "total_qty", 12),
        ("Inv Value (VND)", "inv_value", 18),
    ])

    # Sheet 2: By Brand (for sale)
    ws2 = wb.create_sheet("By Brand (For Sale)")
    by_brand = data.get("for_sale", {}).get("by_brand", [])
    _write_sheet(ws2, by_brand, [
        ("Brand", "brand", 20),
        ("SKU Lines", "sku_lines", 12),
        ("On Hand Qty", "on_hand_qty", 14),
        ("Inv Value (VND)", "inv_value", 18),
    ])

    # Sheet 3: Dead Stock
    ws3 = wb.create_sheet("Dead Stock SKUs")
    dead_top = data.get("for_sale", {}).get("dead", {}).get("top", [])
    _write_sheet(ws3, dead_top, [
        ("Shop", "shop_name", 28),
        ("Product Code", "product_code", 14),
        ("Product Name", "product_name", 30),
        ("Color", "color", 12),
        ("Size", "size", 8),
        ("Brand", "brand", 14),
        ("Year", "year", 8),
        ("Season", "season", 10),
        ("Tag Price", "tag_price", 12),
        ("Qty", "qty", 8),
        ("Value (VND)", "value", 16),
    ])

    # Sheet 4: Gifts (non-commercial)
    ws4 = wb.create_sheet("By Shop (Gifts)")
    gift_by_shop = data.get("gifts", {}).get("by_shop", [])
    _write_sheet(ws4, gift_by_shop, [
        ("Shop", "shop_name", 30),
        ("SKU Lines", "sku_lines", 12),
        ("On Hand Qty", "on_hand_qty", 14),
        ("In Transit", "in_transit_qty", 14),
    ])

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    slug = f"_{shop_group}" if shop_group else ""
    slug += f"_{year}" if year else ""
    slug += f"_{season}" if season else ""
    fname = f"inventory{slug}_{ts}.xlsx"

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    wb.save(resp)
    return resp


@requires_perm("inventory.export")
def export_inventory_dead_stock(request):
    """Download dead stock SKUs as CSV."""
    shop_group = request.GET.get("shop_group", "")
    year_str   = request.GET.get("year", "")
    season     = request.GET.get("season", "")
    year = int(year_str) if year_str.isdigit() else None

    data = get_inventory_overview(
        shop_group=shop_group or None,
        year=year,
        season=season or None,
    )
    dead_top = (data.get("for_sale") or {}).get("dead", {}).get("top", [])

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([
        "Shop", "Product Code", "Product Name", "Product Name VN",
        "Color", "Size", "Category L1", "Category L2", "Category L3", "Gender",
        "Brand", "Year", "Season", "Tag Price", "Qty", "Value (VND)",
    ])
    for row in dead_top:
        writer.writerow([
            row.get("shop_name", ""),
            row.get("product_code", ""),
            row.get("product_name", ""),
            row.get("product_name_vn", ""),
            row.get("color", ""),
            row.get("size", ""),
            row.get("category_l1", ""),
            row.get("category_l2", ""),
            row.get("category_l3", ""),
            row.get("gender", ""),
            row.get("brand", ""),
            row.get("year", ""),
            row.get("season", ""),
            row.get("tag_price", ""),
            row.get("qty", 0),
            row.get("value", 0),
        ])

    response = HttpResponse(buf.getvalue(), content_type="text/csv; charset=utf-8-sig")
    fname = f"dead_stock_{shop_group or 'all'}.csv"
    response["Content-Disposition"] = f'attachment; filename="{fname}"'
    return response
