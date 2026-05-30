"""
tests/test_pages.py — Page-render smoke tests for all uncovered GET routes.

Covers: home, formulas, upload pages, customer-detail, users, admin-logs,
        coupons/chart, coupons/campaigns, cnv/sync-status, shop-detail/export,
        analytics/chart/export, coupons/chart/export, cnv/customer-chart/export,
        upload jobs JSON endpoints, POST trigger endpoints.

No fixture data is loaded here — every page must handle an empty DB gracefully.
Period-filter pages are tested in both variants: all-time + 2025.

Run:
  cd SemirDashboard && python manage.py test tests.test_pages -v 2
"""
import json
import time

from django.contrib.auth.models import User
from django.urls import reverse

from tests.base import SnapshotTestCase

PERIOD_2025_PARAMS = "?start_date=2025-01-01&end_date=2025-12-31"


class PageRenderTest(SnapshotTestCase):
    """
    Smoke tests: every GET route must return HTTP 200.
    No fixture data — tests graceful empty-state rendering.
    Period-filter routes tested in both all-time and 2025 variants.
    """

    @classmethod
    def setUpTestData(cls):
        cls.superuser = User.objects.create_superuser(
            username="testadmin", password="testpass", email="test@test.com"
        )

    def setUp(self):
        self.client.force_login(self.superuser)

    # ── Static / home pages ───────────────────────────────────────────────────

    def test_home_200(self):
        t = self.timer("home")
        r = self.client.get(reverse("home"))
        t.checkpoint("GET /")
        t.report()
        self.assertEqual(r.status_code, 200)

    def test_formulas_200(self):
        r = self.client.get(reverse("formulas"))
        self.assertEqual(r.status_code, 200)

    # ── Upload pages (GET only — form render, no file submitted) ─────────────

    def test_upload_customers_200(self):
        r = self.client.get(reverse("upload_customers"))
        self.assertEqual(r.status_code, 200)

    def test_upload_sales_200(self):
        r = self.client.get(reverse("upload_sales"))
        self.assertEqual(r.status_code, 200)

    def test_upload_coupons_200(self):
        r = self.client.get(reverse("upload_coupons"))
        self.assertEqual(r.status_code, 200)

    def test_upload_inventory_200(self):
        r = self.client.get(reverse("upload_inventory"))
        self.assertEqual(r.status_code, 200)

    def test_product_dashboard_empty_200(self):
        # With no sale detail data, redirects to upload_sales — follow it.
        r = self.client.get(reverse("product_dashboard"), follow=True)
        self.assertEqual(r.status_code, 200)

    def test_product_tab_smoke(self):
        r = self.client.get(
            reverse("product_tab", kwargs={"tab": "brand"}),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 302])

    def test_shop_detail_inventory_partial_smoke(self):
        r = self.client.get(
            reverse("shop_detail_inventory_partial"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(r.status_code, 200)

    def test_upload_used_points_redirect(self):
        # upload_used_points is POST-only; GET redirects to upload_customers
        r = self.client.get(reverse("upload_used_points"), follow=True)
        self.assertEqual(r.status_code, 200)

    def test_upload_jobs_list_json(self):
        r = self.client.get(reverse("upload_jobs_list"))
        self.assertEqual(r.status_code, 200)
        data = json.loads(r.content)
        self.assertIn("jobs", data)

    # ── Customer detail ───────────────────────────────────────────────────────

    def test_customer_detail_empty_200(self):
        # No search params → show search form only
        r = self.client.get(reverse("customer_detail"))
        self.assertEqual(r.status_code, 200)

    def test_customer_detail_not_found_200(self):
        # VIP ID that doesn't exist → "no customer found" state, still 200
        r = self.client.get(reverse("customer_detail") + "?vip_id=XXXXNOTEXIST")
        self.assertEqual(r.status_code, 200)

    # ── Admin pages ───────────────────────────────────────────────────────────

    def test_users_200(self):
        r = self.client.get(reverse("user_management"))
        self.assertEqual(r.status_code, 200)

    def test_admin_logs_200(self):
        # Reads log files — graceful if logs/ is empty
        r = self.client.get(reverse("admin_logs"))
        self.assertEqual(r.status_code, 200)

    # ── Analytics pages (period-filter) ──────────────────────────────────────

    def test_analytics_dashboard_200(self):
        # With no fixture data, the view redirects to upload_sales — follow it.
        t = self.timer("analytics_dashboard_alltime")
        r = self.client.get(reverse("analytics_dashboard"), follow=True)
        t.checkpoint("GET /analytics/ (follow)")
        t.report()
        self.assertEqual(r.status_code, 200)

    def test_analytics_dashboard_2025_200(self):
        r = self.client.get(reverse("analytics_dashboard") + PERIOD_2025_PARAMS, follow=True)
        self.assertEqual(r.status_code, 200)

    def test_analytics_chart_200(self):
        r = self.client.get(reverse("analytics_chart"))
        self.assertEqual(r.status_code, 200)

    def test_analytics_chart_2025_200(self):
        r = self.client.get(reverse("analytics_chart") + PERIOD_2025_PARAMS)
        self.assertEqual(r.status_code, 200)

    def test_analytics_tab_smoke(self):
        # AJAX tab — representative tab: season
        r = self.client.get(
            reverse("analytics_tab", kwargs={"tab": "season"}),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 204])

    # ── Coupon pages (period-filter) ──────────────────────────────────────────

    def test_coupon_dashboard_200(self):
        t = self.timer("coupon_dashboard_alltime")
        r = self.client.get(reverse("coupon_dashboard"))
        t.checkpoint("GET /coupons/")
        t.report()
        self.assertEqual(r.status_code, 200)

    def test_coupon_dashboard_2025_200(self):
        r = self.client.get(reverse("coupon_dashboard") + PERIOD_2025_PARAMS)
        self.assertEqual(r.status_code, 200)

    def test_coupon_tab_smoke(self):
        # 'shop' tab is rejected by coupon_tab (handled by coupon_dashboard); use 'detail'
        r = self.client.get(
            reverse("coupon_tab", kwargs={"tab": "detail"}),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 204])

    def test_coupon_chart_alltime_200(self):
        t = self.timer("coupon_chart_alltime")
        r = self.client.get(reverse("coupon_chart"))
        t.checkpoint("GET /coupons/chart/")
        t.report()
        self.assertEqual(r.status_code, 200)

    def test_coupon_chart_period_2025_200(self):
        t = self.timer("coupon_chart_2025")
        r = self.client.get(reverse("coupon_chart") + PERIOD_2025_PARAMS)
        t.checkpoint("GET /coupons/chart/ with 2025 filter")
        t.report()
        self.assertEqual(r.status_code, 200)

    def test_coupon_campaigns_200(self):
        r = self.client.get(reverse("manage_campaigns"))
        self.assertEqual(r.status_code, 200)

    # ── Upload jobs ───────────────────────────────────────────────────────────

    def test_upload_job_status_404(self):
        # Non-existent job_id → 404 JSON, no crash
        r = self.client.get(reverse("upload_job_status", kwargs={"job_id": "nonexistent-job-id"}))
        self.assertEqual(r.status_code, 404)

    # ── Auth pages ────────────────────────────────────────────────────────────

    def test_login_200(self):
        self.client.logout()
        r = self.client.get(reverse("login"))
        self.assertEqual(r.status_code, 200)

    def test_register_200(self):
        # register_view requires @login_required — superuser stays logged in
        r = self.client.get(reverse("register"))
        self.assertEqual(r.status_code, 200)

    # ── CNV pages ─────────────────────────────────────────────────────────────

    def test_cnv_customer_analytics_200(self):
        t = self.timer("cnv_customer_analytics_alltime")
        r = self.client.get(reverse("cnv:customer_analytics"))
        t.checkpoint("GET /cnv/customer-analytics/")
        t.report()
        self.assertEqual(r.status_code, 200)

    def test_cnv_customer_analytics_2025_200(self):
        r = self.client.get(reverse("cnv:customer_analytics") + PERIOD_2025_PARAMS)
        self.assertEqual(r.status_code, 200)

    def test_cnv_customer_tab_smoke(self):
        r = self.client.get(
            reverse("cnv:customer_tab", kwargs={"tab": "bd_season"}),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 204])

    def test_cnv_customer_chart_200(self):
        r = self.client.get(reverse("cnv:customer_chart"))
        self.assertEqual(r.status_code, 200)

    def test_cnv_customer_chart_2025_200(self):
        r = self.client.get(reverse("cnv:customer_chart") + PERIOD_2025_PARAMS)
        self.assertEqual(r.status_code, 200)

    def test_cnv_sync_status_200(self):
        t = self.timer("cnv_sync_status")
        r = self.client.get(reverse("cnv:sync_status"))
        t.checkpoint("GET /cnv/sync-status/")
        t.report()
        self.assertEqual(r.status_code, 200)

    # ── POST trigger endpoints (smoke — just check no crash) ──────────────────

    def test_trigger_sync_post(self):
        # POST to trigger sync — expect redirect or JSON, not 500
        r = self.client.post(
            reverse("cnv:trigger_sync"),
            data={"sync_type": "customers"},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 302, 400, 429])

    def test_trigger_zalo_sync_post(self):
        r = self.client.post(
            reverse("cnv:trigger_zalo_sync"),
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 302, 400, 429])

    def test_sync_cnv_points_post(self):
        # POST with empty phones list — expect JSON response, not 500
        r = self.client.post(
            reverse("cnv:sync_cnv_points"),
            data={"phones": ""},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertIn(r.status_code, [200, 302, 400, 403, 429])


class ExportSmokeTest(SnapshotTestCase):
    """
    Export endpoint smoke tests — verify workbook is returned (Content-Type: Excel).
    These need fixture data so they share a setUpTestData with full import.
    Period-filter exports tested in both all-time and 2025 variants.
    """

    @classmethod
    def setUpTestData(cls):
        import io
        from App.services import (
            process_customer_file, process_sales_file, process_coupon_file,
            process_inventory_file, process_sale_detail_file,
        )
        from tests.base import INPUT_DIR

        cls.superuser = User.objects.create_superuser(
            username="exportadmin", password="testpass", email="export@test.com"
        )

        def _named(path):
            with open(path, "rb") as f:
                data = f.read()
            class _N(io.BytesIO):
                pass
            obj = _N(data)
            obj.name = path.name
            return obj

        customer_file  = INPUT_DIR / "customer.xlsx"
        sale_files     = [INPUT_DIR / "Sale 2024.xlsx", INPUT_DIR / "Sale 2025.xlsx", INPUT_DIR / "Sale 2026.xlsx"]
        coupon_file    = INPUT_DIR / "coupon_1 (1).xlsx"
        inventory_file = INPUT_DIR / "inventory.xlsx"
        sd_file        = INPUT_DIR / "sale detail.xlsx"

        if customer_file.exists():
            process_customer_file(_named(customer_file))
        for path in sale_files:
            if path.exists():
                process_sales_file(_named(path))
        if coupon_file.exists():
            process_coupon_file(_named(coupon_file))
        if inventory_file.exists():
            process_inventory_file(_named(inventory_file))
        if sd_file.exists():
            process_sale_detail_file(_named(sd_file))

    def setUp(self):
        self.client.force_login(self.superuser)

    EXCEL_CONTENT_TYPES = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    )

    def _assert_excel(self, response):
        self.assertEqual(response.status_code, 200)
        ct = response.get("Content-Type", "")
        self.assertTrue(
            any(ct.startswith(t) for t in self.EXCEL_CONTENT_TYPES),
            f"Expected Excel Content-Type, got: {ct}",
        )

    # ── Analytics chart export [period-filter] ────────────────────────────────

    def test_analytics_chart_export_alltime(self):
        t = self.timer("analytics_chart_export_alltime")
        r = self.client.get(reverse("export_sales_chart_excel"))
        t.checkpoint("GET /analytics/chart/export/ all-time")
        t.report()
        self._assert_excel(r)

    def test_analytics_chart_export_period_2025(self):
        t = self.timer("analytics_chart_export_2025")
        r = self.client.get(reverse("export_sales_chart_excel") + PERIOD_2025_PARAMS)
        t.checkpoint("GET /analytics/chart/export/ 2025 filter")
        t.report()
        self._assert_excel(r)

    # ── Coupon chart export [period-filter] ───────────────────────────────────

    def test_coupon_chart_export_alltime(self):
        t = self.timer("coupon_chart_export_alltime")
        r = self.client.get(reverse("export_coupon_chart_excel"))
        t.checkpoint("GET /coupons/chart/export/ all-time")
        t.report()
        self._assert_excel(r)

    def test_coupon_chart_export_period_2025(self):
        t = self.timer("coupon_chart_export_2025")
        r = self.client.get(reverse("export_coupon_chart_excel") + PERIOD_2025_PARAMS)
        t.checkpoint("GET /coupons/chart/export/ 2025 filter")
        t.report()
        self._assert_excel(r)

    # ── Shop detail export [period-filter] ────────────────────────────────────

    def test_shop_detail_export_alltime(self):
        from App.models import SalesTransaction
        shop = SalesTransaction.objects.order_by().values_list("shop_name", flat=True).first()
        if not shop:
            self.skipTest("No sales data in fixture")
        t = self.timer("shop_detail_export_alltime")
        from urllib.parse import urlencode
        params = urlencode({"section": "sales", "sales_shop": shop})
        r = self.client.get(reverse("export_shop_detail_excel") + f"?{params}")
        t.checkpoint(f"GET /shop-detail/export/ shop={shop}")
        t.report()
        self._assert_excel(r)

    def test_shop_detail_export_period_2025(self):
        from App.models import SalesTransaction
        shop = SalesTransaction.objects.order_by().values_list("shop_name", flat=True).first()
        if not shop:
            self.skipTest("No sales data in fixture")
        t = self.timer("shop_detail_export_2025")
        from urllib.parse import urlencode
        params = urlencode({"section": "sales", "sales_shop": shop,
                            "start_date": "2025-01-01", "end_date": "2025-12-31"})
        r = self.client.get(reverse("export_shop_detail_excel") + f"?{params}")
        t.checkpoint(f"GET /shop-detail/export/ shop={shop} 2025 filter")
        t.report()
        self._assert_excel(r)

    # ── CNV customer chart export [period-filter] ─────────────────────────────

    def test_cnv_customer_chart_export_alltime(self):
        t = self.timer("cnv_chart_export_alltime")
        r = self.client.get(reverse("cnv:export_customer_chart_excel"))
        t.checkpoint("GET /cnv/customer-chart/export/ all-time")
        t.report()
        self._assert_excel(r)

    def test_cnv_customer_chart_export_period_2025(self):
        t = self.timer("cnv_chart_export_2025")
        r = self.client.get(reverse("cnv:export_customer_chart_excel") + PERIOD_2025_PARAMS)
        t.checkpoint("GET /cnv/customer-chart/export/ 2025 filter")
        t.report()
        self._assert_excel(r)

    # ── Analytics full export [period-filter] ─────────────────────────────────

    def test_export_analytics_alltime(self):
        t = self.timer("export_analytics_alltime")
        r = self.client.get(reverse("export_analytics"))
        t.checkpoint("GET /analytics/export/ all-time")
        t.report()
        self._assert_excel(r)

    def test_export_analytics_period_2025(self):
        t = self.timer("export_analytics_2025")
        r = self.client.get(reverse("export_analytics") + PERIOD_2025_PARAMS)
        t.checkpoint("GET /analytics/export/ 2025 filter")
        t.report()
        self._assert_excel(r)

    # ── Coupon full export ────────────────────────────────────────────────────

    def test_export_coupons_alltime(self):
        t = self.timer("export_coupons_alltime")
        r = self.client.get(reverse("export_coupons"))
        t.checkpoint("GET /coupons/export/ all-time")
        t.report()
        self._assert_excel(r)

    # ── CNV customer analytics export ─────────────────────────────────────────

    def test_export_customer_analytics_alltime(self):
        t = self.timer("export_customer_analytics_alltime")
        r = self.client.get(reverse("cnv:export_customer_analytics"))
        t.checkpoint("GET /cnv/export-customer-analytics/ all-time")
        t.report()
        self._assert_excel(r)

    def test_export_customer_analytics_zalo_tab(self):
        """?tab=zalo should return Excel with Zalo Mini App, Zalo Not Active, Zalo Follow OA sheets."""
        t = self.timer("export_cnv_zalo_tab")
        r = self.client.get(reverse("cnv:export_customer_analytics") + "?tab=zalo")
        t.checkpoint("GET /cnv/export-customer-analytics/ tab=zalo")
        t.report()
        self._assert_excel(r)
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        sheet_names = wb.sheetnames
        self.assertIn("Zalo Mini App",   sheet_names, "Missing 'Zalo Mini App' sheet")
        self.assertIn("Zalo Not Active", sheet_names, "Missing 'Zalo Not Active' sheet")
        self.assertIn("Zalo Follow OA",  sheet_names, "Missing 'Zalo Follow OA' sheet")
        self.assert_snapshot("cnv_export_zalo_tab_sheets", {"sheets": sheet_names})

    def test_export_customer_analytics_zalo_sheet_structure(self):
        """Full workbook export must include 'Zalo Not Active' sheet alongside 'Zalo Mini App'."""
        r = self.client.get(reverse("cnv:export_customer_analytics"))
        self._assert_excel(r)
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        self.assertIn("Zalo Mini App",   wb.sheetnames)
        self.assertIn("Zalo Not Active", wb.sheetnames)
        self.assertIn("Zalo Follow OA",  wb.sheetnames)

    # ── Product analytics export ──────────────────────────────────────────────

    def test_export_product_analytics_alltime(self):
        t = self.timer("export_product_alltime")
        r = self.client.get(reverse("export_product_analytics"))
        t.checkpoint("GET /products/export/ all-time")
        t.report()
        self._assert_excel(r)

    def test_export_product_analytics_period_filter(self):
        """Export with a date filter scoped to the actual sale detail date range."""
        from App.models import SaleDetail
        from django.db.models import Min, Max
        bounds = SaleDetail.objects.filter(sales_date__isnull=False).aggregate(
            mn=Min('sales_date'), mx=Max('sales_date')
        )
        if not bounds['mn']:
            self.skipTest("No dated sale detail data")
        params = f"?start_date={bounds['mn']}&end_date={bounds['mx']}"
        t = self.timer("export_product_period")
        r = self.client.get(reverse("export_product_analytics") + params)
        t.checkpoint(f"GET /products/export/ {bounds['mn']} → {bounds['mx']}")
        t.report()
        self._assert_excel(r)

    def test_export_product_analytics_tab_brand(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=brand")
        self._assert_excel(r)

    def test_export_product_analytics_tab_category(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=category")
        self._assert_excel(r)

    def test_export_product_analytics_tab_year(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=year")
        self._assert_excel(r)

    def test_export_product_analytics_tab_week(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=week")
        self._assert_excel(r)

    def test_export_product_analytics_tab_sales_season(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=sales_season")
        self._assert_excel(r)

    def test_export_product_analytics_tab_product_season(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=product_season")
        self._assert_excel(r)

    def test_export_product_analytics_tab_vip_grade(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=vip_grade")
        self._assert_excel(r)

    def test_export_product_analytics_tab_shop(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=shop")
        self._assert_excel(r)

    def test_export_product_analytics_tab_product(self):
        r = self.client.get(reverse("export_product_analytics") + "?tab=product")
        self._assert_excel(r)

    def test_export_product_analytics_with_shop_name(self):
        """Export filtered to a specific shop — verifies shop_name param threads through."""
        from App.models import SaleDetail
        shop = SaleDetail.objects.values_list('shop_name', flat=True).order_by('shop_name').first()
        if not shop:
            self.skipTest("No SaleDetail data")
        import urllib.parse
        r = self.client.get(
            reverse("export_product_analytics") + f"?shop_name={urllib.parse.quote(shop)}"
        )
        self._assert_excel(r)

    # ── Inventory dead stock export ───────────────────────────────────────────

    def test_export_inventory_dead_stock_200(self):
        from App.models import InventorySnapshot
        if not InventorySnapshot.objects.exists():
            self.skipTest("No inventory data")
        t = self.timer("export_inventory_dead_stock")
        r = self.client.get(reverse("export_inventory_dead_stock"))
        t.checkpoint("GET /inventory/export/")
        t.report()
        self.assertEqual(r.status_code, 200)
        self.assertIn("text/csv", r.get("Content-Type", ""))

    def test_export_inventory_dead_stock_with_filters(self):
        from App.models import InventorySnapshot
        if not InventorySnapshot.objects.exists():
            self.skipTest("No inventory data")
        year = InventorySnapshot.objects.exclude(year__isnull=True).values_list('year', flat=True).first()
        params = f"?year={year}" if year else ""
        r = self.client.get(reverse("export_inventory_dead_stock") + params)
        self.assertEqual(r.status_code, 200)
        self.assertIn("text/csv", r.get("Content-Type", ""))


class ExportPermissionTest(SnapshotTestCase):
    """Verify export endpoints enforce permissions — unauthorized users get 302."""

    EXPORT_URLS = [
        ("export_analytics",           None),
        ("export_product_analytics",   None),
        ("export_inventory_dead_stock", None),
        ("export_shop_detail_excel",   None),
        ("export_sales_chart_excel",   None),
        ("export_coupon_chart_excel",  None),
        ("export_coupons",             None),
    ]

    @classmethod
    def setUpTestData(cls):
        cls.no_perm_user = User.objects.create_user(
            username="noperm_export", password="testpass"
        )

    def setUp(self):
        self.client.force_login(self.no_perm_user)

    def test_export_endpoints_require_permission(self):
        """All export endpoints redirect a user with no role/permissions."""
        for url_name, kwargs in self.EXPORT_URLS:
            with self.subTest(url=url_name):
                r = self.client.get(reverse(url_name), follow=False)
                self.assertIn(
                    r.status_code, [302, 403],
                    f"{url_name} returned {r.status_code}, expected 302/403 for user without perm",
                )

    def test_cnv_export_requires_permission(self):
        r = self.client.get(reverse("cnv:export_customer_analytics"), follow=False)
        self.assertIn(r.status_code, [302, 403])


class ExportDataParityTest(SnapshotTestCase):
    """Verify exported file content matches what the UI analytics functions return."""

    @classmethod
    def setUpTestData(cls):
        import io
        from App.services import process_inventory_file, process_sale_detail_file
        from tests.base import INPUT_DIR

        cls.superuser = User.objects.create_superuser(
            username="parity_admin", password="testpass", email="parity@test.com"
        )

        def _named(path):
            with open(path, "rb") as f:
                data = f.read()
            class _N(io.BytesIO):
                pass
            obj = _N(data)
            obj.name = path.name
            return obj

        inv_file  = INPUT_DIR / "inventory.xlsx"
        sd_file   = INPUT_DIR / "sale detail.xlsx"

        if inv_file.exists():
            process_inventory_file(_named(inv_file))
        if sd_file.exists():
            process_sale_detail_file(_named(sd_file))

    def setUp(self):
        self.client.force_login(self.superuser)

    # ── Inventory dead stock parity ───────────────────────────────────────────

    def test_inventory_csv_structure(self):
        """CSV structure snapshot: columns, row count, sort order, first-row key fields."""
        from App.models import InventorySnapshot
        if not InventorySnapshot.objects.exists():
            self.skipTest("No inventory data")
        import csv, io
        from App.analytics.inventory_functions import get_inventory_overview

        r = self.client.get(reverse("export_inventory_dead_stock"))
        self.assertEqual(r.status_code, 200)
        content = r.content.decode("utf-8-sig")
        reader = list(csv.reader(io.StringIO(content)))
        header = reader[0]
        data_rows = reader[1:]

        # Always verify column headers (structural contract)
        for col in ("Shop", "Product Code", "Product Name", "Product Name VN",
                    "Color", "Size", "Category L1", "Category L2", "Category L3",
                    "Gender", "Brand", "Year", "Season", "Qty", "Value (VND)"):
            self.assertIn(col, header, f"CSV missing column '{col}'")

        # Always verify sort order
        val_idx = header.index("Value (VND)")
        values = [float(r[val_idx]) for r in data_rows if r[val_idx]]
        self.assertEqual(values, sorted(values, reverse=True), "CSV not sorted by Value descending")

        # Always verify CSV row count matches UI data
        data = get_inventory_overview()
        ui_top = data.get('for_sale', {}).get('dead', {}).get('top', [])
        self.assertEqual(len(data_rows), len(ui_top),
                         f"CSV has {len(data_rows)} rows but UI has {len(ui_top)}")

        # Snapshot: locks in exact row count and first-row identifiers
        if ui_top:
            csv_first = dict(zip(header, data_rows[0]))
            self.assert_snapshot("inventory_csv_structure", {
                "row_count": len(data_rows),
                "columns": header,
                "first_product_code": csv_first.get("Product Code", ""),
                "first_shop": csv_first.get("Shop", ""),
            })

    # ── Product analytics export parity ──────────────────────────────────────

    def _get_excel_data_rows(self, tab_param, sheet_name_fragment, summary_only=False):
        """Download export for a tab and return data rows (skip header row).

        summary_only=True: only period-level rows (first column not empty/blank),
        excludes category sub-rows which have '' in the first column.
        """
        import openpyxl, io
        r = self.client.get(reverse("export_product_analytics") + f"?tab={tab_param}")
        self.assertEqual(r.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = next(
            (wb[n] for n in wb.sheetnames if sheet_name_fragment.lower() in n.lower()),
            wb.active,
        )
        rows = [row for row in ws.iter_rows(min_row=2, values_only=True)
                if any(c is not None for c in row)]
        if summary_only:
            rows = [row for row in rows if row[0] not in (None, '')]
        return rows

    def _assert_product_tab_parity(self, tab, ui_key, sheet_fragment):
        """Shared logic: verify Excel summary rows match UI + snapshot the counts."""
        from App.models import SaleDetail
        if not SaleDetail.objects.exists():
            self.skipTest("No sale detail data")
        from App.analytics.product_analytics import get_product_tab

        ui_data = get_product_tab(tab)
        ui_rows = ui_data.get(ui_key, []) if ui_data else []
        if not ui_rows:
            self.skipTest(f"No {tab} data in fixture")

        xl_rows = self._get_excel_data_rows(tab, sheet_fragment, summary_only=True)
        self.assertEqual(
            len(xl_rows), len(ui_rows),
            f"Excel has {len(xl_rows)} {tab} rows but UI has {len(ui_rows)}",
        )

        # Snapshot locks in the row count and first label so future regressions are visible
        overview = ui_data.get('overview', {}) if ui_data else {}
        self.assert_snapshot(f"product_export_parity_{tab}", {
            "row_count": len(ui_rows),
            "first_label": str(xl_rows[0][0]) if xl_rows else None,
            "total_qty": overview.get('total_qty'),
            "total_amount": overview.get('total_amount'),
        })

    def test_product_export_brand_tab_row_count(self):
        self._assert_product_tab_parity('brand', 'by_brand', 'brand')

    def test_product_export_month_tab_row_count(self):
        self._assert_product_tab_parity('month', 'by_month', 'month')

    def test_product_export_week_tab_row_count(self):
        self._assert_product_tab_parity('week', 'by_week', 'week')

    def test_product_export_sales_season_tab_row_count(self):
        self._assert_product_tab_parity('sales_season', 'by_sales_season', 'season')

    def test_product_export_vip_grade_tab_row_count(self):
        self._assert_product_tab_parity('vip_grade', 'by_vip_grade', 'grade')

    def test_product_export_year_tab_row_count(self):
        self._assert_product_tab_parity('year', 'by_year', 'year')

    def test_product_export_product_season_tab_row_count(self):
        self._assert_product_tab_parity('product_season', 'by_product_season', 'season')
