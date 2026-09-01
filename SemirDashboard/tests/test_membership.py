"""
tests/test_membership.py — Customer Membership snapshot feature tests.

Covers:
  1. next_tier_info() pure-function correctness (calculations.py)
  1B. resolve_grade() — shared grade-resolution rule
  2. compute_annual_spend_map() correctness — year window + as_of_date cutoff
  3. create_auto_snapshot() — snapshots the entire Customer table into
     MembershipSnapshotBatch.grade_counts/grade_members (JSON — redesigned
     2026-09-01, see App/models/membership.py docstring)
  4. Auto-hook wiring into upload_customers
  5. create_backfill_snapshot() never touches the live Customer table;
     also overrides each row's registration_store with the vip_id's
     CURRENT live Customer.registration_store (PO decision 2026-09-02),
     falling back to the file's own value when there's no live match.
     5B. _resolve_live_stores() bulk helper — exactly 1 query.
  6. compare_batches() delta calculation
  7. get_live_customer_tier_table() — live Customer-table counterpart
  8. Web smoke — /membership/ permission gating
  9-13. get_grade_breakdown_by_store(), store= filters, get_snapshot_registration_stores()
  14. get_grade_breakdown_by_store_comparison() — From/To matrix (added 2026-09-01)
  15. get_grade_changes() — grade-change diff feature (added 2026-09-01)
  16. membership_movers_partial web tests (added 2026-09-01)

Run:
  cd SemirDashboard && python manage.py test tests.test_membership -v 2
"""
import io
import json
from datetime import date
from decimal import Decimal
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook
from django.contrib.auth.models import User
from django.core.management import call_command
from django.test import Client, TestCase

from App.models import Customer, SalesTransaction, Role, UserProfile
from App.models.membership import MembershipSnapshotBatch
from App.analytics.calculations import next_tier_info, GRADE_UPGRADE_THRESHOLDS
from App.analytics.customer_utils import resolve_grade
from App.analytics.membership import (
    compute_annual_spend_map, compare_batches,
    get_grade_breakdown, get_all_batch_grade_series, get_grade_breakdown_by_store,
    get_grade_breakdown_by_store_comparison, get_grade_changes,
    get_grade_changes_overview_by_store, get_grade_changes_store_transitions,
    get_live_customer_tier_table, get_snapshot_registration_stores, DISPLAY_GRADES,
)
from App.services.membership_snapshot import (
    create_auto_snapshot, create_backfill_snapshot, _resolve_live_stores,
)

from tests.base import SnapshotTestCase


def _make_xlsx(headers, rows=None):
    buf = io.BytesIO()
    df = pd.DataFrame(rows or [], columns=headers)
    df.to_excel(buf, index=False)
    return buf.getvalue()


def _django_file(data, filename):
    from django.core.files.uploadedfile import SimpleUploadedFile
    return SimpleUploadedFile(filename, data)


def _customer(vip_id, phone, grade='Member', reg_date=None, points=0, store=''):
    return Customer.objects.create(
        vip_id=vip_id, phone=phone, name=f"Cust {vip_id}", vip_grade=grade,
        registration_date=reg_date, points=points, registration_store=store,
    )


def _sale(vip_id, sales_date, amount, invoice):
    return SalesTransaction.objects.create(
        invoice_number=invoice, shop_id="S1", shop_name="Test Shop", country="VN",
        sales_date=sales_date, vip_id=vip_id, vip_name=f"Cust {vip_id}",
        quantity=1, settlement_amount=Decimal(str(amount)), sales_amount=Decimal(str(amount)),
    )


def _add_snapshot_member(batch, vip_id, grade, store=''):
    """
    Test helper — adds one customer into a MembershipSnapshotBatch's
    grade_counts/grade_members JSON fields, mirroring exactly what
    App/services/membership_snapshot.py::_build_rows() does. Lets tests
    express "this batch has these customers" without needing a real
    Customer table + create_auto_snapshot() round-trip for every fixture.
    Mutates and saves `batch`; returns it for chaining.
    """
    store_key = store or '(No Store)'
    gc = batch.grade_counts or {}
    gm = batch.grade_members or {}
    gc.setdefault('overall', {})
    gc['overall'][grade] = gc['overall'].get(grade, 0) + 1
    gc.setdefault('by_store', {}).setdefault(store_key, {})
    gc['by_store'][store_key][grade] = gc['by_store'][store_key].get(grade, 0) + 1
    gm.setdefault('overall', {}).setdefault(grade, []).append(vip_id)
    gm.setdefault('by_store', {}).setdefault(store_key, {}).setdefault(grade, []).append(vip_id)
    batch.grade_counts = gc
    batch.grade_members = gm
    batch.row_count = batch.row_count + 1
    batch.save(update_fields=['grade_counts', 'grade_members', 'row_count'])
    return batch


class _ClearDropdownCacheMixin:
    """Any test hitting GET /membership/ populates shop_detail.py's shared
    "shop_detail_dropdowns" cache (5-min TTL, App/views/shop_detail.py
    _get_dropdown_options()) from that test's transaction-local data. Django's
    cache is not rolled back with the DB transaction, so a stale/empty entry
    would otherwise leak into whichever test runs next and expects a fresh
    computation — found via a real cross-file failure where this leaked an
    empty `sales_shops` into ShopDetailTest.test_main_page_loads."""

    def setUp(self):
        from django.core.cache import cache
        cache.delete("shop_detail_dropdowns")
        super().setUp()

    def tearDown(self):
        from django.core.cache import cache
        cache.delete("shop_detail_dropdowns")
        super().tearDown()


# ---------------------------------------------------------------------------
# 1. next_tier_info() — pure function
# ---------------------------------------------------------------------------

class NextTierInfoTest(TestCase):
    def test_no_grade_targets_silver(self):
        target, remaining = next_tier_info('No Grade', Decimal('0'))
        self.assertEqual(target, 'Silver')
        self.assertEqual(remaining, GRADE_UPGRADE_THRESHOLDS['Silver'])

    def test_member_just_under_silver_threshold(self):
        target, remaining = next_tier_info('Member', Decimal('5999999'))
        self.assertEqual(target, 'Silver')
        self.assertEqual(remaining, Decimal('1'))

    def test_member_exactly_at_silver_threshold(self):
        target, remaining = next_tier_info('Member', GRADE_UPGRADE_THRESHOLDS['Silver'])
        self.assertEqual(target, 'Silver')
        self.assertEqual(remaining, Decimal('0'))

    def test_silver_over_gold_threshold_floors_at_zero(self):
        target, remaining = next_tier_info('Silver', Decimal('99999999'))
        self.assertEqual(target, 'Gold')
        self.assertEqual(remaining, Decimal('0'))

    def test_gold_targets_diamond(self):
        target, remaining = next_tier_info('Gold', Decimal('15000000'))
        self.assertEqual(target, 'Diamond')
        self.assertEqual(remaining, Decimal('5000000'))

    def test_diamond_has_no_next_tier(self):
        target, remaining = next_tier_info('Diamond', Decimal('50000000'))
        self.assertIsNone(target)
        self.assertEqual(remaining, Decimal('0'))

    def test_unrecognized_grade_has_no_next_tier(self):
        target, remaining = next_tier_info('Unknown Grade', Decimal('0'))
        self.assertIsNone(target)
        self.assertEqual(remaining, Decimal('0'))


# ---------------------------------------------------------------------------
# 1B. resolve_grade() — shared by App/services/membership_snapshot.py
#     (_build_rows) and App/analytics/membership.py (get_live_customer_tier_table)
# ---------------------------------------------------------------------------

class ResolveGradeTest(TestCase):
    def test_vip_id_zero_forced_to_no_grade_regardless_of_raw_grade(self):
        self.assertEqual(resolve_grade('0', 'Gold'), 'No Grade')
        self.assertEqual(resolve_grade('0', ''), 'No Grade')

    def test_normal_vip_id_uses_normalize_grade(self):
        self.assertEqual(resolve_grade('V1', 'Silver'), 'Silver')
        self.assertEqual(resolve_grade('V1', 'Golden'), 'Gold')  # normalize_grade() typo handling
        self.assertEqual(resolve_grade('V1', ''), 'No Grade')
        self.assertEqual(resolve_grade('V1', None), 'No Grade')


# ---------------------------------------------------------------------------
# 2. compute_annual_spend_map() — year window + as_of_date cutoff
# ---------------------------------------------------------------------------

class ComputeAnnualSpendMapTest(TestCase):
    def test_only_target_year_counted(self):
        _sale('V1', date(2025, 3, 1), 1000000, 'INV-2025-1')
        _sale('V1', date(2026, 2, 1), 2000000, 'INV-2026-1')  # different year — excluded
        result = compute_annual_spend_map(date(2025, 12, 31))
        self.assertEqual(result['V1']['annual_spend'], Decimal('1000000'))
        self.assertEqual(result['V1']['annual_purchase_count'], 1)

    def test_sale_after_as_of_date_excluded(self):
        _sale('V2', date(2026, 6, 1), 1000000, 'INV-A')
        _sale('V2', date(2026, 6, 2), 5000000, 'INV-B')  # day AFTER as_of_date — excluded
        result = compute_annual_spend_map(date(2026, 6, 1))
        self.assertEqual(result['V2']['annual_spend'], Decimal('1000000'))
        self.assertEqual(result['V2']['annual_purchase_count'], 1)

    def test_sale_on_as_of_date_included(self):
        _sale('V3', date(2026, 6, 1), 3000000, 'INV-C')
        result = compute_annual_spend_map(date(2026, 6, 1))
        self.assertEqual(result['V3']['annual_spend'], Decimal('3000000'))

    def test_multiple_invoices_sum_and_count(self):
        _sale('V4', date(2026, 1, 5), 1000000, 'INV-D1')
        _sale('V4', date(2026, 2, 5), 2000000, 'INV-D2')
        result = compute_annual_spend_map(date(2026, 12, 31))
        self.assertEqual(result['V4']['annual_spend'], Decimal('3000000'))
        self.assertEqual(result['V4']['annual_purchase_count'], 2)

    def test_customer_with_no_sales_absent_from_map(self):
        result = compute_annual_spend_map(date(2026, 6, 1))
        self.assertNotIn('V-NONEXISTENT', result)

    def test_vip_id_zero_excluded(self):
        # vip_id "0" = buyer without info, excluded from grade analytics
        # everywhere else in the codebase — must not accumulate spend here.
        _sale('0', date(2026, 3, 1), 9999999, 'INV-ANON')
        result = compute_annual_spend_map(date(2026, 12, 31))
        self.assertNotIn('0', result)


# ---------------------------------------------------------------------------
# 3. create_auto_snapshot() — snapshots the entire Customer table into JSON
# ---------------------------------------------------------------------------

class CreateAutoSnapshotTest(TestCase):
    def test_snapshots_entire_customer_table(self):
        _customer('C1', 'P1', grade='Silver')
        _customer('C2', 'P2', grade='Gold')
        _customer('C3', 'P3', grade='No Grade')

        batch = create_auto_snapshot(as_of_date=date(2026, 6, 1))

        self.assertEqual(batch.source, 'auto')
        self.assertEqual(batch.row_count, Customer.objects.count())
        overall = batch.grade_counts['overall']
        self.assertEqual(overall.get('Silver', 0), 1)
        self.assertEqual(overall.get('Gold', 0), 1)
        self.assertEqual(overall.get('No Grade', 0), 1)
        self.assertEqual(batch.grade_members['overall']['Silver'], ['C1'])

    def test_vip_id_zero_forced_to_no_grade_regardless_of_stored_grade(self):
        # VIP ID "0" = buyer without info, excluded from grade analytics
        # everywhere else in the codebase — a raw vip_grade of "Gold" on such
        # a row must not leak into the Gold bucket.
        _customer('0', 'P6', grade='Gold')
        batch = create_auto_snapshot(as_of_date=date(2026, 6, 1))
        self.assertEqual(batch.grade_counts['overall'].get('Gold', 0), 0)
        self.assertEqual(batch.grade_counts['overall'].get('No Grade', 0), 1)
        self.assertEqual(batch.grade_members['overall']['No Grade'], ['0'])

    def test_grouped_by_store(self):
        _customer('C7', 'P7', grade='Silver', store='Shop A')
        _customer('C8', 'P8', grade='Gold', store='Shop B')
        batch = create_auto_snapshot(as_of_date=date(2026, 6, 1))
        self.assertEqual(batch.grade_counts['by_store']['Shop A'].get('Silver', 0), 1)
        self.assertEqual(batch.grade_counts['by_store']['Shop B'].get('Gold', 0), 1)

    def test_blank_store_bucketed_as_no_store(self):
        _customer('C9', 'P9', grade='Gold', store='')
        batch = create_auto_snapshot(as_of_date=date(2026, 6, 1))
        self.assertEqual(batch.grade_counts['by_store']['(No Store)'].get('Gold', 0), 1)

    def test_unaffected_by_backfill_live_store_resolution(self):
        # Regression check only (not new logic): create_auto_snapshot()
        # already reads registration_store straight from the live Customer
        # table by construction, so the backfill-only _resolve_live_stores()
        # fix must not touch its behavior or call it.
        _customer('C10', 'P10', grade='Silver', store='Shop A')
        with patch("App.services.membership_snapshot._resolve_live_stores") as mock_resolve:
            batch = create_auto_snapshot(as_of_date=date(2026, 6, 1))
        mock_resolve.assert_not_called()
        self.assertEqual(batch.grade_counts['by_store']['Shop A'].get('Silver', 0), 1)


# ---------------------------------------------------------------------------
# 4. Auto-hook wiring into upload_customers
# ---------------------------------------------------------------------------

class MembershipAutoHookTest(TestCase):
    def setUp(self):
        # _start_thread is mocked in every test below, so the real thread
        # (which normally releases the per-type upload lock in its finally
        # block) never runs — clear the cache each test so the "customers"
        # type lock from a previous test method doesn't block this one.
        from django.core.cache import cache
        cache.clear()
        self.client = Client()
        self.user = User.objects.create_superuser("testadmin", password="pw")
        self.client.force_login(self.user)

    def test_membership_hook_passed_to_start_thread_on_successful_upload(self):
        data = _make_xlsx(["VIP ID", "PHONE NO.", "Name"], [{"VIP ID": "X1", "PHONE NO.": "0900", "Name": "A"}])
        with patch("App.views.upload._start_thread") as mock_thread:
            r = self.client.post(
                "/upload/customers/",
                {"file": _django_file(data, "good.xlsx")},
                SERVER_NAME="localhost",
            )
        self.assertEqual(r.status_code, 302)
        mock_thread.assert_called_once()
        on_done_fn = mock_thread.call_args[0][4]
        self.assertTrue(callable(on_done_fn))

    def test_membership_hook_calls_create_auto_snapshot(self):
        data = _make_xlsx(["VIP ID", "PHONE NO.", "Name"], [{"VIP ID": "X2", "PHONE NO.": "0901", "Name": "B"}])
        with patch("App.views.upload._start_thread") as mock_thread:
            self.client.post(
                "/upload/customers/",
                {"file": _django_file(data, "good.xlsx")},
                SERVER_NAME="localhost",
            )
        on_done_fn = mock_thread.call_args[0][4]
        with patch("App.services.membership_snapshot.create_auto_snapshot") as mock_snap:
            on_done_fn()
        mock_snap.assert_called_once()

    def test_membership_hook_swallows_its_own_exceptions(self):
        """A snapshot failure must NOT propagate — the customer-import job's
        status must stay based on the import result, not the hook."""
        data = _make_xlsx(["VIP ID", "PHONE NO.", "Name"], [{"VIP ID": "X3", "PHONE NO.": "0902", "Name": "C"}])
        with patch("App.views.upload._start_thread") as mock_thread:
            self.client.post(
                "/upload/customers/",
                {"file": _django_file(data, "good.xlsx")},
                SERVER_NAME="localhost",
            )
        on_done_fn = mock_thread.call_args[0][4]
        with patch("App.services.membership_snapshot.create_auto_snapshot", side_effect=RuntimeError("boom")):
            on_done_fn()  # must not raise


# ---------------------------------------------------------------------------
# 5. create_backfill_snapshot() never touches live Customer
# ---------------------------------------------------------------------------

class CreateBackfillSnapshotTest(TestCase):
    def test_does_not_modify_live_customer_table(self):
        live = _customer('B1', 'PB1', grade='Silver', points=100)
        before_count = Customer.objects.count()
        before_grade = live.vip_grade
        before_points = live.points

        data = _make_xlsx(
            ["VIP ID", "PHONE NO.", "Name", "VIP GRADE", "POINTS"],
            [{"VIP ID": "B1", "PHONE NO.": "PB1", "Name": "Backfilled", "VIP GRADE": "Diamond", "POINTS": 9999}],
        )
        f = _django_file(data, "backfill.xlsx")
        result = create_backfill_snapshot(f, snapshot_date=date(2025, 1, 15), note="test backfill")

        self.assertEqual(Customer.objects.count(), before_count)
        live.refresh_from_db()
        self.assertEqual(live.vip_grade, before_grade)
        self.assertEqual(live.points, before_points)

        batch = MembershipSnapshotBatch.objects.get(pk=result['batch_id'])
        self.assertEqual(batch.source, 'manual_import')
        self.assertEqual(batch.grade_counts['overall'].get('Diamond', 0), 1)
        self.assertEqual(batch.grade_members['overall']['Diamond'], ['B1'])

    def test_uses_live_store_not_file_store(self):
        # PO decision 2026-09-02: old historical files can use a different
        # store-naming format than the live table — the backfill snapshot
        # must attribute the row to the CURRENT live store, not the file's.
        _customer('B2', 'PB2', grade='Gold', store='巴拉越南河内市SAVICO MEGAMALL-直营店')

        data = _make_xlsx(
            ["VIP ID", "PHONE NO.", "Name", "VIP GRADE", "REGISTRATION STORE"],
            [{"VIP ID": "B2", "PHONE NO.": "PB2", "Name": "Backfilled",
              "VIP GRADE": "Gold", "REGISTRATION STORE": "Savico Megamall"}],
        )
        f = _django_file(data, "backfill.xlsx")
        result = create_backfill_snapshot(f, snapshot_date=date(2025, 1, 15))

        batch = MembershipSnapshotBatch.objects.get(pk=result['batch_id'])
        self.assertNotIn('Savico Megamall', batch.grade_counts['by_store'])
        self.assertEqual(
            batch.grade_counts['by_store']['巴拉越南河内市SAVICO MEGAMALL-直营店'].get('Gold', 0), 1,
        )
        self.assertEqual(
            batch.grade_members['by_store']['巴拉越南河内市SAVICO MEGAMALL-直营店']['Gold'], ['B2'],
        )

    def test_falls_back_to_file_store_when_no_live_match(self):
        # vip_id in the uploaded file but not in the live Customer table
        # (deleted since, or never re-uploaded) — keep the file's own value.
        data = _make_xlsx(
            ["VIP ID", "PHONE NO.", "Name", "VIP GRADE", "REGISTRATION STORE"],
            [{"VIP ID": "GONE1", "PHONE NO.": "P999", "Name": "Ghost",
              "VIP GRADE": "Silver", "REGISTRATION STORE": "Old Shop Name"}],
        )
        f = _django_file(data, "backfill.xlsx")
        result = create_backfill_snapshot(f, snapshot_date=date(2025, 1, 15))

        batch = MembershipSnapshotBatch.objects.get(pk=result['batch_id'])
        self.assertEqual(batch.grade_counts['by_store']['Old Shop Name'].get('Silver', 0), 1)
        self.assertEqual(batch.grade_members['by_store']['Old Shop Name']['Silver'], ['GONE1'])

    def test_uses_live_store_when_live_store_is_blank(self):
        # Direct test for the blank-live-store edge case (review follow-up,
        # 2026-09-02) -- the file's store is non-blank but the live store is
        # blank; the row must land under the canonical '(No Store)' bucket,
        # not a raw '' key, matching _resolve_live_stores()'s coercion.
        _customer('B3', 'PB3', grade='Silver', store='')

        data = _make_xlsx(
            ["VIP ID", "PHONE NO.", "Name", "VIP GRADE", "REGISTRATION STORE"],
            [{"VIP ID": "B3", "PHONE NO.": "PB3", "Name": "Backfilled",
              "VIP GRADE": "Silver", "REGISTRATION STORE": "Some Old Store"}],
        )
        f = _django_file(data, "backfill.xlsx")
        result = create_backfill_snapshot(f, snapshot_date=date(2025, 1, 15))

        batch = MembershipSnapshotBatch.objects.get(pk=result['batch_id'])
        self.assertIn('(No Store)', batch.grade_counts['by_store'])
        self.assertNotIn('', batch.grade_counts['by_store'])
        self.assertEqual(batch.grade_counts['by_store']['(No Store)'].get('Silver', 0), 1)


# ---------------------------------------------------------------------------
# 5B. _resolve_live_stores() — bulk store-lookup helper
# ---------------------------------------------------------------------------

class ResolveLiveStoresTest(TestCase):
    def test_returns_only_existing_vip_ids_with_correct_values(self):
        _customer('E1', 'PE1', store='Shop A')
        _customer('E2', 'PE2', store='Shop B')

        result = _resolve_live_stores(['E1', 'E2', 'MISSING1', 'MISSING2'])

        self.assertEqual(result, {'E1': 'Shop A', 'E2': 'Shop B'})
        self.assertNotIn('MISSING1', result)
        self.assertNotIn('MISSING2', result)

    def test_exactly_one_query_regardless_of_input_size(self):
        _customer('E3', 'PE3', store='Shop C')
        vip_ids = ['E3'] + [f'NOPE{i}' for i in range(500)]
        with self.assertNumQueries(1):
            _resolve_live_stores(vip_ids)

    def test_blank_live_store_coerced_to_no_store_placeholder(self):
        # Centralized here 2026-09-02 (see docstring) after a real bug where
        # a caller used the raw '' value as a by_store dict key instead of
        # the canonical '(No Store)' placeholder every other store-keying
        # path in this codebase uses.
        _customer('E4', 'PE4', store='')
        result = _resolve_live_stores(['E4'])
        self.assertEqual(result, {'E4': '(No Store)'})


# ---------------------------------------------------------------------------
# 6. compare_batches() delta calculation
# ---------------------------------------------------------------------------

class CompareBatchesTest(TestCase):
    def test_delta_and_delta_pct(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        for i in range(2):
            _add_snapshot_member(b1, f"S{i}", 'Silver')
        for i in range(3):
            _add_snapshot_member(b2, f"S{i}", 'Silver')

        rows = {r['grade']: r for r in compare_batches(b1.id, b2.id)}
        self.assertEqual(rows['Silver']['from_count'], 2)
        self.assertEqual(rows['Silver']['to_count'], 3)
        self.assertEqual(rows['Silver']['delta'], 1)
        self.assertEqual(rows['Silver']['delta_pct'], 50.0)
        self.assertEqual(rows['Gold']['from_count'], 0)
        self.assertIsNone(rows['Gold']['delta_pct'])  # division by zero guarded

    def test_no_from_batch_treats_as_zero(self):
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b2, "Z1", 'Gold')
        rows = {r['grade']: r for r in compare_batches(None, b2.id)}
        self.assertEqual(rows['Gold']['from_count'], 0)
        self.assertEqual(rows['Gold']['to_count'], 1)

    def test_no_grade_excluded_from_breakdown_and_comparison(self):
        # PO feedback 2026-08-14: "No Grade" is noise in the grade-level KPI
        # view — the grade-level summary/comparison/chart excludes it.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, "N1", 'No Grade')
        _add_snapshot_member(b2, "N1", 'No Grade')
        _add_snapshot_member(b2, "N2", 'No Grade')

        breakdown = get_grade_breakdown(b2.id)
        self.assertNotIn('No Grade', breakdown)

        grades = [r['grade'] for r in compare_batches(b1.id, b2.id)]
        self.assertNotIn('No Grade', grades)

        series = get_all_batch_grade_series()
        for entry in series:
            self.assertNotIn('No Grade', entry['counts'])


# ---------------------------------------------------------------------------
# 7. get_live_customer_tier_table() — live Customer-table counterpart, PO
#    feedback 2026-08-31: "Customer Tier Progress" has nothing to do with
#    snapshot. Applies the same resolve_grade() convention as
#    App/services/membership_snapshot.py::_build_rows().
# ---------------------------------------------------------------------------

class GetLiveCustomerTierTableTest(TestCase):
    def test_reads_from_customer_not_snapshot(self):
        _customer('L1', 'PL1', grade='Silver')
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'SNAPONLY', 'Gold')

        rows, total_count = get_live_customer_tier_table()
        vip_ids = [r['vip_id'] for r in rows]
        self.assertIn('L1', vip_ids)
        self.assertNotIn('SNAPONLY', vip_ids)

    def test_works_with_zero_customers_and_zero_batches(self):
        rows, total_count = get_live_customer_tier_table()
        self.assertEqual(rows, [])
        self.assertEqual(total_count, 0)

    def test_annual_spend_computed_live_from_sales(self):
        _customer('L2', 'PL2', grade='Member')
        _sale('L2', date(2026, 3, 1), 7000000, 'INV-LIVE1')
        rows, total_count = get_live_customer_tier_table(as_of_date=date(2026, 6, 1))
        row = next(r for r in rows if r['vip_id'] == 'L2')
        self.assertEqual(row['annual_spend'], Decimal('7000000'))
        self.assertEqual(row['next_grade'], 'Silver')  # already over Silver threshold

    def test_vip_id_zero_forced_to_no_grade(self):
        _customer('0', 'PL3', grade='Gold')
        rows, total_count = get_live_customer_tier_table()
        row = next(r for r in rows if r['vip_id'] == '0')
        self.assertEqual(row['grade'], 'No Grade')

    def test_grade_filter_applies_after_normalization(self):
        _customer('L4', 'PL4', grade='Golden')  # normalize_grade() maps 'Golden' -> 'Gold'
        _customer('L5', 'PL5', grade='Silver')
        rows, total_count = get_live_customer_tier_table(grade_filter='Gold')
        self.assertEqual([r['vip_id'] for r in rows], ['L4'])
        self.assertEqual(total_count, 1)

    def test_shop_filter_is_exact_match(self):
        # The store filter is a <select> of exact DB values, not free text.
        _customer('L6', 'PL6', grade='Silver', store='Bala VN Haiphong AEON MALL- Direct')
        _customer('L7', 'PL7', grade='Silver', store='BL VN North Warehouse')
        rows, total_count = get_live_customer_tier_table(shop_filter='Bala VN Haiphong AEON MALL- Direct')
        self.assertEqual([r['vip_id'] for r in rows], ['L6'])
        # A partial string must NOT match
        partial_rows, partial_count = get_live_customer_tier_table(shop_filter='Haiphong')
        self.assertEqual(partial_rows, [])
        self.assertEqual(partial_count, 0)

    def test_sort_ascending_amount_to_next_tier(self):
        _customer('L8', 'PL8', grade='Member')  # far from Silver (0 spend)
        _sale('L8', date(2026, 3, 1), 5900000, 'INV-LIVE2')
        _customer('L9', 'PL9', grade='Diamond')  # no next tier -> sorts last
        rows, total_count = get_live_customer_tier_table(as_of_date=date(2026, 6, 1))
        vip_order = [r['vip_id'] for r in rows]
        self.assertLess(vip_order.index('L8'), vip_order.index('L9'))


# ---------------------------------------------------------------------------
# 8. Web smoke — permission gating
# ---------------------------------------------------------------------------

class MembershipWebSmokeTest(_ClearDropdownCacheMixin, TestCase):
    def test_dashboard_200_for_superuser(self):
        user = User.objects.create_superuser("membadmin", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", follow=True, SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)

    def test_dashboard_blocked_for_anonymous(self):
        client = Client()
        r = client.get("/membership/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 302)  # @login_required redirects to login

    def test_dashboard_blocked_without_permission(self):
        user = User.objects.create_user("noperm", password="pw")  # no role/perm assigned
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 302)  # requires_perm redirects to "home"

    def test_table_partial_reads_live_customer_table_not_snapshot(self):
        # PO feedback 2026-08-31: "Customer Tier Progress" reads Customer,
        # "không liên quan gì đến snapshot" (nothing to do with snapshot) —
        # a customer that exists ONLY in a snapshot batch (never in the live
        # Customer table) must NOT show up here, and a live Customer with no
        # snapshot at all must show up.
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'SNAPONLY', 'Gold')
        _customer('LIVEONLY', 'PL1', grade='Gold')

        user = User.objects.create_superuser("membadmin2", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/partial/table/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"LIVEONLY", r.content)
        self.assertNotIn(b"SNAPONLY", r.content)

    def test_table_partial_works_with_zero_snapshot_batches(self):
        # Must not require any snapshot to exist at all — e.g. right after
        # this feature is deployed, before the next customer upload triggers
        # the first auto-snapshot.
        self.assertEqual(MembershipSnapshotBatch.objects.count(), 0)
        _customer('NOBATCH1', 'PNB1', grade='Silver')

        user = User.objects.create_superuser("membadmin3", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/partial/table/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"NOBATCH1", r.content)
        self.assertNotIn(b"No snapshot", r.content)


# ---------------------------------------------------------------------------
# 9. get_grade_breakdown_by_store() — one-batch matrix building block
# ---------------------------------------------------------------------------

class GetGradeBreakdownByStoreTest(TestCase):
    def test_counts_grouped_by_store_and_grade(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'S1', 'Silver', 'Shop A')
        _add_snapshot_member(batch, 'S2', 'Gold', 'Shop A')
        _add_snapshot_member(batch, 'S3', 'Silver', 'Shop B')

        rows = {r['store']: r for r in get_grade_breakdown_by_store(batch.id)}
        self.assertEqual(set(rows.keys()), {'Shop A', 'Shop B'})
        a = dict(zip(DISPLAY_GRADES, rows['Shop A']['counts']))
        self.assertEqual(a['Silver'], 1)
        self.assertEqual(a['Gold'], 1)
        self.assertEqual(rows['Shop A']['total'], 2)
        b = dict(zip(DISPLAY_GRADES, rows['Shop B']['counts']))
        self.assertEqual(b['Silver'], 1)
        self.assertEqual(rows['Shop B']['total'], 1)

    def test_blank_store_bucketed_as_no_store(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'S4', 'Gold', '')
        rows = {r['store']: r for r in get_grade_breakdown_by_store(batch.id)}
        self.assertIn('(No Store)', rows)
        self.assertEqual(rows['(No Store)']['total'], 1)

    def test_no_grade_excluded_like_get_grade_breakdown(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'S5', 'No Grade', 'Shop C')
        rows = get_grade_breakdown_by_store(batch.id)
        self.assertEqual(rows, [])  # only row was 'No Grade' -> excluded entirely

    def test_empty_batch_id_returns_empty_list(self):
        self.assertEqual(get_grade_breakdown_by_store(None), [])


# ---------------------------------------------------------------------------
# 10. Registration Store dropdown wired to LIVE Customer DB — PO feedback
#     2026-08-31: "list này đang lấy latest db đúng không?"
# ---------------------------------------------------------------------------

class MembershipDashboardRegistrationStoreDropdownTest(_ClearDropdownCacheMixin, TestCase):
    def test_dashboard_context_has_live_registration_stores_not_snapshot_scoped(self):
        _customer('LIVE1', 'PL1', store='Live Store Only')
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'SNAP1', 'Gold', 'Snapshot-Only Store')

        user = User.objects.create_superuser("membadmin4", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", follow=True, SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        self.assertIn('Live Store Only', r.context['registration_stores'])
        self.assertNotIn('Snapshot-Only Store', r.context['registration_stores'])


# ---------------------------------------------------------------------------
# 11. membership_delete_batch — a MembershipSnapshotBatch row IS the storage
#     now (grade_counts/grade_members live directly on it, no child model).
# ---------------------------------------------------------------------------

class MembershipDeleteBatchTest(_ClearDropdownCacheMixin, TestCase):
    def test_delete_batch_removes_it(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'DEL1', 'Gold')
        batch_id = batch.id

        user = User.objects.create_superuser("delmemb1", password="pw")
        client = Client()
        client.force_login(user)
        r = client.post(f"/membership/delete-batch/{batch_id}/", SERVER_NAME="localhost", follow=True)

        self.assertEqual(r.status_code, 200)
        self.assertFalse(MembershipSnapshotBatch.objects.filter(pk=batch_id).exists())

    def test_delete_batch_blocked_without_permission(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        user = User.objects.create_user("delmemb2", password="pw")  # no role/perm assigned
        client = Client()
        client.force_login(user)
        r = client.post(f"/membership/delete-batch/{batch.id}/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 302)  # requires_perm redirects to "home"
        self.assertTrue(MembershipSnapshotBatch.objects.filter(pk=batch.id).exists())

    def test_delete_batch_blocked_with_adjacent_permissions_but_not_delete(self):
        # Stronger version of the above — a user holding the OTHER two
        # membership permissions (view + import) but explicitly NOT
        # membership.delete must still be blocked.
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        role = Role.objects.create(
            name='membership_no_delete', permissions=['membership.view', 'membership.import'], is_system=False)
        user = User.objects.create_user("delmemb5", password="pw")
        UserProfile.objects.create(user=user, role=role)
        client = Client()
        client.force_login(user)
        r = client.post(f"/membership/delete-batch/{batch.id}/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 302)
        self.assertTrue(MembershipSnapshotBatch.objects.filter(pk=batch.id).exists())

    def test_delete_batch_get_request_does_not_delete(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        user = User.objects.create_superuser("delmemb3", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(f"/membership/delete-batch/{batch.id}/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 302)
        self.assertTrue(MembershipSnapshotBatch.objects.filter(pk=batch.id).exists())

    def test_delete_nonexistent_batch_redirects_without_error(self):
        user = User.objects.create_superuser("delmemb4", password="pw")
        client = Client()
        client.force_login(user)
        r = client.post("/membership/delete-batch/999999/", SERVER_NAME="localhost", follow=True)
        self.assertEqual(r.status_code, 200)  # redirected to dashboard, not a 500

    def test_manage_snapshots_ui_hidden_without_delete_permission(self):
        MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        role = Role.objects.create(name='membership_view_only', permissions=['membership.view'], is_system=False)
        user = User.objects.create_user("delmemb6", password="pw")
        UserProfile.objects.create(user=user, role=role)
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", follow=True, SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        self.assertNotIn(b"Manage Snapshots", r.content)
        self.assertNotIn(b"delete-batch", r.content)

    def test_manage_snapshots_ui_shown_with_delete_permission(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        role = Role.objects.create(
            name='membership_with_delete', permissions=['membership.view', 'membership.delete'], is_system=False)
        user = User.objects.create_user("delmemb7", password="pw")
        UserProfile.objects.create(user=user, role=role)
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", follow=True, SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"Manage Snapshots", r.content)
        self.assertIn(f"/membership/delete-batch/{batch.id}/".encode(), r.content)


# ---------------------------------------------------------------------------
# 12. store= param on get_grade_breakdown / compare_batches / series
# ---------------------------------------------------------------------------

class GradeBreakdownStoreFilterTest(TestCase):
    def setUp(self):
        self.batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(self.batch, 'F1', 'Silver', 'Shop A')
        _add_snapshot_member(self.batch, 'F2', 'Gold', 'Shop A')
        _add_snapshot_member(self.batch, 'F3', 'Silver', 'Shop B')
        _add_snapshot_member(self.batch, 'F4', 'Gold', '')

    def test_no_store_filter_counts_everything(self):
        counts = get_grade_breakdown(self.batch.id)
        self.assertEqual(counts['Silver'], 2)
        self.assertEqual(counts['Gold'], 2)

    def test_store_filter_isolates_one_store(self):
        counts = get_grade_breakdown(self.batch.id, store='Shop A')
        self.assertEqual(counts['Silver'], 1)
        self.assertEqual(counts['Gold'], 1)
        # Shop B's Silver customer must not leak into Shop A's count
        counts_b = get_grade_breakdown(self.batch.id, store='Shop B')
        self.assertEqual(counts_b['Silver'], 1)
        self.assertEqual(counts_b['Gold'], 0)

    def test_no_store_bucket_matches_blank_registration_store(self):
        counts = get_grade_breakdown(self.batch.id, store='(No Store)')
        self.assertEqual(counts['Gold'], 1)
        self.assertEqual(counts['Silver'], 0)

    def test_nonexistent_store_returns_zero_counts(self):
        counts = get_grade_breakdown(self.batch.id, store='Nonexistent Shop')
        self.assertEqual(counts['Silver'], 0)
        self.assertEqual(counts['Gold'], 0)

    def test_compare_batches_store_filter(self):
        batch2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 7, 1), source='auto')
        _add_snapshot_member(batch2, 'F1', 'Silver', 'Shop A')
        _add_snapshot_member(batch2, 'F5', 'Silver', 'Shop A')
        _add_snapshot_member(batch2, 'F3', 'Silver', 'Shop B')

        rows = {r['grade']: r for r in compare_batches(self.batch.id, batch2.id, store='Shop A')}
        self.assertEqual(rows['Silver']['from_count'], 1)
        self.assertEqual(rows['Silver']['to_count'], 2)
        self.assertEqual(rows['Silver']['delta'], 1)

    def test_series_store_filter(self):
        batch2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 7, 1), source='auto')
        _add_snapshot_member(batch2, 'F1', 'Silver', 'Shop A')
        _add_snapshot_member(batch2, 'F3', 'Gold', 'Shop B')

        series = get_all_batch_grade_series(store='Shop A')
        by_batch = {s['batch_id']: s['counts'] for s in series}
        self.assertEqual(by_batch[self.batch.id]['Silver'], 1)
        self.assertEqual(by_batch[self.batch.id]['Gold'], 1)
        self.assertEqual(by_batch[batch2.id]['Silver'], 1)
        self.assertEqual(by_batch[batch2.id]['Gold'], 0)  # batch2's Gold customer is Shop B, not Shop A


# ---------------------------------------------------------------------------
# 13. get_snapshot_registration_stores() — PO feedback 2026-08-31: the
#     snapshot-scoped sections' store dropdown must NOT reuse the live-
#     Customer-sourced list.
# ---------------------------------------------------------------------------

class GetSnapshotRegistrationStoresTest(TestCase):
    def test_returns_distinct_stores_across_all_batches(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'S1', 'Silver', 'Old Store Name')
        _add_snapshot_member(b2, 'S2', 'Gold', 'New Store Name')
        stores = get_snapshot_registration_stores()
        self.assertIn('Old Store Name', stores)
        self.assertIn('New Store Name', stores)

    def test_blank_store_excluded(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(batch, 'S3', 'Gold', '')
        stores = get_snapshot_registration_stores()
        self.assertNotIn('', stores)
        self.assertNotIn('(No Store)', stores)

    def test_differs_from_live_customer_store_list(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        _add_snapshot_member(batch, 'S4', 'Silver', 'Renamed-Since Store')
        _customer('LIVE1', 'PL1', store='Current Live Store Only')

        stores = get_snapshot_registration_stores()
        self.assertIn('Renamed-Since Store', stores)
        self.assertNotIn('Current Live Store Only', stores)  # this store has zero snapshot rows


class MembershipDashboardSnapshotStoresContextTest(_ClearDropdownCacheMixin, TestCase):
    def test_context_has_separate_snapshot_stores_list(self):
        batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        _add_snapshot_member(batch, 'S5', 'Gold', 'Snapshot-Only Store')
        _customer('LIVE2', 'PL2', store='Live-Only Store')

        user = User.objects.create_superuser("membctx1", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", follow=True, SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)
        self.assertIn('Snapshot-Only Store', r.context['snapshot_stores'])
        self.assertNotIn('Live-Only Store', r.context['snapshot_stores'])
        self.assertIn('Live-Only Store', r.context['registration_stores'])
        self.assertNotIn('Snapshot-Only Store', r.context['registration_stores'])


# ---------------------------------------------------------------------------
# 14. get_grade_breakdown_by_store_comparison() + membership_store_breakdown_partial
#     — From/To matrix (added 2026-09-01). Replaced the old single-store
#     drill-down mode (removed — redundant now this matrix shows from/to
#     directly, and get_grade_changes() shows the actual movers).
# ---------------------------------------------------------------------------

class GetGradeBreakdownByStoreComparisonTest(TestCase):
    def test_from_to_pairs_per_store(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'C1', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'C1', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'C2', 'Silver', 'Shop A')

        rows = {r['store']: r for r in get_grade_breakdown_by_store_comparison(b1.id, b2.id)}
        counts = {c['grade']: c for c in rows['Shop A']['counts']}
        self.assertEqual(counts['Silver']['from'], 1)
        self.assertEqual(counts['Silver']['to'], 2)
        self.assertEqual(rows['Shop A']['total_from'], 1)
        self.assertEqual(rows['Shop A']['total_to'], 2)

    def test_store_present_in_only_one_batch_still_appears(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b2, 'C3', 'Gold', 'New Store')  # only exists in b2

        rows = {r['store']: r for r in get_grade_breakdown_by_store_comparison(b1.id, b2.id)}
        self.assertIn('New Store', rows)
        counts = {c['grade']: c for c in rows['New Store']['counts']}
        self.assertEqual(counts['Gold']['from'], 0)
        self.assertEqual(counts['Gold']['to'], 1)

    def test_no_from_batch_treats_from_as_zero(self):
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b2, 'C4', 'Gold', 'Shop X')
        rows = {r['store']: r for r in get_grade_breakdown_by_store_comparison(None, b2.id)}
        self.assertEqual(rows['Shop X']['total_from'], 0)
        self.assertEqual(rows['Shop X']['total_to'], 1)

    def test_no_grade_only_store_excluded_on_both_sides(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'C5', 'No Grade', 'Shop NG')
        _add_snapshot_member(b2, 'C5', 'No Grade', 'Shop NG')
        rows = {r['store']: r for r in get_grade_breakdown_by_store_comparison(b1.id, b2.id)}
        self.assertNotIn('Shop NG', rows)  # total_from == total_to == 0 -> excluded entirely

    def test_all_stores_total_row_appended_last(self):
        # Added 2026-09-02, PO feedback — the last row is the 'All Stores'
        # total, computed from get_grade_breakdown() (the 'overall' bucket),
        # NOT a sum of the per-store rows above it.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'AS1', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'AS1', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'AS2', 'Gold', 'Shop A')
        _add_snapshot_member(b1, 'AS3', 'Silver', 'Shop B')
        _add_snapshot_member(b2, 'AS3', 'Gold', 'Shop B')

        rows = get_grade_breakdown_by_store_comparison(b1.id, b2.id)
        for row in rows[:-1]:
            self.assertFalse(row['is_total'])
        total_row = rows[-1]
        self.assertEqual(total_row['store'], 'All Stores')
        self.assertTrue(total_row['is_total'])

        expected_from = get_grade_breakdown(b1.id)
        expected_to = get_grade_breakdown(b2.id)
        counts_by_grade = {c['grade']: c for c in total_row['counts']}
        for g in DISPLAY_GRADES:
            self.assertEqual(counts_by_grade[g]['from'], expected_from[g])
            self.assertEqual(counts_by_grade[g]['to'], expected_to[g])
        self.assertEqual(total_row['total_from'], sum(expected_from.values()))
        self.assertEqual(total_row['total_to'], sum(expected_to.values()))
        # Cross-check against compare_batches() (no store=) — the same
        # authoritative "all stores combined" source.
        comparison = {c['grade']: c for c in compare_batches(b1.id, b2.id)}
        for g in DISPLAY_GRADES:
            self.assertEqual(counts_by_grade[g]['from'], comparison[g]['from_count'])
            self.assertEqual(counts_by_grade[g]['to'], comparison[g]['to_count'])


class MembershipStoreBreakdownPartialWebTest(TestCase):
    """membership_store_breakdown_partial — always renders the From/To
    matrix now (PO feedback 2026-09-01)."""

    def test_defaults_to_latest_batch_when_no_batch_param(self):
        MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        newest = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(newest, 'SB1', 'Gold', 'Shop X')

        user = User.objects.create_superuser("membadmin5", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/partial/store-breakdown/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"Shop X", r.content)

    def test_shows_from_and_to_columns_with_correct_counts(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'C1', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'C1', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'C2', 'Silver', 'Shop A')

        user = User.objects.create_superuser("storecmp1", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/store-breakdown/?batch={b2.id}&from_batch={b1.id}",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"Shop A", r.content)
        self.assertIn(b"From", r.content)
        self.assertIn(b"To", r.content)
        self.assertIn(b"All Stores", r.content)

    def test_blocked_for_anonymous(self):
        client = Client()
        r = client.get("/membership/partial/store-breakdown/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 401)

    def test_no_batches_shows_warning(self):
        user = User.objects.create_superuser("membadmin6", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get("/membership/partial/store-breakdown/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"No snapshot", r.content)


# ---------------------------------------------------------------------------
# 15. get_grade_changes() — grade-change diff feature (added 2026-09-01, PO
#     feedback: this is the whole reason grade_members stores vip_id lists).
# ---------------------------------------------------------------------------

class GetGradeChangesTest(TestCase):
    def test_upgrade_detected(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'M1', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'M1', 'Silver', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 1)
        self.assertEqual(rows[0]['vip_id'], 'M1')
        self.assertEqual(rows[0]['from_grade'], 'Member')
        self.assertEqual(rows[0]['to_grade'], 'Silver')
        self.assertEqual(rows[0]['direction'], 'upgrade')

    def test_downgrade_detected(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'M2', 'Diamond', 'Shop A')
        _add_snapshot_member(b2, 'M2', 'Gold', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(rows[0]['direction'], 'downgrade')

    def test_unchanged_grade_excluded(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'M3', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'M3', 'Silver', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 0)

    def test_customer_present_in_only_one_batch_excluded(self):
        # A "new" or "removed" customer is a different event, not a grade
        # change — must be in BOTH snapshots to count.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'ONLY_IN_FROM', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'ONLY_IN_TO', 'Gold', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 0)

    def test_vip_id_zero_excluded(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, '0', 'Member', 'Shop A')
        _add_snapshot_member(b2, '0', 'Gold', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 0)

    def test_no_grade_transition_excluded(self):
        # Matches the DISPLAY_GRADES convention used by every other
        # grade-level view on this page — 'No Grade' is noise.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'M4', 'No Grade', 'Shop A')
        _add_snapshot_member(b2, 'M4', 'Member', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 0)

    def test_store_filter_matches_same_store_both_sides(self):
        # Still true under the new OR-semantics (store=='X' if from_store==X
        # OR to_store==X): a customer whose store is unchanged is matched by
        # that one store's filter, and NOT by an unrelated store's filter.
        # See test_store_filter_or_semantics_matches_either_side (below) for
        # the actual behavior-change case (store differs between snapshots).
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'M5', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'M5', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'M6', 'Member', 'Shop B')
        _add_snapshot_member(b2, 'M6', 'Silver', 'Shop B')

        rows, total = get_grade_changes(b1.id, b2.id, store='Shop A')
        self.assertEqual(total, 1)
        self.assertEqual(rows[0]['vip_id'], 'M5')

    def test_grade_filter_matches_new_grade(self):
        # `grade` filters on the NEW (to) grade — a single filter covering
        # both directions, e.g. grade='Silver' shows Member->Silver upgrades
        # AND Gold->Silver downgrades.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'UP', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'UP', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'DOWN', 'Gold', 'Shop A')
        _add_snapshot_member(b2, 'DOWN', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'OTHER', 'Silver', 'Shop A')
        _add_snapshot_member(b2, 'OTHER', 'Gold', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id, grade='Silver')
        self.assertEqual(total, 2)
        vip_ids = {r['vip_id'] for r in rows}
        self.assertEqual(vip_ids, {'UP', 'DOWN'})

    def test_direction_filter(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'UP2', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'UP2', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'DOWN2', 'Gold', 'Shop A')
        _add_snapshot_member(b2, 'DOWN2', 'Member', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id, direction='upgrade')
        self.assertEqual(total, 1)
        self.assertEqual(rows[0]['vip_id'], 'UP2')

    def test_total_count_reflects_full_set_not_page(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        for i in range(5):
            _add_snapshot_member(b1, f'P{i}', 'Member', 'Shop A')
            _add_snapshot_member(b2, f'P{i}', 'Silver', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id, limit=2)
        self.assertEqual(len(rows), 2)
        self.assertEqual(total, 5)

    def test_name_phone_joined_from_live_customer(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _customer('M7', 'P123', grade='Silver', store='Shop A')
        _add_snapshot_member(b1, 'M7', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'M7', 'Silver', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(rows[0]['name'], 'Cust M7')
        self.assertEqual(rows[0]['phone'], 'P123')

    def test_missing_live_customer_yields_none_name_phone_not_error(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        # No matching Customer row created for 'GHOST' — deleted since the snapshot.
        _add_snapshot_member(b1, 'GHOST', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'GHOST', 'Silver', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 1)
        self.assertIsNone(rows[0]['name'])
        self.assertIsNone(rows[0]['phone'])

    def test_from_store_to_store_fields_same_store(self):
        # Store unchanged between snapshots -> from_store == to_store.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'SS1', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'SS1', 'Silver', 'Shop A')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 1)
        self.assertEqual(rows[0]['from_store'], 'Shop A')
        self.assertEqual(rows[0]['to_store'], 'Shop A')

    def test_from_store_to_store_fields_different_store(self):
        # Store-name-drift case (e.g. a customer re-import that reformatted a
        # store's name between the two snapshots) -> from_store != to_store,
        # both populated correctly rather than the change being invisible.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'DS1', 'Member', 'Old Store Name')
        _add_snapshot_member(b2, 'DS1', 'Silver', 'New Store Name')

        rows, total = get_grade_changes(b1.id, b2.id)
        self.assertEqual(total, 1)
        self.assertEqual(rows[0]['from_store'], 'Old Store Name')
        self.assertEqual(rows[0]['to_store'], 'New Store Name')

    def test_store_filter_or_semantics_matches_either_side(self):
        # PO feedback: a customer who moved FROM store X TO store Y (X != Y)
        # must now be findable via EITHER store's filter — this is the
        # regression-fixing behavior change from the old AND/scoped-both-
        # sides semantics, which made these customers unfindable by any
        # store filter.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'OR1', 'Member', 'X')
        _add_snapshot_member(b2, 'OR1', 'Silver', 'Y')

        rows_x, total_x = get_grade_changes(b1.id, b2.id, store='X')
        self.assertEqual(total_x, 1)
        self.assertEqual(rows_x[0]['vip_id'], 'OR1')

        rows_y, total_y = get_grade_changes(b1.id, b2.id, store='Y')
        self.assertEqual(total_y, 1)
        self.assertEqual(rows_y[0]['vip_id'], 'OR1')

        rows_z, total_z = get_grade_changes(b1.id, b2.id, store='Z')
        self.assertEqual(total_z, 0)


# ---------------------------------------------------------------------------
# 15B. get_grade_changes_overview_by_store() — aggregate Downgrade/Upgrade
#      overview table, added 2026-09-02 (PO feedback). Sits above the
#      individual-customer list get_grade_changes() powers.
# ---------------------------------------------------------------------------

class GetGradeChangesOverviewByStoreTest(TestCase):
    def test_missing_batch_returns_empty_list(self):
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        self.assertEqual(get_grade_changes_overview_by_store(None, b2.id), [])
        self.assertEqual(get_grade_changes_overview_by_store(b2.id, None), [])
        self.assertEqual(get_grade_changes_overview_by_store(None, None), [])

    def test_per_store_downgrade_upgrade_counts(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        # Shop A: two upgrades to Silver, one downgrade to Silver.
        _add_snapshot_member(b1, 'OV1', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'OV1', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'OV2', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'OV2', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'OV3', 'Gold', 'Shop A')
        _add_snapshot_member(b2, 'OV3', 'Silver', 'Shop A')
        # Shop B: one upgrade to Gold.
        _add_snapshot_member(b1, 'OV4', 'Silver', 'Shop B')
        _add_snapshot_member(b2, 'OV4', 'Gold', 'Shop B')
        # Shop C: configured members, but NO grade change between batches —
        # must still appear in the list with all-zero counts.
        _add_snapshot_member(b1, 'OV5', 'Silver', 'Shop C')
        _add_snapshot_member(b2, 'OV5', 'Silver', 'Shop C')

        rows = get_grade_changes_overview_by_store(b1.id, b2.id)
        by_store = {r['store']: r for r in rows}

        self.assertIn('Shop A', by_store)
        counts_a = {c['grade']: c for c in by_store['Shop A']['counts']}
        self.assertEqual(counts_a['Silver']['upgrade'], 2)
        self.assertEqual(counts_a['Silver']['downgrade'], 1)
        self.assertEqual(by_store['Shop A']['total_upgrade'], 2)
        self.assertEqual(by_store['Shop A']['total_downgrade'], 1)
        self.assertFalse(by_store['Shop A']['is_total'])

        counts_b = {c['grade']: c for c in by_store['Shop B']['counts']}
        self.assertEqual(counts_b['Gold']['upgrade'], 1)
        self.assertEqual(counts_b['Gold']['downgrade'], 0)

        # Shop C: zero changes, must still appear with all-zero counts.
        self.assertIn('Shop C', by_store)
        self.assertEqual(by_store['Shop C']['total_upgrade'], 0)
        self.assertEqual(by_store['Shop C']['total_downgrade'], 0)
        for c in by_store['Shop C']['counts']:
            self.assertEqual(c['upgrade'], 0)
            self.assertEqual(c['downgrade'], 0)

    def test_all_stores_row_last_and_matches_independent_tally(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'AT1', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'AT1', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'AT2', 'Gold', 'Shop A')
        _add_snapshot_member(b2, 'AT2', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'AT3', 'Silver', 'Shop B')
        _add_snapshot_member(b2, 'AT3', 'Gold', 'Shop B')
        _add_snapshot_member(b1, 'AT4', 'Silver', 'Shop B')  # unchanged, not counted
        _add_snapshot_member(b2, 'AT4', 'Silver', 'Shop B')

        rows = get_grade_changes_overview_by_store(b1.id, b2.id)
        total_row = rows[-1]
        self.assertEqual(total_row['store'], 'All Stores')
        self.assertTrue(total_row['is_total'])
        for row in rows[:-1]:
            self.assertFalse(row['is_total'])

        # Independent verification path: tally get_grade_changes() directly
        # (limit=None — full set) by to_grade/direction, rather than summing
        # the per-store rows already computed above.
        changes, total = get_grade_changes(b1.id, b2.id, limit=None)
        self.assertEqual(total, 3)
        expected_up = {g: 0 for g in DISPLAY_GRADES}
        expected_down = {g: 0 for g in DISPLAY_GRADES}
        for c in changes:
            bucket = expected_up if c['direction'] == 'upgrade' else expected_down
            bucket[c['to_grade']] += 1

        counts_by_grade = {c['grade']: c for c in total_row['counts']}
        for g in DISPLAY_GRADES:
            self.assertEqual(counts_by_grade[g]['upgrade'], expected_up[g])
            self.assertEqual(counts_by_grade[g]['downgrade'], expected_down[g])
        self.assertEqual(total_row['total_upgrade'], sum(expected_up.values()))
        self.assertEqual(total_row['total_downgrade'], sum(expected_down.values()))

    def test_db_query_count_is_exactly_two(self):
        # Critical performance constraint — must be one query per batch
        # (via _grade_members_json()), NOT one query per store.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        for i in range(10):
            store = f"Shop {i}"
            _add_snapshot_member(b1, f'Q{i}', 'Member', store)
            _add_snapshot_member(b2, f'Q{i}', 'Silver', store)

        with self.assertNumQueries(2):
            get_grade_changes_overview_by_store(b1.id, b2.id)


# ---------------------------------------------------------------------------
# 15D. get_grade_changes_store_transitions() — itemized (from_store, to_store)
#      appendix for changes get_grade_changes_overview_by_store() cannot
#      attribute to a single store (store-name-drift cases). Added 2026-09.
# ---------------------------------------------------------------------------

def _build_transitions_scenario():
    """Shared fixture for grouping/reconciliation tests below.

    Same-store changes (attributable by the main overview table):
      SAME1: Shop A -> Shop A, Member -> Silver (upgrade)
      SAME2: Shop A -> Shop A, Gold -> Silver (downgrade)
      SAME3: Shop F -> Shop F, Silver -> Gold (upgrade)
      -> 3 changes total, all invisible to get_grade_changes_store_transitions().

    Different-store changes (the "invisible" remainder this function surfaces):
      Shop A -> Shop B (3 members, the largest group — must sort first):
        T1: Member -> Silver (upgrade)
        T2: Diamond -> Gold (downgrade)
        T3: Silver -> Gold (upgrade)
      Shop C -> Shop D (1 member):
        T4: Member -> Silver (upgrade)
      (No Store) -> Shop E (1 member, blank store on the from-side):
        T5: Gold -> Diamond (upgrade)
      -> 5 changes total.

    Grand total across both groups: 8 changes.
    """
    b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
    b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')

    _add_snapshot_member(b1, 'SAME1', 'Member', 'Shop A')
    _add_snapshot_member(b2, 'SAME1', 'Silver', 'Shop A')
    _add_snapshot_member(b1, 'SAME2', 'Gold', 'Shop A')
    _add_snapshot_member(b2, 'SAME2', 'Silver', 'Shop A')
    _add_snapshot_member(b1, 'SAME3', 'Silver', 'Shop F')
    _add_snapshot_member(b2, 'SAME3', 'Gold', 'Shop F')

    _add_snapshot_member(b1, 'T1', 'Member', 'Shop A')
    _add_snapshot_member(b2, 'T1', 'Silver', 'Shop B')
    _add_snapshot_member(b1, 'T2', 'Diamond', 'Shop A')
    _add_snapshot_member(b2, 'T2', 'Gold', 'Shop B')
    _add_snapshot_member(b1, 'T3', 'Silver', 'Shop A')
    _add_snapshot_member(b2, 'T3', 'Gold', 'Shop B')

    _add_snapshot_member(b1, 'T4', 'Member', 'Shop C')
    _add_snapshot_member(b2, 'T4', 'Silver', 'Shop D')

    _add_snapshot_member(b1, 'T5', 'Gold', '')  # blank -> '(No Store)'
    _add_snapshot_member(b2, 'T5', 'Diamond', 'Shop E')

    return b1, b2


class GetGradeChangesStoreTransitionsTest(TestCase):
    def test_missing_batch_returns_empty_list(self):
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        self.assertEqual(get_grade_changes_store_transitions(None, b2.id), [])
        self.assertEqual(get_grade_changes_store_transitions(b2.id, None), [])
        self.assertEqual(get_grade_changes_store_transitions(None, None), [])

    def test_grouping_counts_sort_order_and_same_store_exclusion(self):
        b1, b2 = _build_transitions_scenario()

        transitions = get_grade_changes_store_transitions(b1.id, b2.id)
        pairs = {(t['from_store'], t['to_store']): t for t in transitions}

        # Same-store customers must NOT appear here — they belong in the
        # main overview table, not this appendix.
        self.assertNotIn(('Shop A', 'Shop A'), pairs)
        self.assertNotIn(('Shop F', 'Shop F'), pairs)
        self.assertEqual(len(transitions), 3)

        # Grouping + per-grade counts.
        self.assertIn(('Shop A', 'Shop B'), pairs)
        ab = pairs[('Shop A', 'Shop B')]
        counts_ab = {c['grade']: c for c in ab['counts']}
        self.assertEqual(counts_ab['Silver']['upgrade'], 1)   # T1
        self.assertEqual(counts_ab['Gold']['downgrade'], 1)   # T2
        self.assertEqual(counts_ab['Gold']['upgrade'], 1)     # T3
        self.assertEqual(ab['total_upgrade'], 2)
        self.assertEqual(ab['total_downgrade'], 1)

        self.assertIn(('Shop C', 'Shop D'), pairs)
        cd = pairs[('Shop C', 'Shop D')]
        self.assertEqual(cd['total_upgrade'], 1)
        self.assertEqual(cd['total_downgrade'], 0)

        self.assertIn(('(No Store)', 'Shop E'), pairs)
        ne = pairs[('(No Store)', 'Shop E')]
        self.assertEqual(ne['total_upgrade'], 1)
        self.assertEqual(ne['total_downgrade'], 0)

        # Sort order: highest total (Shop A -> Shop B, total=3) must be first.
        self.assertEqual((transitions[0]['from_store'], transitions[0]['to_store']), ('Shop A', 'Shop B'))
        totals = [t['total_downgrade'] + t['total_upgrade'] for t in transitions]
        self.assertEqual(totals, sorted(totals, reverse=True))

    def test_db_query_count_is_exactly_two(self):
        b1, b2 = _build_transitions_scenario()
        with self.assertNumQueries(2):
            get_grade_changes_store_transitions(b1.id, b2.id)

    def test_reconciliation_with_overview_by_store(self):
        # Core correctness property of this whole feature: this function's
        # rows are a strict partition of "the changes NOT captured by
        # get_grade_changes_overview_by_store()'s per-store rows" — so
        # summing both, non-total-row totals from the overview plus every
        # row's totals from this function, must exactly equal the overview's
        # 'All Stores' row.
        b1, b2 = _build_transitions_scenario()

        transitions = get_grade_changes_store_transitions(b1.id, b2.id)
        overview = get_grade_changes_overview_by_store(b1.id, b2.id)

        transitions_total = sum(t['total_downgrade'] + t['total_upgrade'] for t in transitions)
        overview_non_total = sum(
            r['total_downgrade'] + r['total_upgrade'] for r in overview if not r['is_total']
        )
        all_stores_row = next(r for r in overview if r['is_total'])
        all_stores_total = all_stores_row['total_downgrade'] + all_stores_row['total_upgrade']

        self.assertEqual(all_stores_total, 8)  # 3 same-store + 5 different-store
        self.assertEqual(transitions_total, 5)
        self.assertEqual(overview_non_total, 3)
        self.assertEqual(transitions_total + overview_non_total, all_stores_total)


class MembershipMoversPartialWebTest(TestCase):
    def test_shows_changed_members(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'MV1', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'MV1', 'Silver', 'Shop A')

        user = User.objects.create_superuser("movers1", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers/?from_batch={b1.id}&to_batch={b2.id}",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"MV1", r.content)
        self.assertIn(b"Upgraded", r.content)

    def test_requires_both_batches(self):
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        user = User.objects.create_superuser("movers2", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers/?to_batch={b2.id}",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"Select both", r.content)

    def test_store_and_grade_filters(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'MV2', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'MV2', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'MV3', 'Member', 'Shop B')
        _add_snapshot_member(b2, 'MV3', 'Gold', 'Shop B')

        user = User.objects.create_superuser("movers3", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers/?from_batch={b1.id}&to_batch={b2.id}&store=Shop+A",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"MV2", r.content)
        self.assertNotIn(b"MV3", r.content)

    def test_blocked_for_anonymous(self):
        client = Client()
        r = client.get("/membership/partial/movers/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 401)

    def test_direction_filter(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'MV4', 'Member', 'Shop A')   # upgrade
        _add_snapshot_member(b2, 'MV4', 'Silver', 'Shop A')
        _add_snapshot_member(b1, 'MV5', 'Gold', 'Shop A')     # downgrade
        _add_snapshot_member(b2, 'MV5', 'Silver', 'Shop A')

        user = User.objects.create_superuser("movers4", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers/?from_batch={b1.id}&to_batch={b2.id}&direction=downgrade",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"MV5", r.content)
        self.assertNotIn(b"MV4", r.content)

    def test_store_filter_matches_either_from_or_to_store(self):
        # OR-semantics regression test at the HTTP/view layer (added
        # 2026-09-02 review follow-up) — a customer whose store changed
        # between snapshots must now be findable by filtering on EITHER
        # their old or their new store, not excluded from both like the
        # pre-2026-09-02 AND-semantics did.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'MV6', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'MV6', 'Silver', 'Shop B')  # store changed A -> B

        user = User.objects.create_superuser("movers6", password="pw")
        client = Client()
        client.force_login(user)
        r_a = client.get(
            f"/membership/partial/movers/?from_batch={b1.id}&to_batch={b2.id}&store=Shop+A",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertIn(b"MV6", r_a.content)
        r_b = client.get(
            f"/membership/partial/movers/?from_batch={b1.id}&to_batch={b2.id}&store=Shop+B",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertIn(b"MV6", r_b.content)
        r_c = client.get(
            f"/membership/partial/movers/?from_batch={b1.id}&to_batch={b2.id}&store=Shop+C",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertNotIn(b"MV6", r_c.content)


# ---------------------------------------------------------------------------
# 15C. membership_movers_overview_partial web tests (added 2026-09-02)
# ---------------------------------------------------------------------------

class MembershipMoversOverviewPartialWebTest(TestCase):
    def test_shows_downgrade_upgrade_counts(self):
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'MO1', 'Member', 'Shop A')
        _add_snapshot_member(b2, 'MO1', 'Silver', 'Shop A')

        user = User.objects.create_superuser("moversov1", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers-overview/?from_batch={b1.id}&to_batch={b2.id}",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"Shop A", r.content)
        self.assertIn(b"All Stores", r.content)

    def test_drifted_store_shows_as_single_arrow_row_in_the_one_table(self):
        # Merged into ONE table (2026-09-02, PO feedback: "đừng làm nó
        # complicated") — a customer whose store differs between From/To no
        # longer gets a separate appendix table; they appear as one "A -> B"
        # row inside the same table as the real per-store rows.
        b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(b1, 'MO2', 'Member', 'Old Store Name')
        _add_snapshot_member(b2, 'MO2', 'Silver', 'New Store Name')

        user = User.objects.create_superuser("moversov3", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers-overview/?from_batch={b1.id}&to_batch={b2.id}",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn("Old Store Name → New Store Name".encode(), r.content)
        # No separate appendix section/heading anymore -- one table only.
        self.assertNotIn(b"Store Name Changes", r.content)

    def test_requires_both_batches(self):
        b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        user = User.objects.create_superuser("moversov2", password="pw")
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/partial/movers-overview/?to_batch={b2.id}",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        self.assertIn(b"Select both", r.content)

    def test_blocked_for_anonymous(self):
        client = Client()
        r = client.get("/membership/partial/movers-overview/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 401)


# ---------------------------------------------------------------------------
# 16. membership_trend_partial — JSON endpoint for the chart's store filter
# ---------------------------------------------------------------------------

class MembershipTrendPartialWebTest(TestCase):
    def setUp(self):
        self.batch = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(self.batch, 'T1', 'Gold', 'Shop A')
        _add_snapshot_member(self.batch, 'T2', 'Gold', 'Shop B')
        self.user = User.objects.create_superuser("trend1", password="pw")
        self.client_ = Client()
        self.client_.force_login(self.user)

    def test_no_store_returns_full_json(self):
        r = self.client_.get("/membership/partial/trend/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        data = json.loads(r.content)
        series = {s['batch_id']: s['counts'] for s in data['series']}
        self.assertEqual(series[self.batch.id]['Gold'], 2)

    def test_store_filter_scopes_json(self):
        r = self.client_.get(
            "/membership/partial/trend/?store=Shop+A",
            SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 200)
        data = json.loads(r.content)
        series = {s['batch_id']: s['counts'] for s in data['series']}
        self.assertEqual(series[self.batch.id]['Gold'], 1)

    def test_blocked_for_anonymous(self):
        client = Client()
        r = client.get("/membership/partial/trend/", SERVER_NAME="localhost", HTTP_X_REQUESTED_WITH="XMLHttpRequest")
        self.assertEqual(r.status_code, 401)


# ---------------------------------------------------------------------------
# 17. export_membership_excel — per-section Excel download buttons
#     (added 2026-09-01: comparison | store | movers | trend | tier)
# ---------------------------------------------------------------------------

class ExportMembershipExcelTest(_ClearDropdownCacheMixin, TestCase):
    def setUp(self):
        super().setUp()
        self.b1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 1, 1), source='auto')
        self.b2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 6, 1), source='auto')
        _add_snapshot_member(self.b1, 'EXP1', 'Member', 'Shop A')
        _add_snapshot_member(self.b2, 'EXP1', 'Silver', 'Shop A')
        self.user = User.objects.create_superuser("exportadmin", password="pw")
        self.client_ = Client()
        self.client_.force_login(self.user)

    def _get(self, params):
        return self.client_.get(f"/membership/export/?{params}", SERVER_NAME="localhost")

    def test_comparison_export_returns_xlsx(self):
        r = self._get(f"section=comparison&from_batch={self.b1.id}&to_batch={self.b2.id}")
        self.assertEqual(r.status_code, 200)
        self.assertEqual(
            r["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(r.content)
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        self.assertEqual(ws.title, "Grade Comparison")
        header = [c.value for c in ws[3]]
        self.assertEqual(header, ["Grade", "From", "To", "Diff", "% Change"])
        rows = {row[0]: row for row in ws.iter_rows(min_row=4, values_only=True)}
        self.assertIn('Silver', rows)
        self.assertEqual(rows['Silver'][1], 0)   # from_count
        self.assertEqual(rows['Silver'][2], 1)   # to_count

    def test_comparison_export_missing_batch_redirects(self):
        r = self._get("section=comparison")
        self.assertEqual(r.status_code, 302)

    def test_store_export_returns_xlsx(self):
        r = self._get(f"section=store&from_batch={self.b1.id}&to_batch={self.b2.id}")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.active.title, "By Registration Store")
        self.assertGreater(wb.active.max_row, 1)
        header = [c.value for c in wb.active[1]]
        self.assertEqual(header[0], "Store")
        self.assertIn("Silver From", header)
        self.assertIn("Silver To", header)

    def test_movers_export_returns_xlsx(self):
        r = self._get(f"section=movers&from_batch={self.b1.id}&to_batch={self.b2.id}")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.active.title, "Grade Changes")
        header = [c.value for c in wb.active[1]]
        self.assertEqual(
            header,
            ["VIP ID", "Name", "Phone", "Store", "From Store", "To Store",
             "From Grade", "To Grade", "Direction"],
        )
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 'EXP1')
        self.assertEqual(rows[0][4], 'Shop A')  # from_store
        self.assertEqual(rows[0][5], 'Shop A')  # to_store
        self.assertEqual(rows[0][6], 'Member')  # from_grade
        self.assertEqual(rows[0][7], 'Silver')  # to_grade

    def test_movers_export_full_dataset_not_capped_at_ui_limit(self):
        # UI partial view/table is limited for display; the export must
        # contain the FULL filtered set. Use >20 changed members to prove it.
        b3 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 2, 1), source='auto')
        b4 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 7, 1), source='auto')
        for i in range(25):
            vid = f"MOV{i}"
            _add_snapshot_member(b3, vid, 'Member', 'Shop A')
            _add_snapshot_member(b4, vid, 'Silver', 'Shop A')
        r = self._get(f"section=movers&from_batch={b3.id}&to_batch={b4.id}")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 25)

    def test_movers_export_requires_both_batches(self):
        r = self._get(f"section=movers&to_batch={self.b2.id}")
        self.assertEqual(r.status_code, 302)

    def test_trend_export_returns_xlsx(self):
        r = self._get("section=trend")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.active.title, "Grade Trend")
        header = [c.value for c in wb.active[1]]
        self.assertEqual(header[0], "Snapshot Date")
        self.assertIn("Silver", header)

    def test_tier_export_returns_xlsx(self):
        _customer('TIER1', 'PT1', grade='Gold')
        r = self._get("section=tier")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.active.title, "Customer Tier Progress")
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertTrue(any(row[0] == 'TIER1' for row in rows))

    def test_tier_export_full_dataset_not_capped_at_ui_limit(self):
        # UI partial caps get_live_customer_tier_table() at limit=500; the
        # export must pass limit=None and return every row.
        for i in range(505):
            _customer(f"TCUST{i}", f"PTC{i}", grade='Member')
        r = self._get("section=tier")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertGreater(len(rows), 500)

    def test_invalid_section_redirects(self):
        r = self._get("section=bogus")
        self.assertEqual(r.status_code, 302)

    def test_export_blocked_without_permission(self):
        user = User.objects.create_user("noexport", password="pw")  # no role/perm assigned
        client = Client()
        client.force_login(user)
        r = client.get(
            f"/membership/export/?section=comparison&from_batch={self.b1.id}&to_batch={self.b2.id}",
            SERVER_NAME="localhost",
        )
        self.assertEqual(r.status_code, 302)  # requires_perm redirects to "home"

    def test_export_blocked_for_anonymous(self):
        client = Client()
        r = client.get(
            f"/membership/export/?section=comparison&from_batch={self.b1.id}&to_batch={self.b2.id}",
            SERVER_NAME="localhost",
        )
        self.assertEqual(r.status_code, 302)  # @login_required redirects to login

    def test_export_blocked_with_view_but_not_export_permission(self):
        # Distinct from test_export_blocked_without_permission (no role at
        # all) — this user CAN see the page (membership.view) but must still
        # be blocked from the export endpoint without membership.export.
        role = Role.objects.create(name='membership_view_no_export', permissions=['membership.view'], is_system=False)
        user = User.objects.create_user("viewnoexport", password="pw")
        UserProfile.objects.create(user=user, role=role)
        client = Client()
        client.force_login(user)
        r = client.get("/membership/", SERVER_NAME="localhost")
        self.assertEqual(r.status_code, 200)  # can view the page
        r = client.get(
            f"/membership/export/?section=comparison&from_batch={self.b1.id}&to_batch={self.b2.id}",
            SERVER_NAME="localhost",
        )
        self.assertEqual(r.status_code, 302)  # but not export

    def test_store_export_with_no_per_store_data_still_shows_all_stores_total(self):
        # Two batches with zero snapshot members each -> no per-store rows,
        # but get_grade_breakdown_by_store_comparison() now always appends an
        # 'All Stores' total row (added for the by-Store overview feature) ->
        # the export gracefully shows that all-zero total instead of
        # redirecting, consistent with how section=comparison already
        # behaves for a batch with no data.
        b_empty1 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 9, 1), source='auto')
        b_empty2 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 10, 1), source='auto')
        r = self._get(f"section=store&from_batch={b_empty1.id}&to_batch={b_empty2.id}")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 'All Stores')

    def test_movers_export_with_store_grade_direction_filters(self):
        b3 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 3, 1), source='auto')
        b4 = MembershipSnapshotBatch.objects.create(snapshot_date=date(2026, 8, 1), source='auto')
        _add_snapshot_member(b3, 'FMOV1', 'Member', 'Shop A')
        _add_snapshot_member(b4, 'FMOV1', 'Silver', 'Shop A')   # upgrade, Shop A
        _add_snapshot_member(b3, 'FMOV2', 'Gold', 'Shop B')
        _add_snapshot_member(b4, 'FMOV2', 'Silver', 'Shop B')   # downgrade, Shop B
        r = self._get(
            f"section=movers&from_batch={b3.id}&to_batch={b4.id}"
            "&store=Shop+A&grade=Silver&direction=upgrade"
        )
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0][0], 'FMOV1')

    def test_movers_export_empty_filtered_result_redirects(self):
        r = self._get(
            f"section=movers&from_batch={self.b1.id}&to_batch={self.b2.id}&direction=downgrade"
        )  # the only real change in setUp (EXP1) is an upgrade
        self.assertEqual(r.status_code, 302)

    def test_trend_export_with_store_filter(self):
        r = self._get("section=trend&store=Shop+A")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        self.assertEqual(wb.active.title, "Grade Trend")

    def test_trend_export_with_nonmatching_store_filter_returns_zeroed_series(self):
        # get_all_batch_grade_series() always returns one row per EXISTING
        # batch (never an empty list while batches exist) -- a store filter
        # that matches nothing just zeroes every grade's count rather than
        # redirecting. Confirms the filter is applied, not ignored.
        r = self._get("section=trend&store=NoSuchStoreAnywhere")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(rows), 2)  # b1 and b2 from setUp
        for row in rows:
            self.assertEqual(sum(row[2:]), 0)  # all grade counts zeroed

    def test_tier_export_with_grade_and_shop_filters(self):
        _customer('TIERF1', 'PTF1', grade='Gold', store='Shop X')
        _customer('TIERF2', 'PTF2', grade='Silver', store='Shop Y')
        r = self._get("section=tier&grade=Gold&shop=Shop+X")
        self.assertEqual(r.status_code, 200)
        wb = load_workbook(io.BytesIO(r.content))
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        self.assertTrue(any(row[0] == 'TIERF1' for row in rows))
        self.assertFalse(any(row[0] == 'TIERF2' for row in rows))

    def test_tier_export_empty_filtered_result_redirects(self):
        r = self._get("section=tier&grade=Diamond&shop=NoSuchShopAnywhere")
        self.assertEqual(r.status_code, 302)


# ---------------------------------------------------------------------------
# 17. normalize_membership_stores management command — one-time retroactive
#     fix for EXISTING MembershipSnapshotBatch rows' `by_store` attribution,
#     mirroring the live-Customer-store rule create_backfill_snapshot() now
#     applies to future manual-import backfills via _resolve_live_stores()
#     (see App/management/commands/normalize_membership_stores.py docstring).
# ---------------------------------------------------------------------------

class NormalizeMembershipStoresCommandTest(TestCase):
    def setUp(self):
        from django.core.cache import cache
        cache.clear()

    def _run(self, **kwargs):
        out = io.StringIO()
        call_command('normalize_membership_stores', stdout=out, **kwargs)
        return out.getvalue()

    def test_dry_run_writes_nothing(self):
        # Live store differs from what the manual-import batch has recorded.
        _customer('N1', 'PN1', store='New Store Name')
        batch = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='manual_import')
        _add_snapshot_member(batch, 'N1', 'Silver', 'Old Store Name')

        before_counts = json.loads(json.dumps(batch.grade_counts))
        before_members = json.loads(json.dumps(batch.grade_members))

        self._run()  # no apply=True -> dry-run

        batch.refresh_from_db()
        self.assertEqual(batch.grade_counts, before_counts)
        self.assertEqual(batch.grade_members, before_members)

    def test_apply_updates_store_attribution_and_counts(self):
        _customer('N2', 'PN2', store='New Store Name')
        batch = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='manual_import')
        _add_snapshot_member(batch, 'N2', 'Silver', 'Old Store Name')

        self._run(apply=True)

        batch.refresh_from_db()
        self.assertNotIn('Old Store Name', batch.grade_counts['by_store'])
        self.assertEqual(batch.grade_counts['by_store']['New Store Name'].get('Silver', 0), 1)
        self.assertEqual(batch.grade_members['by_store']['New Store Name']['Silver'], ['N2'])
        # 'overall' is grade-level, not store-level — untouched by the move.
        self.assertEqual(batch.grade_counts['overall']['Silver'], 1)
        self.assertEqual(batch.grade_members['overall']['Silver'], ['N2'])

    def test_blank_live_store_coerced_to_no_store_placeholder(self):
        # Regression test (found via live-data reconciliation, 2026-09-02):
        # a customer whose LIVE registration_store is genuinely blank ('')
        # must be bucketed under the canonical '(No Store)' placeholder, not
        # a raw empty-string key -- matching _build_rows()'s own
        # `store_key = c['registration_store'] or '(No Store)'` convention.
        # A raw '' key silently diverged from '(No Store)' downstream (e.g.
        # get_grade_changes_store_transitions()'s `store or '(No Store)'`
        # comparison treated them as equal and incorrectly skipped a real
        # transition, breaking the overview-table reconciliation invariant).
        _customer('N4', 'PN4', store='')  # live store is blank
        batch = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='manual_import')
        _add_snapshot_member(batch, 'N4', 'Silver', 'Old Store Name')

        self._run(apply=True)

        batch.refresh_from_db()
        self.assertIn('(No Store)', batch.grade_counts['by_store'])
        self.assertNotIn('', batch.grade_counts['by_store'])
        self.assertEqual(batch.grade_counts['by_store']['(No Store)'].get('Silver', 0), 1)
        self.assertEqual(batch.grade_members['by_store']['(No Store)']['Silver'], ['N4'])

    def test_vip_id_not_in_live_customer_kept_under_original_store(self):
        batch = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='manual_import')
        _add_snapshot_member(batch, 'GHOST1', 'Gold', 'Ghost Old Store')
        # No live Customer row created for GHOST1 at all.

        self._run(apply=True)

        batch.refresh_from_db()
        self.assertEqual(batch.grade_counts['by_store']['Ghost Old Store'].get('Gold', 0), 1)
        self.assertEqual(batch.grade_members['by_store']['Ghost Old Store']['Gold'], ['GHOST1'])

    def test_default_scope_excludes_auto_batches_include_auto_widens_it(self):
        _customer('N3', 'PN3', store='Live Store')
        auto_batch = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='auto')
        _add_snapshot_member(auto_batch, 'N3', 'Silver', 'Stale Store')

        self._run(apply=True)  # default scope: manual_import only

        auto_batch.refresh_from_db()
        self.assertIn('Stale Store', auto_batch.grade_counts['by_store'])
        self.assertNotIn('Live Store', auto_batch.grade_counts['by_store'])

        self._run(apply=True, include_auto=True)

        auto_batch.refresh_from_db()
        self.assertNotIn('Stale Store', auto_batch.grade_counts['by_store'])
        self.assertIn('Live Store', auto_batch.grade_counts['by_store'])

    def test_batch_id_targets_only_that_batch(self):
        _customer('N4', 'PN4', store='Live Store A')
        _customer('N5', 'PN5', store='Live Store B')
        target = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='manual_import')
        _add_snapshot_member(target, 'N4', 'Silver', 'Old A')
        other = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 2, 1), source='manual_import')
        _add_snapshot_member(other, 'N5', 'Silver', 'Old B')

        self._run(apply=True, batch_id=target.id)

        target.refresh_from_db()
        other.refresh_from_db()
        self.assertIn('Live Store A', target.grade_counts['by_store'])
        self.assertIn('Old B', other.grade_counts['by_store'])
        self.assertNotIn('Live Store B', other.grade_counts['by_store'])

    def test_unknown_batch_id_raises_command_error(self):
        from django.core.management.base import CommandError
        with self.assertRaises(CommandError):
            self._run(batch_id=999999)

    def test_live_store_lookup_is_one_query_per_batch(self):
        # Expected 3 queries total for one batch, regardless of vip_id count:
        #   1. fetch the in-scope batch list (the command's scoping queryset)
        #   2. _resolve_live_stores() — exactly ONE bulk Customer query for
        #      the whole batch's vip_ids (never one per vip_id)
        #   3. batch.save(update_fields=[...]) — the UPDATE for this batch
        # cache.delete() is not a DB query.
        batch = MembershipSnapshotBatch.objects.create(
            snapshot_date=date(2026, 1, 1), source='manual_import')
        for i in range(50):
            vid = f'BULK{i}'
            _customer(vid, f'PBULK{i}', store=f'Live Store {i % 3}')
            _add_snapshot_member(batch, vid, 'Silver', f'Old Store {i % 5}')

        with self.assertNumQueries(3):
            call_command('normalize_membership_stores', apply=True, stdout=io.StringIO())
